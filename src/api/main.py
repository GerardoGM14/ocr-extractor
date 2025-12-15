"""
API Main - FastAPI application principal
Responsabilidad: Endpoints REST para procesamiento OCR
"""

import os
import sys
import time
import uuid
import tempfile
import logging
import json
import asyncio
import secrets
import hashlib
import string
from pathlib import Path
from typing import Optional, Dict, Any, AsyncGenerator, Tuple
from datetime import datetime, timedelta
from threading import Lock
from pydantic import BaseModel

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, status, Query, Header, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.exceptions import RequestValidationError
from starlette.requests import Request as StarletteRequest
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from .models import (
    ProcessPDFResponse,
    ProcessStatusResponse,
    ErrorResponse,
    HealthResponse,
    PageResult,
    UploadPDFResponse,
    DeleteUploadResponse,
    UploadedFilesResponse,
    ProcessedFilesResponse,
    UploadedFileInfo,
    ErrorsResponse,
    ErrorsSummaryResponse,
    AnalysisResponse,
    PromptsResponse,
    ApplyPromptResponse,
    ErrorInfo,
    PromptVersionInfo,
    TokenTestResponse,
    TokenUsageStats,
    PageTokenStats,
    # Dashboard Models
    DashboardStatsResponse,
    DashboardAnalyticsResponse,
    AnalyticsItem,
    DepartamentoItem,
    DisciplinaItem,
    RejectedConceptsResponse,
    RejectedConcept,
    # Periodos Models
    PeriodoInfo,
    PeriodoArchivoInfo,
    CreatePeriodoRequest,
    PeriodoResponse,
    PeriodosListResponse,
    PeriodoDetailResponse,
    PeriodoResumenPSResponse,
    PeriodoResumenPSItem,
    # Batch Models
    BatchProcessRequest,
    BatchProcessResponse,
    BatchJobInfo,
    ProcessSelectedRequest,
    # Auth Models
    LoginRequest,
    LoginResponse,
    LogoutRequest,
    LogoutResponse,
    # User Management Models
    UserInfo,
    UserCreateRequest,
    UserUpdateRequest,
    UserUpdateByAdminRequest,
    UserPasswordResetRequest,
    UsersListResponse,
    UserResponse,
    # Export Models
    BulkExportRequest,
    # Maestros Models
    ApartadoInfo,
    MaestrosSaveRequest,
    MaestrosResponse
)
from .dependencies import (
    get_ocr_extractor,
    get_file_manager,
    get_gemini_service,
    get_upload_manager,
    get_archive_manager,
    get_processed_tracker,
    is_email_allowed,
    add_email_to_allowed_list,
    remove_email_from_allowed_list,
    get_learning_system,
    get_periodo_manager,
    get_database_service
)
from .processing_worker import get_worker_manager, ProcessingJob
from .middleware import AuthMiddleware

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Suprimir warnings de archivos temporales en consola
logging.getLogger(__name__).setLevel(logging.INFO)

# Suprimir errores de asyncio en Windows (errores de limpieza al cerrar)
if sys.platform == 'win32':
    logging.getLogger('asyncio').setLevel(logging.CRITICAL)

# Rate Limiter
limiter = Limiter(key_func=get_remote_address)

# Crear aplicación FastAPI
app = FastAPI(
    title="ExtractorOCR API",
    description="API REST para procesamiento de PDFs con OCR usando Gemini Vision",
    version="1.0.0",
    docs_url="/docs",
    redoc_url="/redoc"
)

# Agregar rate limiter a la app
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Exception handler para errores de validación de Pydantic
@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: StarletteRequest, exc: RequestValidationError):
    """
    Maneja errores de validación de Pydantic y retorna mensajes amigables.
    """
    errors = exc.errors()
    error_messages = []
    
    for error in errors:
        field = ".".join(str(loc) for loc in error.get("loc", []))
        error_type = error.get("type", "")
        error_msg = error.get("msg", "")
        
        # Mensajes personalizados según el campo y tipo de error
        if "periodo" in field.lower():
            if error_type == "value_error.missing":
                error_messages.append("Periodo vacío, considerar insertar uno")
            elif error_type == "value_error.str.min_length":
                error_messages.append("Periodo vacío, considerar insertar uno")
            elif "not_empty" in str(error_msg) or "Periodo vacío" in str(error_msg):
                error_messages.append("Periodo vacío, considerar insertar uno")
            else:
                error_messages.append("Periodo vacío, considerar insertar uno")
        elif "tipo" in field.lower():
            if error_type == "value_error.missing":
                error_messages.append("Periodo vacío, considerar insertar uno")
            elif "tipo_valid" in str(error_msg) or "onshore" in str(error_msg).lower() or "offshore" in str(error_msg).lower():
                error_messages.append("El tipo debe ser 'onshore' o 'offshore'")
            else:
                error_messages.append(f"Error en el campo Tipo: {error_msg}")
        else:
            # Mensaje genérico para otros campos
            if error_type == "value_error.missing":
                error_messages.append(f"El campo '{field}' es obligatorio")
            else:
                error_messages.append(f"Error en '{field}': {error_msg}")
    
    # Si no hay mensajes personalizados, usar el primero disponible
    if not error_messages:
        if errors:
            first_error = errors[0]
            error_messages.append(first_error.get("msg", "Error de validación"))
    
    # Retornar mensaje en formato compatible con FastAPI (detail como string)
    error_message = error_messages[0] if error_messages else "Error de validación"
    
    # FastAPI espera detail como string o lista de objetos, pero para compatibilidad con frontend
    # retornamos como string simple
    return JSONResponse(
        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
        content={
            "detail": error_message  # Formato estándar de FastAPI - string simple para compatibilidad con frontend
        }
    )

# CORS middleware - Permitir todos los orígenes
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Permitir todos los orígenes
    allow_credentials=True,  # Permitir credenciales para todos los orígenes
    allow_methods=["*"],  # Permitir todos los métodos HTTP
    allow_headers=["*"],  # Permitir todos los headers
    expose_headers=["*"],  # Exponer todos los headers
)


# Middleware de autenticación (preparado pero desactivado)
auth_middleware = AuthMiddleware()
# app.middleware("http")(auth_middleware)  # Descomentar cuando se active auth

# ===== Sistema de Autenticación Simple =====
# Almacenamiento en memoria de tokens activos
_active_tokens: Dict[str, Dict[str, Any]] = {}  # token -> {email, expires_at, created_at}
_tokens_lock = Lock()
TOKEN_EXPIRATION_HOURS = 24  # Tokens expiran después de 24 horas

# ===== Sistema de edición concurrente de periodos (en memoria) =====
class PeriodEditingRequest(BaseModel):
    email: str
    nombre: Optional[str] = None


_period_edit_locks: Dict[str, Dict[str, Any]] = {}  # periodo_id -> {email, nombre, since, last_seen}
_period_edit_lock = Lock()
PERIOD_EDIT_TTL_MINUTES = 5


def _cleanup_period_edit_locks() -> None:
    """
    Limpia bloqueos de edición expirados para evitar que se queden colgados
    si un usuario cierra el navegador sin salir explícitamente.
    """
    if not _period_edit_locks:
        return

    now = datetime.utcnow()
    expire_before = now - timedelta(minutes=PERIOD_EDIT_TTL_MINUTES)

    with _period_edit_lock:
        to_delete = []
        for periodo_id, info in _period_edit_locks.items():
            last_seen_str = info.get("last_seen")
            try:
                if last_seen_str:
                    last_seen = datetime.fromisoformat(last_seen_str)
                else:
                    # Si no hay last_seen, usar since como fallback
                    since_str = info.get("since")
                    last_seen = datetime.fromisoformat(since_str) if since_str else None
            except Exception:
                last_seen = None

            if last_seen is None or last_seen < expire_before:
                to_delete.append(periodo_id)

        for periodo_id in to_delete:
            _period_edit_locks.pop(periodo_id, None)


def generate_auth_token() -> str:
    """Genera un token de autenticación seguro."""
    return secrets.token_urlsafe(32)


def truncate_request_id_for_folder(request_id: str, max_length: int = 30) -> str:
    """
    Trunca el request_id para usarlo como nombre de carpeta, evitando rutas demasiado largas.
    
    Args:
        request_id: ID del request completo
        max_length: Longitud máxima (default: 30 caracteres)
        
    Returns:
        request_id truncado a max_length caracteres
    """
    if len(request_id) <= max_length:
        return request_id
    # Truncar y mantener los primeros caracteres (más importantes)
    return request_id[:max_length]

def generate_secure_password(length: int = 10) -> str:
    """
    Genera una contraseña aleatoria cumpliendo requisitos de complejidad.
    """
    lowercase_chars = string.ascii_lowercase
    uppercase_chars = string.ascii_uppercase
    numbers = string.digits
    special_chars = "!@#$%&*?"
    all_chars = lowercase_chars + uppercase_chars + numbers + special_chars

    while True:
        password = ''.join(secrets.choice(all_chars) for _ in range(length))
        if (
            any(c in lowercase_chars for c in password)
            and any(c in uppercase_chars for c in password)
            and any(c in numbers for c in password)
            and any(c in special_chars for c in password)
        ):
            return password


@app.post("/api/v1/periodos/{periodo_id}/editing/enter", tags=["Periodos"])
async def enter_period_editing(periodo_id: str, body: PeriodEditingRequest):
    """
    Marca que un usuario ha entrado a editar un periodo.
    No bloquea el acceso, solo informa si ya había otra persona editando.
    """
    _cleanup_period_edit_locks()
    now = datetime.utcnow().isoformat()

    with _period_edit_lock:
        existing = _period_edit_locks.get(periodo_id)

        # Si ya hay alguien distinto editando, informar pero no reemplazar su lock
        if existing and existing.get("email") != body.email:
            return {
                "success": True,
                "already_in_use": True,
                "editor_email": existing.get("email"),
                "editor_nombre": existing.get("nombre"),
                "since": existing.get("since"),
            }

        # Registrar / actualizar lock para este usuario
        since = existing.get("since") if existing and existing.get("email") == body.email else now
        _period_edit_locks[periodo_id] = {
            "email": body.email,
            "nombre": body.nombre,
            "since": since,
            "last_seen": now,
        }

    return {
        "success": True,
        "already_in_use": False,
        "editor_email": body.email,
        "editor_nombre": body.nombre,
        "since": since,
    }


@app.post("/api/v1/periodos/{periodo_id}/editing/heartbeat", tags=["Periodos"])
async def heartbeat_period_editing(periodo_id: str, body: PeriodEditingRequest):
    """
    Actualiza el heartbeat de un usuario que sigue con el periodo abierto.
    Si el lock ya no existe o pertenece a otro usuario, se devuelve active=False.
    """
    _cleanup_period_edit_locks()
    now = datetime.utcnow().isoformat()

    with _period_edit_lock:
        existing = _period_edit_locks.get(periodo_id)
        if existing and existing.get("email") == body.email:
            existing["last_seen"] = now
            return {"success": True, "active": True}

    return {"success": True, "active": False}


@app.post("/api/v1/periodos/{periodo_id}/editing/leave", tags=["Periodos"])
async def leave_period_editing(periodo_id: str, body: PeriodEditingRequest):
    """
    Libera explícitamente el lock de edición cuando el usuario sale del periodo.
    """
    with _period_edit_lock:
        existing = _period_edit_locks.get(periodo_id)
        if existing and existing.get("email") == body.email:
            _period_edit_locks.pop(periodo_id, None)

    return {"success": True}


def generate_password_from_email(email: str) -> str:
    """
    Genera una contraseña compleja de 8 caracteres basada en el email.
    
    La contraseña incluye:
    - Minúsculas
    - Mayúsculas
    - Números
    - Caracteres especiales
    - Longitud exacta de 8 caracteres
    
    La contraseña es determinística (siempre la misma para el mismo email).
    
    Args:
        email: Email del usuario
        
    Returns:
        Contraseña generada de 8 caracteres
    """
    email_normalized = email.lower().strip()
    
    # Crear un hash del email para obtener valores determinísticos
    hash_obj = hashlib.md5(email_normalized.encode())
    hash_hex = hash_obj.hexdigest()
    
    # Mapeo de caracteres para asegurar diversidad
    lowercase_chars = "abcdefghijklmnopqrstuvwxyz"
    uppercase_chars = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
    numbers = "0123456789"
    special_chars = "!@#$%&*"
    
    # Usar el hash para seleccionar caracteres de manera determinística
    # Asegurar que tenga al menos: 1 minúscula, 1 mayúscula, 1 número, 1 especial
    password_chars = []
    
    # Posición 0: minúscula (usar hash[0])
    password_chars.append(lowercase_chars[int(hash_hex[0], 16) % len(lowercase_chars)])
    
    # Posición 1: mayúscula (usar hash[1])
    password_chars.append(uppercase_chars[int(hash_hex[1], 16) % len(uppercase_chars)])
    
    # Posición 2: número (usar hash[2])
    password_chars.append(numbers[int(hash_hex[2], 16) % len(numbers)])
    
    # Posición 3: especial (usar hash[3])
    password_chars.append(special_chars[int(hash_hex[3], 16) % len(special_chars)])
    
    # Posiciones 4-7: mezcla de todos los tipos (usar hash[4-7])
    all_chars = lowercase_chars + uppercase_chars + numbers + special_chars
    for i in range(4, 8):
        password_chars.append(all_chars[int(hash_hex[i], 16) % len(all_chars)])
    
    # Mezclar las posiciones para que no sea predecible el orden
    # Usar más valores del hash para determinar el orden
    hash_int = int(hash_hex[8:16], 16)
    shuffled = list(password_chars)
    
    # Aplicar un shuffle simple basado en el hash
    for i in range(7, 0, -1):
        j = hash_int % (i + 1)
        shuffled[i], shuffled[j] = shuffled[j], shuffled[i]
        hash_int = hash_int // (i + 1)
    
    password = ''.join(shuffled)
    
    # Asegurar que tiene al menos uno de cada tipo (verificación final)
    has_lower = any(c.islower() for c in password)
    has_upper = any(c.isupper() for c in password)
    has_digit = any(c.isdigit() for c in password)
    has_special = any(c in special_chars for c in password)
    
    # Si falta algún tipo, reemplazar en una posición específica
    if not has_lower:
        password = password[0].lower() + password[1:]
    if not has_upper:
        password = password[:1] + password[1].upper() + password[2:]
    if not has_digit:
        password = password[:2] + numbers[int(hash_hex[4], 16) % len(numbers)] + password[3:]
    if not has_special:
        password = password[:3] + special_chars[int(hash_hex[5], 16) % len(special_chars)] + password[4:]
    
    return password[:8]  # Asegurar exactamente 8 caracteres


def get_or_generate_password(email: str) -> str:
    """
    Obtiene la contraseña generada para un email, o la genera si no existe.
    
    La contraseña se guarda en config/user_passwords.json para que sea consistente.
    
    Args:
        email: Email del usuario
        
    Returns:
        Contraseña del usuario (generada o existente)
    """
    email_normalized = email.lower().strip()
    config_paths = [
        Path(__file__).parent.parent.parent.parent / "config" / "user_passwords.json",
        Path("./config/user_passwords.json"),
        Path(__file__).parent.parent.parent / "config" / "user_passwords.json"
    ]
    
    passwords_file = None
    for path in config_paths:
        if path.exists():
            passwords_file = path
            break
    
    # Si no existe el archivo, crear uno nuevo
    if not passwords_file:
        passwords_file = Path(__file__).parent.parent.parent.parent / "config" / "user_passwords.json"
        passwords_file.parent.mkdir(parents=True, exist_ok=True)
        with open(passwords_file, 'w', encoding='utf-8') as f:
            json.dump({"passwords": {}}, f, indent=2, ensure_ascii=False)
    
    # Leer contraseñas existentes
    try:
        with open(passwords_file, 'r', encoding='utf-8') as f:
            passwords_data = json.load(f)
    except Exception:
        passwords_data = {"passwords": {}}
    
    passwords = passwords_data.get("passwords", {})
    
    # Si ya existe la contraseña, retornarla
    if email_normalized in passwords:
        return passwords[email_normalized]
    
    # Si no existe, generar una nueva y guardarla
    new_password = generate_password_from_email(email_normalized)
    passwords[email_normalized] = new_password
    passwords_data["passwords"] = passwords
    
    # Guardar en el archivo
    try:
        with open(passwords_file, 'w', encoding='utf-8') as f:
            json.dump(passwords_data, f, indent=2, ensure_ascii=False)
        logger.info(f"Contraseña generada para {email_normalized}")
        
        # Sincronizar a BD como respaldo (no crítico si falla)
        try:
            from .dependencies import get_database_service
            db_service = get_database_service()
            nombre = format_name_from_email(email_normalized)
            db_service.sync_usuario_to_db(email_normalized, new_password, nombre)
        except Exception as e:
            # No es crítico si falla la sincronización a BD
            logger.debug(f"No se pudo sincronizar usuario a BD (no crítico): {e}")
    except Exception as e:
        logger.error(f"Error guardando contraseña: {e}")
    
    return new_password


def verify_password(email: str, provided_password: str) -> Tuple[bool, str]:
    """
    Verifica si la contraseña proporcionada es correcta para el email.
    
    TEMPORALMENTE: Intenta validar desde BD primero, si falla usa JSON.
    Después de pruebas, volveremos a usar solo JSON como fuente de verdad.
    
    Args:
        email: Email del usuario
        provided_password: Contraseña proporcionada
        
    Returns:
        Tupla (bool, str): (True/False si es válida, fuente_usada: "BD" o "JSON")
    """
    email_normalized = email.lower().strip()
    
    # TEMPORAL: Intentar validar desde BD primero (solo para pruebas)
    try:
        from .dependencies import get_database_service
        db_service = get_database_service()
        if db_service.is_enabled():
            # Intentar validar desde BD
            if db_service.verify_password_from_db(email_normalized, provided_password):
                logger.info(f"✓ [BD] Contraseña validada desde BASE DE DATOS para: {email_normalized}")
                print(f"✓ [BD] Validación desde BASE DE DATOS para: {email_normalized}")
                return True, "BD"
            # Si BD está habilitada pero falla, continuar con JSON como fallback
            logger.info(f"⚠ [BD→JSON] Validación desde BD falló, usando JSON como fallback para: {email_normalized}")
            print(f"⚠ [BD→JSON] BD falló, usando JSON para: {email_normalized}")
    except Exception as e:
        logger.info(f"⚠ [JSON] No se pudo validar desde BD (usando JSON): {e}")
        print(f"⚠ [JSON] No se pudo validar desde BD, usando JSON: {e}")
    
    # Fuente de verdad: JSON (o fallback si BD no está disponible)
    expected_password = get_or_generate_password(email_normalized)
    
    # Comparar contraseñas (case-sensitive)
    is_valid = provided_password == expected_password
    
    if is_valid:
        logger.info(f"✓ [JSON] Contraseña validada desde JSON para: {email_normalized}")
        print(f"✓ [JSON] Validación desde JSON para: {email_normalized}")
    else:
        logger.warning(f"✗ [JSON] Contraseña incorrecta desde JSON para: {email_normalized}")
        print(f"✗ [JSON] Contraseña incorrecta desde JSON para: {email_normalized}")
    
    return is_valid, "JSON"


def format_name_from_email(email: str) -> str:
    """
    Extrae y formatea el nombre desde un email.
    
    Ejemplo:
        "victor.cabeza@newmont.com" -> "Victor Cabeza"
        "mariadelosangeles.abanto@newmont.com" -> "Mariadelosangeles Abanto"
    
    Args:
        email: Email del usuario
        
    Returns:
        Nombre formateado con primera letra en mayúscula
    """
    # Extraer la parte antes del @
    email_part = email.split('@')[0] if '@' in email else email
    
    # Dividir por punto y capitalizar cada palabra
    name_parts = email_part.split('.')
    formatted_parts = [part.capitalize() for part in name_parts if part]
    
    # Unir con espacio
    return ' '.join(formatted_parts)


def get_user_passwords_file() -> Path:
    """Obtiene la ruta del archivo user_passwords.json."""
    config_paths = [
        Path(__file__).parent.parent.parent.parent / "config" / "user_passwords.json",
        Path("./config/user_passwords.json"),
        Path(__file__).parent.parent.parent / "config" / "user_passwords.json"
    ]
    
    for path in config_paths:
        if path.exists():
            return path
    
    # Si no existe, crear uno nuevo
    passwords_file = Path(__file__).parent.parent.parent.parent / "config" / "user_passwords.json"
    passwords_file.parent.mkdir(parents=True, exist_ok=True)
    with open(passwords_file, 'w', encoding='utf-8') as f:
        json.dump({"passwords": {}, "users": {}}, f, indent=2, ensure_ascii=False)
    return passwords_file


def load_users_data() -> Dict[str, Any]:
    """Carga los datos de usuarios desde user_passwords.json."""
    passwords_file = get_user_passwords_file()
    try:
        with open(passwords_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {"passwords": {}, "users": {}}


def save_users_data(data: Dict[str, Any]) -> bool:
    """Guarda los datos de usuarios en user_passwords.json."""
    passwords_file = get_user_passwords_file()
    try:
        with open(passwords_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logger.error(f"Error guardando datos de usuarios: {e}")
        return False


# ===== Maestros Helper Functions =====

def get_maestros_file() -> Path:
    """Obtiene la ruta del archivo maestros_apartados.json dentro de ExtractorOCRv1/config/."""
    # Desde main.py: ExtractorOCRv1/src/api/main.py
    # Subir 3 niveles: src/api -> src -> ExtractorOCRv1
    # Luego ir a config/maestros_apartados.json
    base_path = Path(__file__).parent.parent.parent  # ExtractorOCRv1
    config_paths = [
        base_path / "config" / "maestros_apartados.json",  # ExtractorOCRv1/config/maestros_apartados.json
        Path("./config/maestros_apartados.json"),
        Path("ExtractorOCRv1/config/maestros_apartados.json"),
    ]
    
    for path in config_paths:
        if path.exists():
            return path.resolve()
    
    # Si no existe, crear uno nuevo en ExtractorOCRv1/config/
    maestros_file = base_path / "config" / "maestros_apartados.json"
    maestros_file.parent.mkdir(parents=True, exist_ok=True)
    with open(maestros_file, 'w', encoding='utf-8') as f:
        json.dump({"apartados": []}, f, indent=2, ensure_ascii=False)
    return maestros_file.resolve()


def load_maestros_data() -> Dict[str, Any]:
    """Carga los datos de maestros desde maestros_apartados.json."""
    maestros_file = get_maestros_file()
    try:
        with open(maestros_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {"apartados": []}


def save_maestros_data(data: Dict[str, Any]) -> bool:
    """Guarda los datos de maestros en maestros_apartados.json."""
    maestros_file = get_maestros_file()
    try:
        with open(maestros_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logger.error(f"Error guardando datos de maestros: {e}")
        return False


def get_user_role(email: str) -> str:
    """Obtiene el rol de un usuario."""
    email_normalized = email.lower().strip()
    data = load_users_data()
    users = data.get("users", {})
    if email_normalized in users:
        return users[email_normalized].get("role", "user")
    # Si no está en users, verificar si es admin por defecto
    if email_normalized == "luis.saenz@newmont.com":
        return "admin"
    return "user"


def is_admin(email: str) -> bool:
    """Verifica si un usuario es admin."""
    return get_user_role(email) == "admin"


def create_auth_token(email: str) -> str:
    """
    Crea un token de autenticación para un email.
    
    Args:
        email: Email del usuario
        
    Returns:
        Token generado
    """
    token = generate_auth_token()
    expires_at = datetime.now() + timedelta(hours=TOKEN_EXPIRATION_HOURS)
    
    with _tokens_lock:
        _active_tokens[token] = {
            "email": email.lower().strip(),
            "expires_at": expires_at.isoformat(),
            "created_at": datetime.now().isoformat()
        }
    
    logger.info(f"Token creado para email: {email}")
    return token


def validate_auth_token(token: str) -> Optional[str]:
    """
    Valida un token de autenticación y retorna el email asociado.
    
    Args:
        token: Token a validar
        
    Returns:
        Email asociado al token, o None si el token es inválido/expirado
    """
    if not token:
        return None
    
    with _tokens_lock:
        token_data = _active_tokens.get(token)
        
        if not token_data:
            return None
        
        # Verificar expiración
        expires_at = datetime.fromisoformat(token_data["expires_at"])
        if datetime.now() > expires_at:
            # Token expirado, eliminarlo
            _active_tokens.pop(token, None)
            logger.info(f"Token expirado eliminado: {token[:8]}...")
            return None
        
        return token_data["email"]


def revoke_auth_token(token: str) -> bool:
    """
    Revoca (invalida) un token de autenticación.
    
    Args:
        token: Token a revocar
        
    Returns:
        True si se revocó, False si no existía
    """
    with _tokens_lock:
        if token in _active_tokens:
            del _active_tokens[token]
            logger.info(f"Token revocado: {token[:8]}...")
            return True
        return False


def cleanup_expired_tokens():
    """Limpia tokens expirados del almacenamiento."""
    now = datetime.now()
    expired_count = 0
    
    with _tokens_lock:
        tokens_to_remove = []
        for token, data in _active_tokens.items():
            expires_at = datetime.fromisoformat(data["expires_at"])
            if now > expires_at:
                tokens_to_remove.append(token)
        
        for token in tokens_to_remove:
            del _active_tokens[token]
            expired_count += 1
    
    if expired_count > 0:
        logger.info(f"Limpiados {expired_count} tokens expirados")


def _normalize_month(month: str) -> str:
    """
    Normaliza el mes a formato estándar.
    
    Args:
        month: Mes en formato string o int
        
    Returns:
        Nombre del mes en español (normalizado)
    """
    month_lower = month.lower().strip()
    
    # Si es número, convertir a nombre
    month_map_int = {
        '1': 'enero', '2': 'febrero', '3': 'marzo', '4': 'abril',
        '5': 'mayo', '6': 'junio', '7': 'julio', '8': 'agosto',
        '9': 'septiembre', '10': 'octubre', '11': 'noviembre', '12': 'diciembre'
    }
    
    if month_lower in month_map_int:
        return month_map_int[month_lower].capitalize()
    
    # Mapeo de meses en diferentes idiomas/formats
    month_map = {
        'enero': 'Enero', 'january': 'Enero', 'jan': 'Enero',
        'febrero': 'Febrero', 'february': 'Febrero', 'feb': 'Febrero',
        'marzo': 'Marzo', 'march': 'Marzo', 'mar': 'Marzo',
        'abril': 'Abril', 'april': 'Abril', 'apr': 'Abril',
        'mayo': 'Mayo', 'may': 'Mayo',
        'junio': 'Junio', 'june': 'Junio', 'jun': 'Junio',
        'julio': 'Julio', 'july': 'Julio', 'jul': 'Julio',
        'agosto': 'Agosto', 'august': 'Agosto', 'aug': 'Agosto',
        'septiembre': 'Septiembre', 'september': 'Septiembre', 'sep': 'Septiembre',
        'octubre': 'Octubre', 'october': 'Octubre', 'oct': 'Octubre',
        'noviembre': 'Noviembre', 'november': 'Noviembre', 'nov': 'Noviembre',
        'diciembre': 'Diciembre', 'december': 'Diciembre', 'dec': 'Diciembre'
    }
    
    return month_map.get(month_lower, month.capitalize())


@app.get("/", tags=["General"])
async def root():
    """Endpoint raíz con información de la API."""
    return {
        "name": "ExtractorOCR API",
        "version": "1.0.0",
        "status": "running",
        "docs": "/docs"
    }


@app.get("/health", response_model=HealthResponse, tags=["General"])
async def health_check():
    """
    Health check endpoint.
    Verifica que la API esté funcionando correctamente.
    """
    return HealthResponse(
        status="healthy",
        version="1.0.0",
        timestamp=datetime.now()
    )


# ===== Auth Setup =====
security = HTTPBearer(auto_error=False)


@app.post("/api/v1/test-tokens", response_model=TokenTestResponse, tags=["Testing"])
async def test_token_usage(
    pdf_file: UploadFile = File(..., description="Archivo PDF para probar"),
    page_number: Optional[int] = Form(default=None, description="Número de página específica a procesar (1-indexed). Si es None o 0, procesa todas las páginas"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Endpoint de prueba para calcular el uso de tokens de Gemini API.
    
    Procesa una página específica o todas las páginas del PDF y calcula:
    - Tokens de entrada (imagen + prompt)
    - Tokens de salida (respuesta)
    - Estadísticas detalladas por página y totales
    
    Args:
        pdf_file: Archivo PDF a probar
        page_number: Número de página a procesar (None o 0 = todas las páginas, 1-N = página específica)
        credentials: Credenciales de autenticación (opcional)
        
    Returns:
        TokenTestResponse con estadísticas de uso de tokens por página y totales
    """
    import time
    from PIL import Image
    from src.core.pdf_processor import PDFProcessor
    from src.core.file_manager import FileManager
    
    start_time = time.time()
    
    try:
        # Validar autenticación si está habilitada
        if credentials:
            token = credentials.credentials
            if not validate_auth_token(token):
                raise HTTPException(
                    status_code=status.HTTP_401_UNAUTHORIZED,
                    detail="Token inválido o expirado"
                )
        
        # Guardar PDF temporalmente
        file_manager = get_file_manager()
        temp_folder_str = file_manager.get_temp_folder()
        temp_folder = Path(temp_folder_str)  # Convertir string a Path
        temp_folder.mkdir(parents=True, exist_ok=True)
        
        temp_pdf_path = temp_folder / f"test_{uuid.uuid4()}.pdf"
        with open(temp_pdf_path, 'wb') as f:
            content = await pdf_file.read()
            f.write(content)
        
        # Extraer página específica como imagen
        pdf_processor = PDFProcessor()
        if not pdf_processor.open_pdf(str(temp_pdf_path)):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No se pudo abrir el PDF"
            )
        
        total_pages = pdf_processor.get_page_count()
        
        # Determinar qué páginas procesar
        if page_number is None or page_number == 0:
            # Procesar todas las páginas
            pages_to_process = list(range(1, total_pages + 1))
        else:
            # Validar página específica
            if page_number < 1 or page_number > total_pages:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Página {page_number} no válida. El PDF tiene {total_pages} página(s)"
                )
            pages_to_process = [page_number]
        
        # Obtener servicio Gemini
        gemini_service = get_gemini_service()
        prompt = gemini_service._create_ocr_and_structure_prompt()
        prompt_length_chars = len(prompt)
        
        # Obtener configuración de DPI y delay desde config
        image_dpi = gemini_service.config.get("image_dpi", 200)  # Default: 200 DPI (reducido de 300)
        page_delay = gemini_service.config.get("page_delay_seconds", 0.5)  # Default: 0.5 segundos
        
        # Procesar cada página
        page_stats_list = []
        total_input_tokens_estimated = 0
        total_output_tokens_estimated = 0
        total_input_tokens_actual = 0
        total_output_tokens_actual = 0
        total_prompt_length = 0
        total_response_length = 0
        total_image_size = 0
        
        for idx, current_page in enumerate(pages_to_process):
            page_start_time = time.time()
            temp_img_path = None
            
            # Agregar delay entre páginas para evitar exceder cuota (excepto la primera)
            # Delay dinámico basado en tokens estimados: más tokens = más delay
            if idx > 0:
                # Calcular delay basado en tokens estimados de la página anterior
                if page_stats_list:
                    last_tokens = page_stats_list[-1].token_stats.input_tokens_estimated
                    # Delay base desde config, más 0.01s por cada 1000 tokens estimados
                    delay_seconds = page_delay + (last_tokens / 1000) * 0.01
                    delay_seconds = min(delay_seconds, 3.0)  # Máximo 3 segundos
                else:
                    delay_seconds = page_delay
                time.sleep(delay_seconds)
            
            try:
                # Extraer página como imagen (convertir de 1-indexed a 0-indexed)
                # Usar DPI configurado (default: 200 en lugar de 300) para reducir tamaño de imagen y tokens
                temp_img_path = temp_folder / f"test_page_{current_page}_{uuid.uuid4()}.png"
                page_index = current_page - 1  # Convertir de 1-indexed a 0-indexed
                if not pdf_processor.save_page_as_image(page_index, temp_img_path, dpi=image_dpi):
                    page_stats_list.append(PageTokenStats(
                        page_number=current_page,
                        token_stats=TokenUsageStats(
                            input_tokens_estimated=0,
                            output_tokens_estimated=0,
                            total_tokens_estimated=0,
                            prompt_length_chars=prompt_length_chars,
                            response_length_chars=0,
                            image_size_bytes=0,
                            image_dimensions=None
                        ),
                        processing_time_seconds=0,
                        error="No se pudo extraer la página como imagen"
                    ))
                    continue
                
                # Obtener información de la imagen
                img = Image.open(temp_img_path)
                image_size_bytes = temp_img_path.stat().st_size
                image_dimensions = {"width": img.size[0], "height": img.size[1]}
                
                # Estimar tokens de entrada
                image_tokens_estimated = max(1, image_size_bytes // 256)
                prompt_tokens_estimated = prompt_length_chars // 4
                input_tokens_estimated = image_tokens_estimated + prompt_tokens_estimated
                
                # Llamar a Gemini
                try:
                    response = gemini_service.model.generate_content([prompt, img])
                    
                    # Intentar obtener tokens reales
                    input_tokens_actual = None
                    output_tokens_actual = None
                    
                    if hasattr(response, 'usage_metadata'):
                        usage = response.usage_metadata
                        if hasattr(usage, 'prompt_token_count'):
                            input_tokens_actual = usage.prompt_token_count
                        if hasattr(usage, 'candidates_token_count'):
                            output_tokens_actual = usage.candidates_token_count
                    
                    if input_tokens_actual is None:
                        input_tokens_actual = input_tokens_estimated
                    
                    response_text = response.text if response.text else ""
                    response_length_chars = len(response_text)
                    output_tokens_estimated = response_length_chars // 4
                    
                    if output_tokens_actual is None:
                        output_tokens_actual = output_tokens_estimated
                    
                    # Acumular totales
                    total_input_tokens_estimated += input_tokens_estimated
                    total_output_tokens_estimated += output_tokens_estimated
                    total_input_tokens_actual += input_tokens_actual
                    total_output_tokens_actual += output_tokens_actual
                    total_prompt_length += prompt_length_chars
                    total_response_length += response_length_chars
                    total_image_size += image_size_bytes
                    
                    page_processing_time = time.time() - page_start_time
                    
                    page_stats_list.append(PageTokenStats(
                        page_number=current_page,
                        token_stats=TokenUsageStats(
                            input_tokens_estimated=input_tokens_estimated,
                            output_tokens_estimated=output_tokens_estimated,
                            total_tokens_estimated=input_tokens_estimated + output_tokens_estimated,
                            input_tokens_actual=input_tokens_actual,
                            output_tokens_actual=output_tokens_actual,
                            total_tokens_actual=input_tokens_actual + output_tokens_actual,
                            prompt_length_chars=prompt_length_chars,
                            response_length_chars=response_length_chars,
                            image_size_bytes=image_size_bytes,
                            image_dimensions=image_dimensions
                        ),
                        processing_time_seconds=page_processing_time,
                        error=None
                    ))
                    
                except Exception as api_error:
                    error_msg = str(api_error)
                    page_processing_time = time.time() - page_start_time
                    
                    # Acumular tokens estimados incluso si hay error (para estadísticas)
                    total_input_tokens_estimated += input_tokens_estimated
                    total_prompt_length += prompt_length_chars
                    total_image_size += image_size_bytes
                    
                    page_stats_list.append(PageTokenStats(
                        page_number=current_page,
                        token_stats=TokenUsageStats(
                            input_tokens_estimated=input_tokens_estimated,
                            output_tokens_estimated=0,
                            total_tokens_estimated=input_tokens_estimated,
                            prompt_length_chars=prompt_length_chars,
                            response_length_chars=0,
                            image_size_bytes=image_size_bytes,
                            image_dimensions=image_dimensions
                        ),
                        processing_time_seconds=page_processing_time,
                        error=error_msg
                    ))
                    
                    # Si es error 429, esperar antes de continuar con la siguiente página
                    if "429" in error_msg or "quota" in error_msg.lower() or "exhausted" in error_msg.lower():
                        wait_time = 2  # Esperar 2 segundos entre páginas si hay error 429
                        print(f"Page {current_page} - Quota error, waiting {wait_time}s before next page...")
                        time.sleep(wait_time)
                    
            finally:
                # Limpiar imagen temporal
                if temp_img_path and temp_img_path.exists():
                    try:
                        temp_img_path.unlink()
                    except Exception:
                        pass
        
        # Calcular totales
        total_processing_time = time.time() - start_time
        
        # Limpiar PDF temporal
        try:
            temp_pdf_path.unlink()
        except Exception:
            pass
        
        # Crear estadísticas totales
        total_token_stats = TokenUsageStats(
            input_tokens_estimated=total_input_tokens_estimated,
            output_tokens_estimated=total_output_tokens_estimated,
            total_tokens_estimated=total_input_tokens_estimated + total_output_tokens_estimated,
            input_tokens_actual=total_input_tokens_actual if total_input_tokens_actual > 0 else None,
            output_tokens_actual=total_output_tokens_actual if total_output_tokens_actual > 0 else None,
            total_tokens_actual=(total_input_tokens_actual + total_output_tokens_actual) if (total_input_tokens_actual > 0 or total_output_tokens_actual > 0) else None,
            prompt_length_chars=total_prompt_length,
            response_length_chars=total_response_length,
            image_size_bytes=total_image_size,
            image_dimensions=None
        )
        
        return TokenTestResponse(
            success=True,
            filename=pdf_file.filename,
            total_pages=total_pages,
            pages_processed=len(pages_to_process),
            page_stats=page_stats_list,
            total_token_stats=total_token_stats,
            total_processing_time_seconds=total_processing_time,
            model_used=gemini_service.model_name,
            max_output_tokens=gemini_service.config.get("max_output_tokens", 65536),
            error=None
        )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en test-tokens: {e}")
        import traceback
        logger.debug(f"Traceback: {traceback.format_exc()}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error procesando PDF: {str(e)}"
        )


# ===== Auth Endpoints =====

def get_current_user_email(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
) -> Optional[str]:
    """
    Dependencia para obtener el email del usuario autenticado desde el token.
    
    Args:
        credentials: Credenciales HTTP Bearer (opcional)
        
    Returns:
        Email del usuario autenticado, o None si no hay token válido
        
    Uso:
        @app.post("/api/v1/endpoint")
        async def my_endpoint(user_email: str = Depends(get_current_user_email)):
            if not user_email:
                raise HTTPException(status_code=401, detail="No autenticado")
            # Usar user_email...
    """
    if not credentials:
        return None
    
    token = credentials.credentials
    email = validate_auth_token(token)
    return email


@app.post("/api/v1/login", response_model=LoginResponse, tags=["Auth"])
async def login(request: LoginRequest):
    """
    Endpoint de login que valida el email y contraseña contra la lista de correos autorizados.
    
    Si el email está autorizado y la contraseña es correcta, genera un token de sesión
    que puede usarse en otros endpoints en lugar de proporcionar el email cada vez.
    
    La contraseña se genera automáticamente la primera vez basada en el email.
    Se guarda en config/user_passwords.json para mantener consistencia.
    
    Args:
        request: LoginRequest con el email y contraseña del usuario
        
    Returns:
        LoginResponse con token de autenticación y fecha de expiración
        
    Ejemplo:
        POST /api/v1/login
        {
            "email": "usuario@newmont.com",
            "password": "AZEBAC.ROTIV2024"
        }
        
        Respuesta:
        {
            "success": true,
            "token": "abc123...",
            "email": "usuario@newmont.com",
            "message": "Login exitoso",
            "expires_at": "2025-11-19T14:30:00"
        }
    """
    email = request.email.lower().strip()
    
    # Validar que el email esté autorizado
    if not is_email_allowed(email):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": f"Correo no autorizado: {email}. Acceso denegado.",
                "success": False
            }
        )
    
    # Validar contraseña
    if not request.password:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "error": "Contraseña requerida. Por favor, proporciona tu contraseña.",
                "success": False
            }
        )
    
    # Verificar contraseña (retorna tupla: (bool, fuente))
    is_valid, fuente_validacion = verify_password(email, request.password)
    
    if not is_valid:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={
                "error": "Contraseña incorrecta. Por favor, verifica tus credenciales.",
                "success": False,
                "fuente_validacion": fuente_validacion
            }
        )
    
    # Limpiar tokens expirados antes de crear uno nuevo
    cleanup_expired_tokens()
    
    # Crear token de autenticación
    token = create_auth_token(email)
    token_data = _active_tokens.get(token)
    
    # Obtener nombre, rol y estado del usuario
    data = load_users_data()
    users = data.get("users", {})
    user_data = users.get(email)

    if user_data:
        user_status = user_data.get("status", "active")
        if user_status != "active":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail={
                    "error": f"El usuario {email} está inhabilitado. Contacta a un administrador para reactivarlo.",
                    "success": False,
                },
            )

        nombre = user_data.get("nombre", format_name_from_email(email))
        role = user_data.get("role", "user")
    else:
        nombre = format_name_from_email(email)
        role = get_user_role(email)
    
    # Mensaje con información de la fuente usada
    mensaje = f"Login exitoso. Validado desde {fuente_validacion}. Usa este token en el header 'Authorization: Bearer <token>' para autenticarte."
    
    return LoginResponse(
        success=True,
        token=token,
        email=email,
        nombre=nombre,
        message=mensaje,
        expires_at=token_data["expires_at"] if token_data else None,
        fuente_validacion=fuente_validacion,
        role=role
    )


def get_current_user_email(credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)) -> Optional[str]:
    """Obtiene el email del usuario actual desde el token."""
    if not credentials:
        return None
    token = credentials.credentials
    token_data = _active_tokens.get(token)
    if token_data:
        email = token_data.get("email")
        if email:
            # Verificar si el usuario está activo antes de retornar el email
            data = load_users_data()
            users = data.get("users", {})
            user_data = users.get(email)
            
            if not user_data:
                # Usuario eliminado - invalidar token
                _active_tokens.pop(token, None)
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail={
                        "error": "Tu cuenta ha sido eliminada. Contacta con el administrador.",
                        "success": False,
                        "account_suspended": True
                    }
                )
            
            user_status = user_data.get("status", "active")
            if user_status != "active":
                # Usuario inhabilitado - invalidar token
                _active_tokens.pop(token, None)
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN,
                    detail={
                        "error": "Tu cuenta ha sido suspendida. Contacta con el administrador.",
                        "success": False,
                        "account_suspended": True
                    }
                )
        
        return email
    return None


@app.get("/api/v1/auth/me", response_model=UserResponse, tags=["Auth"])
async def get_current_user(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Obtiene la información del usuario actual autenticado.
    Verifica el estado del usuario y retorna error si está inhabilitado o eliminado.
    
    Returns:
        Información del usuario actual
    """
    current_email = get_current_user_email(credentials)
    
    if not current_email:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail={
                "error": "No autenticado. Por favor, inicia sesión.",
                "success": False
            }
        )
    
    # Verificar si el usuario existe y está activo
    data = load_users_data()
    users = data.get("users", {})
    user_data = users.get(current_email)
    
    if not user_data:
        # Usuario eliminado - invalidar token
        token = credentials.credentials if credentials else None
        if token and token in _active_tokens:
            _active_tokens.pop(token, None)
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": "Tu cuenta ha sido eliminada. Contacta con el administrador.",
                "success": False,
                "account_suspended": True
            }
        )
    
    user_status = user_data.get("status", "active")
    if user_status != "active":
        # Usuario inhabilitado - invalidar token
        token = credentials.credentials if credentials else None
        if token and token in _active_tokens:
            _active_tokens.pop(token, None)
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail={
                "error": "Tu cuenta ha sido suspendida. Contacta con el administrador.",
                "success": False,
                "account_suspended": True
            }
        )
    
    return UserResponse(
        success=True,
        message="Usuario autenticado",
        user=UserInfo(
            email=current_email,
            nombre=user_data.get("nombre", format_name_from_email(current_email)),
            role=user_data.get("role", "user"),
            status=user_data.get("status", "active"),
            ultima_edicion=user_data.get("ultima_edicion")
        )
    )


@app.get("/api/v1/users", response_model=UsersListResponse, tags=["Users"])
async def list_users(
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Lista todos los usuarios (solo admin).
    
    Returns:
        Lista de usuarios con su información
    """
    current_email = get_current_user_email(credentials)
    if not current_email or not is_admin(current_email):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo los administradores pueden ver la lista de usuarios"
        )
    
    data = load_users_data()
    users = data.get("users", {})
    passwords = data.get("passwords", {})
    
    user_list = []
    for email, user_data in users.items():
        user_list.append(UserInfo(
            email=email,
            nombre=user_data.get("nombre", format_name_from_email(email)),
            role=user_data.get("role", "user"),
            status=user_data.get("status", "active"),
            ultima_edicion=user_data.get("ultima_edicion")
        ))
    
    # Incluir usuarios que están en passwords pero no en users (backward compatibility)
    for email, password in passwords.items():
        if email not in users:
            user_list.append(UserInfo(
                email=email,
                nombre=format_name_from_email(email),
                role=get_user_role(email),
                status="active",
                ultima_edicion=None
            ))
    
    return UsersListResponse(
        success=True,
        users=user_list,
        total=len(user_list)
    )


@app.post("/api/v1/users", response_model=UserResponse, tags=["Users"])
async def create_user(
    request: UserCreateRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Crea un nuevo usuario (solo admin).
    
    Args:
        request: Datos del usuario a crear
        
    Returns:
        Usuario creado
    """
    current_email = get_current_user_email(credentials)
    if not current_email or not is_admin(current_email):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo los administradores pueden crear usuarios"
        )
    
    email_normalized = request.email.lower().strip()
    data = load_users_data()
    users = data.get("users", {})
    passwords = data.get("passwords", {})
    
    if email_normalized in users:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"El usuario {email_normalized} ya existe"
        )
    
    # Generar contraseña si no se proporciona
    password = request.password or generate_password_from_email(email_normalized)
    
    # Crear usuario
    users[email_normalized] = {
        "password": password,
        "role": request.role,
        "status": "active",
        "nombre": request.nombre,
        "ultima_edicion": datetime.now().isoformat()
    }
    passwords[email_normalized] = password
    
    data["users"] = users
    data["passwords"] = passwords
    
    if not save_users_data(data):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al guardar el usuario"
        )
    
    # Agregar email a la lista de correos autorizados
    if not add_email_to_allowed_list(email_normalized):
        logger.warning(f"No se pudo agregar {email_normalized} a allowed_emails.json, pero el usuario fue creado")
    
    return UserResponse(
        success=True,
        message=f"Usuario {email_normalized} creado exitosamente",
        user=UserInfo(
            email=email_normalized,
            nombre=request.nombre,
            role=request.role,
            status="active",
            ultima_edicion=users[email_normalized]["ultima_edicion"]
        )
    )


@app.put("/api/v1/users/{email}", response_model=UserResponse, tags=["Users"])
async def update_user(
    email: str,
    request: UserUpdateRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Actualiza el perfil del usuario (nombre, contraseña, avatar).
    Los usuarios solo pueden actualizar su propio perfil.
    
    Args:
        email: Email del usuario a actualizar
        request: Datos a actualizar
        
    Returns:
        Usuario actualizado
    """
    current_email = get_current_user_email(credentials)
    if not current_email:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Token de autenticación requerido"
        )
    
    email_normalized = email.lower().strip()
    
    # Los usuarios solo pueden actualizar su propio perfil (a menos que sean admin)
    if email_normalized != current_email.lower() and not is_admin(current_email):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo puedes actualizar tu propio perfil"
        )
    
    data = load_users_data()
    users = data.get("users", {})
    passwords = data.get("passwords", {})
    
    if email_normalized not in users:
        # Si no está en users, crear entrada básica
        users[email_normalized] = {
            "password": passwords.get(email_normalized, generate_password_from_email(email_normalized)),
            "role": get_user_role(email_normalized),
            "status": "active",
            "nombre": format_name_from_email(email_normalized),
            "ultima_edicion": None
        }
    
    # Actualizar campos
    if request.nombre:
        users[email_normalized]["nombre"] = request.nombre
    if request.password:
        users[email_normalized]["password"] = request.password
        passwords[email_normalized] = request.password
    if request.avatar_url:
        users[email_normalized]["avatar_url"] = request.avatar_url
    
    users[email_normalized]["ultima_edicion"] = datetime.now().isoformat()
    
    data["users"] = users
    data["passwords"] = passwords
    
    if not save_users_data(data):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al guardar los cambios"
        )
    
    return UserResponse(
        success=True,
        message="Perfil actualizado exitosamente",
        user=UserInfo(
            email=email_normalized,
            nombre=users[email_normalized]["nombre"],
            role=users[email_normalized]["role"],
            status=users[email_normalized]["status"],
            ultima_edicion=users[email_normalized]["ultima_edicion"]
        )
    )


@app.patch("/api/v1/users/{email}", response_model=UserResponse, tags=["Users"])
async def update_user_by_admin(
    email: str,
    request: UserUpdateByAdminRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Actualiza un usuario (solo admin): nombre, rol, estado.
    
    Args:
        email: Email del usuario a actualizar
        request: Datos a actualizar
        
    Returns:
        Usuario actualizado
    """
    current_email = get_current_user_email(credentials)
    if not current_email or not is_admin(current_email):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo los administradores pueden actualizar usuarios"
        )
    
    email_normalized = email.lower().strip()
    data = load_users_data()
    users = data.get("users", {})
    
    if email_normalized not in users:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Usuario {email_normalized} no encontrado"
        )
    
    # Actualizar campos
    if request.nombre:
        users[email_normalized]["nombre"] = request.nombre
    if request.role:
        users[email_normalized]["role"] = request.role
    if request.status:
        users[email_normalized]["status"] = request.status
        # Si se inhabilita el usuario, invalidar todos sus tokens activos
        if request.status == "inactive":
            tokens_to_remove = [
                token for token, token_data in _active_tokens.items()
                if token_data.get("email") == email_normalized
            ]
            for token in tokens_to_remove:
                _active_tokens.pop(token, None)
            logger.info(f"Tokens invalidados para usuario inhabilitado: {email_normalized}")
    
    users[email_normalized]["ultima_edicion"] = datetime.now().isoformat()
    
    data["users"] = users
    
    if not save_users_data(data):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al guardar los cambios"
        )
    
    return UserResponse(
        success=True,
        message="Usuario actualizado exitosamente",
        user=UserInfo(
            email=email_normalized,
            nombre=users[email_normalized]["nombre"],
            role=users[email_normalized]["role"],
            status=users[email_normalized]["status"],
            ultima_edicion=users[email_normalized]["ultima_edicion"]
        )
    )


@app.patch("/api/v1/users/{email}/password/reset", response_model=UserResponse, tags=["Users"])
async def reset_user_password(
    email: str,
    request: UserPasswordResetRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Restaura la contraseña de un usuario (solo admin).
    
    Args:
        email: Email del usuario
        request: Nueva contraseña (opcional, se genera si no se proporciona)
        
    Returns:
        Usuario actualizado
    """
    current_email = get_current_user_email(credentials)
    if not current_email or not is_admin(current_email):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo los administradores pueden restaurar contraseñas"
        )
    
    email_normalized = email.lower().strip()
    data = load_users_data()
    users = data.get("users", {})
    passwords = data.get("passwords", {})
    
    if email_normalized not in users:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Usuario {email_normalized} no encontrado"
        )
    
    # Generar o usar contraseña proporcionada (si no se envía, crear una nueva aleatoria)
    new_password = request.new_password or generate_secure_password()
    users[email_normalized]["password"] = new_password
    passwords[email_normalized] = new_password
    users[email_normalized]["ultima_edicion"] = datetime.now().isoformat()
    
    data["users"] = users
    data["passwords"] = passwords
    
    if not save_users_data(data):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al guardar la contraseña"
        )
    
    return UserResponse(
        success=True,
        message=f"Contraseña restaurada para {email_normalized}",
        user=UserInfo(
            email=email_normalized,
            nombre=users[email_normalized]["nombre"],
            role=users[email_normalized]["role"],
            status=users[email_normalized]["status"],
            ultima_edicion=users[email_normalized]["ultima_edicion"]
        ),
        new_password=new_password
    )


@app.delete("/api/v1/users/{email}", response_model=UserResponse, tags=["Users"])
async def delete_user(
    email: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Elimina un usuario (solo admin).

    Args:
        email: Email del usuario a eliminar

    Returns:
        Usuario eliminado
    """
    current_email = get_current_user_email(credentials)
    if not current_email or not is_admin(current_email):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Solo los administradores pueden eliminar usuarios"
        )

    email_normalized = email.lower().strip()

    if email_normalized == current_email.lower():
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No puedes eliminar tu propio perfil"
        )

    data = load_users_data()
    users = data.get("users", {})
    passwords = data.get("passwords", {})

    if email_normalized not in users:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Usuario {email_normalized} no encontrado"
        )

    deleted_user = users.pop(email_normalized)
    passwords.pop(email_normalized, None)

    data["users"] = users
    data["passwords"] = passwords

    # Invalidar tokens activos del usuario eliminado
    tokens_to_remove = [
        token for token, token_data in _active_tokens.items()
        if token_data.get("email") == email_normalized
    ]
    for token in tokens_to_remove:
        _active_tokens.pop(token, None)

    if not save_users_data(data):
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Error al eliminar el usuario"
        )
    
    # Remover email de la lista de correos autorizados
    if not remove_email_from_allowed_list(email_normalized):
        logger.warning(f"No se pudo remover {email_normalized} de allowed_emails.json, pero el usuario fue eliminado")
    
    return UserResponse(
        success=True,
        message=f"Usuario {email_normalized} eliminado exitosamente",
        user=UserInfo(
            email=email_normalized,
            nombre=deleted_user.get("nombre", format_name_from_email(email_normalized)),
            role=deleted_user.get("role", "user"),
            status=deleted_user.get("status", "inactive"),
            ultima_edicion=deleted_user.get("ultima_edicion")
        )
    )


@app.post("/api/v1/logout", response_model=LogoutResponse, tags=["Auth"])
async def logout(
    request: Optional[LogoutRequest] = None,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security),
    token: Optional[str] = Query(None, description="Token a invalidar (opcional, también puede enviarse en body o header)")
):
    """
    Endpoint de logout que invalida el token de sesión.
    
    El token puede enviarse de 3 formas (en orden de prioridad):
    1. En el header: Authorization: Bearer <token>
    2. En el body: {"token": "<token>"}
    3. Como query parameter: ?token=<token>
    
    Args:
        request: LogoutRequest con token en body (opcional)
        credentials: Credenciales HTTP Bearer con el token (opcional)
        token: Token como query parameter (opcional)
        
    Returns:
        LogoutResponse confirmando el logout
        
    Ejemplos:
        # Opción 1: Header
        POST /api/v1/logout
        Headers: Authorization: Bearer abc123...
        
        # Opción 2: Body
        POST /api/v1/logout
        Body: {"token": "abc123..."}
        
        # Opción 3: Query parameter
        POST /api/v1/logout?token=abc123...
    """
    # Obtener token de cualquiera de las 3 fuentes (en orden de prioridad)
    token_to_revoke = None
    
    # 1. Intentar desde header Authorization
    if credentials:
        token_to_revoke = credentials.credentials
    
    # 2. Intentar desde body
    if not token_to_revoke and request and request.token:
        token_to_revoke = request.token
    
    # 3. Intentar desde query parameter
    if not token_to_revoke and token:
        token_to_revoke = token
    
    # Validar que se proporcionó token
    if not token_to_revoke:
        return LogoutResponse(
            success=False,
            message="No se proporcionó token. Envía el token en el header 'Authorization: Bearer <token>', en el body como '{\"token\": \"<token>\"}', o como query parameter '?token=<token>'."
        )
    
    # Revocar token
    revoked = revoke_auth_token(token_to_revoke)
    
    if revoked:
        return LogoutResponse(
            success=True,
            message="Logout exitoso. Token invalidado."
        )
    else:
        return LogoutResponse(
            success=False,
            message="Token no encontrado o ya invalidado."
        )


@app.post("/api/v1/upload-pdf", response_model=UploadPDFResponse, tags=["Upload"])
async def upload_pdf(
    pdf_file: UploadFile = File(..., description="Archivo PDF a subir"),
    email: str = Form(..., description="Email del usuario"),
    year: int = Form(..., description="Año a procesar (2000-2100)"),
    month: str = Form(..., description="Mes a procesar (ej: 'Marzo' o '3')"),
    periodo_id: Optional[str] = Form(default=None, description="ID del periodo para asociar este archivo (opcional, ej: '2025-11-onshore')"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Sube y valida un PDF para procesamiento posterior.
    
    Puedes asociar el archivo a un periodo desde el momento de la subida,
    o hacerlo después al procesarlo.
    
    Args:
        pdf_file: Archivo PDF a subir
        email: Email del usuario
        year: Año a procesar
        month: Mes a procesar (string o int)
        periodo_id: ID del periodo para asociar este archivo (opcional)
        
    Returns:
        UploadPDFResponse con file_id para usar en process-pdf
        
    Notas:
        - Si proporcionas periodo_id, el archivo se asociará automáticamente al periodo al procesarlo
        - Si no proporcionas periodo_id, puedes especificarlo al procesar con process-pdf o process-batch
        - También puedes usar process-all para procesar todos los archivos de un periodo automáticamente
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    # Validar que sea PDF
    if not pdf_file.filename.lower().endswith('.pdf'):
        error_response = ErrorResponse(
            request_id="",
            error="El archivo debe ser un PDF",
            details={"filename": pdf_file.filename}
        )
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=error_response.model_dump()
        )
    
    # Normalizar mes
    normalized_month = _normalize_month(month)
    
    # Leer contenido del PDF
    try:
        pdf_content = await pdf_file.read()
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": f"Error leyendo archivo: {str(e)}"}
        )
    
    # Validar periodo_id si se proporciona
    if periodo_id:
        periodo_manager = get_periodo_manager()
        periodo_data = periodo_manager.get_periodo(periodo_id)
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail={"error": f"Periodo '{periodo_id}' no encontrado. Verifica que el periodo exista antes de asociar archivos."}
            )
        
        # Validar que año y mes coincidan con el periodo
        periodo_str = periodo_data.get("periodo", "")
        if "/" in periodo_str:
            mes_str, anio_str = periodo_str.split("/")
            try:
                periodo_year = int(anio_str)
                periodo_month_num = int(mes_str)
                periodo_month_normalized = _normalize_month(mes_str)
                
                if periodo_year != year:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail={"error": f"El año del archivo ({year}) no coincide con el año del periodo ({periodo_year})"}
                    )
                
                if periodo_month_normalized != normalized_month:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail={"error": f"El mes del archivo ({normalized_month}) no coincide con el mes del periodo ({periodo_month_normalized})"}
                    )
            except ValueError:
                logger.warning(f"No se pudo validar año/mes del periodo {periodo_id}")
    
    # Guardar PDF y metadata
    upload_manager = get_upload_manager()
    metadata = {
        "email": email,
        "year": year,
        "month": normalized_month
    }
    
    # Agregar periodo_id a metadata si se proporcionó
    if periodo_id:
        metadata["periodo_id"] = periodo_id
    
    try:
        # Guardar PDF y metadata (operación síncrona que completa antes de continuar)
        file_id = upload_manager.save_uploaded_pdf(
            pdf_content,
            pdf_file.filename,
            metadata
        )
        
        # Verificar que el archivo se guardó correctamente antes de responder
        # Esto asegura que la respuesta solo se devuelve cuando el archivo está completamente guardado
        pdf_path = upload_manager.get_uploaded_pdf_path(file_id)
        saved_metadata = upload_manager.get_uploaded_metadata(file_id)
        
        if not pdf_path or not pdf_path.exists():
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail={"error": "Error: El archivo no se guardó correctamente en el servidor"}
            )
        
        if not saved_metadata:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail={"error": "Error: La metadata no se guardó correctamente"}
            )
        
        # Verificar que el tamaño del archivo guardado coincide
        actual_file_size = pdf_path.stat().st_size
        if actual_file_size != len(pdf_content):
            logger.warning(f"Tamaño del archivo guardado ({actual_file_size}) no coincide con el esperado ({len(pdf_content)})")
        
        # Obtener timestamp de cuando se guardó realmente
        uploaded_at = datetime.now()
        if saved_metadata.get("uploaded_at"):
            try:
                uploaded_at = datetime.fromisoformat(saved_metadata["uploaded_at"])
            except Exception:
                pass
        
        logger.info(f"Archivo {file_id} ({pdf_file.filename}) guardado correctamente. Tamaño: {actual_file_size} bytes")
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error guardando PDF: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"error": f"Error guardando archivo en servidor: {str(e)}"}
        )
    
    # Respuesta solo se devuelve después de confirmar que el archivo está completamente guardado
    # Incluir información adicional para que el frontend pueda actualizar la UI inmediatamente
    # sin necesidad de recargar la página o hacer otra consulta
    response = UploadPDFResponse(
        success=True,
        file_id=file_id,
        filename=pdf_file.filename,
        uploaded_at=uploaded_at,
        metadata=metadata,
        file_size_bytes=len(pdf_content)
    )
    
    # Convertir a dict y agregar campos adicionales que el frontend necesita
    # para actualizar la tabla sin recargar
    response_dict = response.model_dump()
    
    # Agregar información adicional en formato compatible con PeriodoArchivoInfo
    # Esto permite que el frontend agregue el archivo directamente a la tabla
    response_dict["archivo_id"] = file_id[:8] if len(file_id) >= 8 else file_id  # Primeros 8 caracteres para mostrar
    response_dict["request_id"] = file_id  # request_id = file_id (para compatibilidad)
    response_dict["estado"] = "pendiente"  # Estado inicial después de subir
    
    return response_dict


@app.delete("/api/v1/upload-pdf/{file_id}", response_model=DeleteUploadResponse, tags=["Upload"])
async def delete_uploaded_pdf(
    file_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Elimina un PDF subido y su metadata.
    
    Este endpoint elimina tanto el archivo PDF como su metadata asociada.
    Si el archivo está asociado a un periodo, también se removerá de ese periodo.
    
    Args:
        file_id: ID del archivo a eliminar (obtenido de upload-pdf)
        
    Returns:
        DeleteUploadResponse confirmando la eliminación
        
    Notas:
        - Solo se pueden eliminar archivos que aún no han sido procesados
        - Si el archivo está en proceso, se recomienda esperar a que termine antes de eliminarlo
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    upload_manager = get_upload_manager()
    
    # Verificar que el archivo existe
    if not upload_manager.file_exists(file_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"Archivo con file_id '{file_id}' no encontrado"}
        )
    
    # Obtener metadata antes de eliminar (para información de respuesta y periodo)
    metadata = upload_manager.get_uploaded_metadata(file_id)
    filename = None
    periodo_id = None
    request_id = None
    
    if metadata:
        filename = metadata.get("filename")
        periodo_id = metadata.get("metadata", {}).get("periodo_id")
        request_id = metadata.get("request_id")  # Si ya fue procesado, tiene request_id
        
        # Verificar si está procesado
        is_processed = metadata.get("processed", False)
        if is_processed:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail={
                    "error": "No se puede eliminar un archivo que ya ha sido procesado",
                    "file_id": file_id,
                    "request_id": request_id,
                    "message": "Los archivos procesados no pueden ser eliminados por seguridad"
                }
            )
    
    # Si tiene periodo_id y request_id (fue procesado y asociado), remover del periodo
    if periodo_id and request_id:
        try:
            periodo_manager = get_periodo_manager()
            periodo_manager.remove_archivo_from_periodo(periodo_id, request_id)
            logger.info(f"Archivo {file_id} removido del periodo {periodo_id}")
        except Exception as e:
            logger.warning(f"Error removiendo archivo {file_id} del periodo {periodo_id}: {e}")
    
    # Eliminar archivo y metadata
    success = upload_manager.delete_uploaded_pdf(file_id)
    
    if not success:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"error": "Error eliminando archivo. Algunos archivos pueden no haberse eliminado correctamente"}
        )
    
    logger.info(f"Archivo eliminado: {file_id} ({filename})")
    
    return DeleteUploadResponse(
        success=True,
        message=f"Archivo '{filename or file_id}' eliminado correctamente",
        file_id=file_id,
        filename=filename
    )


@app.post("/api/v1/process-pdf", response_model=ProcessPDFResponse, tags=["Processing"])
async def process_pdf(
    request: Request,
    file_id: str = Form(..., description="ID del archivo subido (obtener de /api/v1/uploaded-files)"),
    save_files: bool = Form(default=True, description="Guardar archivos en disco"),
    output_folder: str = Form(default="api", description="Subcarpeta de salida"),
    periodo_id: Optional[str] = Form(default=None, description="ID del periodo para asociar el archivo procesado")
):
    """
    Procesa un PDF que fue previamente subido con upload-pdf (ASÍNCRONO).
    
    Primero debes subir el PDF con POST /api/v1/upload-pdf y obtener el file_id.
    Luego usa ese file_id aquí para procesarlo.
    
    Este endpoint ahora es ASÍNCRONO: responde inmediatamente con un request_id.
    Usa GET /api/v1/process-status/{request_id} para consultar el estado del procesamiento.
    
    Para ver la lista de archivos disponibles, usa GET /api/v1/uploaded-files.
    
    Args:
        file_id: ID del archivo subido (obligatorio, obtener de upload-pdf)
        save_files: Si guardar archivos en disco (default: True)
        output_folder: Subcarpeta dentro de output/ (default: "api")
        periodo_id: ID del periodo para asociar el archivo procesado (opcional)
        
    Returns:
        ProcessPDFResponse con request_id y status inicial (queued/processing)
    """
    # Usar file_id como request_id
    request_id = file_id
    
    upload_manager = get_upload_manager()
    
    # Validar que el file_id existe
    if not upload_manager.file_exists(file_id):
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"Archivo con file_id '{file_id}' no encontrado. Usa GET /api/v1/uploaded-files para ver archivos disponibles."}
        )
    
    # Obtener PDF y metadata
    pdf_path = upload_manager.get_uploaded_pdf_path(file_id)
    metadata = upload_manager.get_uploaded_metadata(file_id)
    
    if not pdf_path or not metadata:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"Error obteniendo archivo o metadata para file_id '{file_id}'"}
        )
    
    # Obtener datos del PDF subido
    pdf_filename = metadata["filename"]
    email = metadata["metadata"]["email"]
    year = metadata["metadata"]["year"]
    normalized_month = metadata["metadata"]["month"]
    
    # Obtener periodo_id de metadata si no se proporcionó explícitamente
    # (puede haber sido especificado en upload-pdf)
    periodo_id_to_use = periodo_id
    if not periodo_id_to_use:
        periodo_id_to_use = metadata.get("metadata", {}).get("periodo_id")
    
    # Variable para usar en el scope de los lotes
    periodo_id_for_batches = periodo_id_to_use
    
    logger.info(f"[{request_id}] Agregando PDF a cola de procesamiento - file_id: {file_id}")
    logger.info(f"[{request_id}] Email: {email}, Año: {year}, Mes: {normalized_month}")
    if periodo_id_to_use:
        logger.info(f"[{request_id}] Periodo asociado: {periodo_id_to_use}")
    
    # Contar páginas del PDF para determinar si necesita procesamiento por lotes
    from .dependencies import get_file_manager
    file_manager = get_file_manager()
    from src.core.pdf_processor import PDFProcessor
    pdf_processor = PDFProcessor()
    
    total_pages = 0
    if pdf_processor.open_pdf(str(pdf_path)):
        total_pages = pdf_processor.get_page_count()
        pdf_processor.close()
    
    logger.info(f"[{request_id}] PDF tiene {total_pages} páginas")
    
    # Crear job(s) y agregarlo(s) a la cola
    worker_manager = get_worker_manager()
    BATCH_THRESHOLD = 200  # Umbral de páginas para procesamiento por lotes
    max_workers = worker_manager.max_workers
    
    if total_pages > BATCH_THRESHOLD:
        # Dividir en lotes
        logger.info(f"[{request_id}] PDF grande detectado ({total_pages} páginas). Dividiendo en lotes...")
        
        # Calcular tamaño de lote (dividir entre workers disponibles)
        pages_per_batch = (total_pages + max_workers - 1) // max_workers  # Redondeo hacia arriba
        
        batch_jobs = []
        for batch_num in range(max_workers):
            start_page = batch_num * pages_per_batch + 1  # 1-indexed
            end_page = min((batch_num + 1) * pages_per_batch, total_pages)  # 1-indexed
            
            if start_page > total_pages:
                break  # No más lotes necesarios
            
            # Crear request_id único para cada lote
            batch_request_id = f"{request_id}_batch_{batch_num + 1}"
            
            batch_job = ProcessingJob(
                request_id=batch_request_id,
                file_id=file_id,
                pdf_path=Path(pdf_path),
                metadata={
                    "filename": pdf_filename,
                    "email": email,
                    "year": year,
                    "month": normalized_month
                },
                save_files=save_files,
                output_folder=f"api/{truncate_request_id_for_folder(request_id)}",  # Carpeta específica por request_id maestro (truncado a 30 chars)
                periodo_id=periodo_id_to_use,
                start_page=start_page,
                end_page=end_page,
                batch_id=request_id,  # Mismo batch_id para todos los lotes
                is_batch_job=True
            )
            
            batch_jobs.append(batch_job)
            worker_manager.add_job(batch_job)
            logger.info(f"[{request_id}] Lote {batch_num + 1} creado: páginas {start_page}-{end_page}")
        
        # Crear job maestro que coordinará la consolidación
        master_job = ProcessingJob(
            request_id=request_id,
            file_id=file_id,
            pdf_path=Path(pdf_path),
            metadata={
                "filename": pdf_filename,
                "email": email,
                "year": year,
                "month": normalized_month
            },
            save_files=save_files,
            output_folder=f"api/{truncate_request_id_for_folder(request_id)}",  # Carpeta específica por request_id (truncado a 30 chars)
            periodo_id=periodo_id_to_use,
            batch_id=request_id,
            is_batch_job=False  # Este es el job maestro
        )
        
        # Agregar job maestro (pero no a la cola, solo para tracking)
        with worker_manager.jobs_lock:
            worker_manager.jobs[request_id] = master_job
            master_job.status = "queued"
            master_job.message = f"Esperando {len(batch_jobs)} lotes..."
        
        job = master_job
        initial_status = "queued"
    else:
        # PDF normal, procesar normalmente
        job = ProcessingJob(
            request_id=request_id,
            file_id=file_id,
            pdf_path=Path(pdf_path),
            metadata={
                "filename": pdf_filename,
                "email": email,
                "year": year,
                "month": normalized_month
            },
            save_files=save_files,
            output_folder=f"api/{truncate_request_id_for_folder(request_id)}",  # Carpeta específica por request_id (truncado a 30 chars)
            periodo_id=periodo_id_to_use
        )
        
        # Agregar job a la cola
        worker_manager.add_job(job)
        
        # Determinar estado inicial (puede ser "queued" o "processing" si hay workers disponibles)
        initial_status = "queued"
        if worker_manager.get_active_jobs_count() < worker_manager.max_workers:
            initial_status = "processing"
            job.status = "processing"
            job.message = "Iniciando procesamiento..."
    
    # Retornar respuesta inmediata
    return ProcessPDFResponse(
        success=True,
        request_id=request_id,
        status=initial_status,
        message="PDF agregado a cola de procesamiento. Usa GET /api/v1/process-status/{request_id} para consultar el estado."
    )


@app.post("/api/v1/process-batch", response_model=BatchProcessResponse, tags=["Processing"])
async def process_batch(request: BatchProcessRequest):
    """
    Procesa múltiples PDFs en paralelo (hasta 3 workers simultáneos).
    
    Este endpoint permite procesar varios archivos a la vez. Los archivos se agregan
    a la cola de procesamiento y se procesan hasta 3 en paralelo automáticamente.
    El resto espera en cola hasta que haya un worker disponible.
    
    Primero debes subir los PDFs con POST /api/v1/upload-pdf y obtener los file_ids.
    Luego usa este endpoint con la lista de file_ids para procesarlos todos.
    
    Este endpoint es ASÍNCRONO: responde inmediatamente con los request_ids.
    Usa GET /api/v1/process-status/{request_id} para consultar el estado de cada procesamiento.
    
    Args:
        request: BatchProcessRequest con lista de file_ids y opciones de procesamiento
        
    Returns:
        BatchProcessResponse con lista de jobs creados (cada uno con su request_id)
    """
    upload_manager = get_upload_manager()
    worker_manager = get_worker_manager()
    
    jobs_creados = []
    errores = []
    
    logger.info(f"Procesando batch de {len(request.file_ids)} archivos")
    
    # Procesar cada file_id
    for file_id in request.file_ids:
        try:
            # Validar que el file_id existe
            if not upload_manager.file_exists(file_id):
                errores.append({
                    "file_id": file_id,
                    "error": f"Archivo con file_id '{file_id}' no encontrado"
                })
                continue
            
            # Obtener PDF y metadata
            pdf_path = upload_manager.get_uploaded_pdf_path(file_id)
            metadata = upload_manager.get_uploaded_metadata(file_id)
            
            if not pdf_path or not metadata:
                errores.append({
                    "file_id": file_id,
                    "error": f"Error obteniendo archivo o metadata para file_id '{file_id}'"
                })
                continue
            
            # Obtener datos del PDF subido
            pdf_filename = metadata["filename"]
            email = metadata["metadata"]["email"]
            year = metadata["metadata"]["year"]
            normalized_month = metadata["metadata"]["month"]
            
            # Usar file_id como request_id
            request_id = file_id
            
            logger.info(f"[{request_id}] Agregando PDF a cola de procesamiento batch - file_id: {file_id}")
            logger.info(f"[{request_id}] Email: {email}, Año: {year}, Mes: {normalized_month}")
            
            # Crear job
            job = ProcessingJob(
                request_id=request_id,
                file_id=file_id,
                pdf_path=Path(pdf_path),
                metadata={
                    "filename": pdf_filename,
            "email": email,
            "year": year,
                    "month": normalized_month
                },
                save_files=request.save_files,
                output_folder=f"api/{truncate_request_id_for_folder(request_id)}",  # Carpeta específica por request_id (truncado a 30 chars)
                periodo_id=request.periodo_id
            )
            
            # Agregar job a la cola
            worker_manager.add_job(job)
            
            # Determinar estado inicial
            initial_status = "queued"
            initial_message = "Esperando en cola de procesamiento..."
            
            # Si hay workers disponibles, el job puede empezar inmediatamente
            if worker_manager.get_active_jobs_count() < worker_manager.max_workers:
                initial_status = "processing"
                initial_message = "Iniciando procesamiento..."
                job.status = "processing"
                job.message = "Iniciando procesamiento..."
            
            # Agregar a la lista de jobs creados
            jobs_creados.append(BatchJobInfo(
                file_id=file_id,
                request_id=request_id,
                status=initial_status,
                message=initial_message
            ))
            
        except Exception as e:
            logger.error(f"Error procesando file_id '{file_id}' en batch: {e}")
            errores.append({
                "file_id": file_id,
                "error": f"Error inesperado: {str(e)}"
            })
    
    logger.info(f"Batch completado: {len(jobs_creados)} jobs creados, {len(errores)} errores")
    
    return BatchProcessResponse(
        success=True,
        total=len(request.file_ids),
        procesados=len(jobs_creados),
        jobs=jobs_creados,
        errores=errores if errores else None
    )


@app.get("/api/v1/process-status/{request_id}", response_model=ProcessStatusResponse, tags=["Processing"])
async def get_process_status(request_id: str):
    """
    Consulta el estado de un procesamiento de PDF.
    
    Args:
        request_id: ID del request de procesamiento (obtenido de POST /api/v1/process-pdf)
        
    Returns:
        ProcessStatusResponse con estado actual del procesamiento
    """
    worker_manager = get_worker_manager()
    job = worker_manager.get_job_status(request_id)
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"Request ID '{request_id}' no encontrado. Verifica que el ID sea correcto."}
        )
    
    return ProcessStatusResponse(
        success=True,
        request_id=job.request_id,
        status=job.status,
        progress=job.progress,
        message=job.message,
        pages_processed=job.pages_processed,
        processing_time=job.processing_time,
        download_url=job.download_url,
        excel_download_url=job.excel_download_url,
        error=job.error
    )


async def _stream_job_status(request_id: str) -> AsyncGenerator[str, None]:
    """
    Generator para streaming de estado de un job individual.
    
    Args:
        request_id: ID del job a monitorear
        
    Yields:
        Eventos SSE en formato text/event-stream
    """
    worker_manager = get_worker_manager()
    last_status = None
    last_progress = -1
    timeout_seconds = 18000000  # 300000 minutos máximo
    start_time = time.time()
    poll_interval = 0.5  # Polling cada 0.5 segundos
    
    # Verificar que el job existe
    job = worker_manager.get_job_status(request_id)
    if not job:
        error_data = {
            "error": f"Request ID '{request_id}' no encontrado",
            "request_id": request_id
        }
        yield f"data: {json.dumps(error_data)}\n\n"
        return
    
    # Enviar estado inicial
    initial_data = {
        "request_id": job.request_id,
        "status": job.status,
        "progress": job.progress,
        "message": job.message,
        "pages_processed": job.pages_processed,
        "processing_time": job.processing_time,
        "download_url": job.download_url,
        "excel_download_url": job.excel_download_url,
        "error": job.error
    }
    yield f"data: {json.dumps(initial_data)}\n\n"
    last_status = job.status
    last_progress = job.progress
    
    # Monitorear cambios hasta que termine
    while True:
        # Verificar timeout
        if time.time() - start_time > timeout_seconds:
            timeout_data = {
                "request_id": request_id,
                "status": "timeout",
                "error": "Timeout: conexión SSE cerrada después de 30 minutos",
                "message": "El procesamiento puede continuar, pero la conexión SSE ha expirado"
            }
            yield f"data: {json.dumps(timeout_data)}\n\n"
            break
        
        # Obtener estado actual
        job = worker_manager.get_job_status(request_id)
        if not job:
            # Job eliminado o no encontrado
            error_data = {
                "request_id": request_id,
                "status": "not_found",
                "error": "Job no encontrado (puede haber sido eliminado)"
            }
            yield f"data: {json.dumps(error_data)}\n\n"
            break
        
        # Verificar si hay cambios
        status_changed = job.status != last_status
        progress_changed = job.progress != last_progress
        
        # Enviar actualización si hay cambios o cada 2 segundos (heartbeat)
        if status_changed or progress_changed or (time.time() - start_time) % 2 < poll_interval:
            update_data = {
                "request_id": job.request_id,
                "status": job.status,
                "progress": job.progress,
                "message": job.message,
                "pages_processed": job.pages_processed,
                "processing_time": job.processing_time,
                "download_url": job.download_url,
                "excel_download_url": job.excel_download_url,
                "error": job.error
            }
            yield f"data: {json.dumps(update_data)}\n\n"
            last_status = job.status
            last_progress = job.progress
        
        # Si el job terminó, cerrar conexión
        if job.status in ["completed", "failed"]:
            # Enviar estado final y cerrar
            final_data = {
                "request_id": job.request_id,
                "status": job.status,
                "progress": job.progress,
                "message": job.message,
                "pages_processed": job.pages_processed,
                "processing_time": job.processing_time,
                "download_url": job.download_url,
                "excel_download_url": job.excel_download_url,
                "error": job.error,
                "finished": True
            }
            yield f"data: {json.dumps(final_data)}\n\n"
            break
        
        await asyncio.sleep(poll_interval)


@app.get("/api/v1/process-status-stream/{request_id}", tags=["Processing"])
async def stream_process_status(request_id: str):
    """
    Stream de estado de procesamiento usando Server-Sent Events (SSE).
    
    Este endpoint mantiene una conexión abierta y envía actualizaciones en tiempo real
    del estado del procesamiento. Ideal para mostrar progreso en el frontend sin polling.
    
    El stream se cierra automáticamente cuando el job termina (completed/failed)
    o después de 30 minutos de inactividad.
    
    Args:
        request_id: ID del request de procesamiento (obtenido de POST /api/v1/process-pdf)
        
    Returns:
        StreamingResponse con eventos SSE (text/event-stream)
        
    Ejemplo de uso en frontend:
        const eventSource = new EventSource('/api/v1/process-status-stream/abc123');
        eventSource.onmessage = (event) => {
            const data = JSON.parse(event.data);
            console.log('Estado:', data.status, 'Progreso:', data.progress);
        };
    """
    return StreamingResponse(
        _stream_job_status(request_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no"  # Deshabilitar buffering en nginx
        }
    )


def _get_archivo_info_from_json(request_id: str) -> Optional[Dict[str, Any]]:
    """
    Obtiene información completa de un archivo desde los JSONs estructurados.
    Similar a lo que se hace en get_periodo_detail.
    
    Args:
        request_id: ID del request
        
    Returns:
        Diccionario con información del archivo (PeriodoArchivoInfo) o None
    """
    try:
        file_manager = get_file_manager()
        base_output = file_manager.get_output_folder() or "./output"
        
        # Para batch jobs, extraer el request_id maestro (antes de _batch_)
        request_id_to_search = request_id
        if "_batch_" in request_id:
            request_id_to_search = request_id.split("_batch_")[0]
        
        # Truncar request_id para buscar en la carpeta (mismo truncamiento que al crear)
        request_id_folder = truncate_request_id_for_folder(request_id_to_search)
        
        # Buscar en la carpeta específica por request_id
        structured_folder = Path(base_output) / "api" / request_id_folder / "structured"
        
        if not structured_folder.exists():
            return None
        
        # Buscar JSONs estructurados en esta carpeta específica
        # Todos los JSONs en esta carpeta pertenecen a este request_id (o sus batches)
        json_files = list(structured_folder.glob("*_structured.json"))
        if not json_files:
            return None
        
        # Usar el primer JSON encontrado (todos tienen la misma metadata)
        json_file = json_files[0]
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                json_data = json.load(f)
            
            metadata = json_data.get("metadata", {})
            # Verificar que el request_id coincide (puede ser el maestro o un batch)
            json_request_id = metadata.get("request_id", "")
            if json_request_id != request_id and not json_request_id.startswith(request_id + "_batch_"):
                # Si no coincide, puede ser que estemos buscando un batch pero encontramos el maestro
                # o viceversa, en ese caso continuar con el primer JSON encontrado
                pass
            
            # Extraer información del archivo (igual que en get_periodo_detail)
            # Las tablas ahora están en el nivel raíz, no en additional_data
            mresumen = json_data.get("mresumen", [])
            mcomprobante = json_data.get("mcomprobante", [])
            
            job_no = None
            source_ref = None
            entered_curr = None
            entered_amount = None
            total_usd = None
            fecha_valoracion = None
            
            # Intentar extraer de mresumen o mcomprobante
            if mresumen:
                first_item = mresumen[0]
                if isinstance(first_item, dict):
                    job_no = first_item.get("tjobno") or first_item.get("job_no")
                    source_ref = first_item.get("source_reference")
                    entered_curr = first_item.get("tDivisa")
                    entered_amount = first_item.get("nMonto")
                    if entered_curr and entered_curr.upper() == "USD" and entered_amount:
                        total_usd = float(entered_amount)
                    fecha_valoracion = first_item.get("fecha_valoracion")
            
            if mcomprobante and not job_no:
                first_item = mcomprobante[0]
                if isinstance(first_item, dict):
                    job_no = first_item.get("_stamp_name") or first_item.get("tSequentialNumber")
                    source_ref = first_item.get("tNumero")
            
            # Buscar total_usd en mresumen si no se encontró
            if not total_usd and mresumen:
                for item in mresumen:
                    if isinstance(item, dict):
                        monto = item.get("nMonto")
                        divisa = item.get("tDivisa")
                        if monto and divisa and divisa.upper() == "USD":
                            total_usd = float(monto)
                            break
            
            return {
                "archivo_id": request_id[:8],
                "request_id": request_id,
                "filename": metadata.get("filename", "unknown"),
                "estado": "procesado",
                "progress": 100,
                "status": "completed",
                "job_no": job_no,
                "type": metadata.get("document_type"),
                "source_reference": source_ref,
                "source_ref_id": source_ref,
                "entered_curr": entered_curr,
                "entered_amount": entered_amount,
                "total_usd": total_usd,
                "fecha_valoracion": fecha_valoracion,
                "processed_at": metadata.get("processed_at")
            }
        except Exception:
            return None
        
        return None
    except Exception:
        return None


async def _stream_periodo_status(periodo_id: str) -> AsyncGenerator[str, None]:
    """
    Generator para streaming de estado de todos los jobs de un periodo.
    
    Versión simplificada:
    - Durante procesamiento: solo envía progreso en %
    - Cuando finaliza: envía información completa del archivo (como PeriodoArchivoInfo)
    
    Args:
        periodo_id: ID del periodo a monitorear
        
    Yields:
        Eventos SSE en formato text/event-stream con estado consolidado
    """
    worker_manager = get_worker_manager()
    periodo_manager = get_periodo_manager()
    last_jobs_state = {}
    last_periodo_estado = None  # Para detectar cambios en el estado del periodo
    last_total_archivos = 0  # Para detectar cuando se sube un archivo nuevo
    timeout_seconds = 18000000  # 300000 minutos máximo
    start_time = time.time()
    poll_interval = 0.5  # Polling cada 0.5 segundos para actualizaciones más rápidas
    
    # Verificar que el periodo existe
    periodo_data = periodo_manager.get_periodo(periodo_id)
    if not periodo_data:
        error_data = {
            "error": f"Periodo '{periodo_id}' no encontrado",
            "periodo_id": periodo_id
        }
        error_event = f"data: {json.dumps(error_data)}\n\n"
        yield error_event
        await asyncio.sleep(0)
        return
    
    # Obtener request_ids asociados al periodo (jobs completados)
    request_ids_completados = set(periodo_manager.get_archivos_from_periodo(periodo_id))
    
    # Obtener jobs activos del worker_manager (queued/processing) con este periodo_id
    jobs_activos = worker_manager.get_jobs_by_periodo_id(periodo_id)
    request_ids_activos = {job.request_id for job in jobs_activos}
    
    # Combinar ambos: jobs completados + jobs activos
    request_ids = list(request_ids_completados | request_ids_activos)
    
    if not request_ids:
        # No hay jobs aún, pero el periodo existe
        initial_data = {
            "periodo_id": periodo_id,
            "total_jobs": 0,
            "completed": 0,
            "processing": 0,
            "queued": 0,
            "failed": 0,
            "jobs": [],
            "message": "No hay archivos procesados en este periodo aún"
        }
        no_jobs_event = f"data: {json.dumps(initial_data)}\n\n"
        yield no_jobs_event
        await asyncio.sleep(0)
        # Esperar a que se agreguen jobs (timeout más corto si no hay jobs)
        await asyncio.sleep(2)
        # Reintentar después de esperar
        request_ids_completados = set(periodo_manager.get_archivos_from_periodo(periodo_id))
        jobs_activos = worker_manager.get_jobs_by_periodo_id(periodo_id)
        request_ids_activos = {job.request_id for job in jobs_activos}
        request_ids = list(request_ids_completados | request_ids_activos)
        if not request_ids:
            return
    
    # Función helper para obtener estado consolidado (simplificado)
    def get_consolidated_status():
        jobs_info = []
        status_counts = {"completed": 0, "processing": 0, "queued": 0, "failed": 0}
        
        # Obtener todos los request_ids (completados + activos)
        request_ids_completados = set(periodo_manager.get_archivos_from_periodo(periodo_id))
        jobs_activos = worker_manager.get_jobs_by_periodo_id(periodo_id)
        request_ids_activos = {job.request_id for job in jobs_activos}
        all_request_ids = list(request_ids_completados | request_ids_activos)
        
        for req_id in all_request_ids:
            job = worker_manager.get_job_status(req_id)
            if job:
                # Job activo en worker_manager
                if job.status in ["queued", "processing"]:
                    # Durante procesamiento: solo progreso
                    job_info = {
                        "request_id": job.request_id,
                        "progress": job.progress,  # Solo progreso en %
                        "status": job.status
                    }
                elif job.status == "completed":
                    # Job completado: obtener info completa del archivo
                    archivo_info = _get_archivo_info_from_json(req_id)
                    if archivo_info:
                        job_info = archivo_info
                        job_info["progress"] = 100
                    else:
                        # Fallback si no se encuentra la info aún
                        job_info = {
                            "request_id": job.request_id,
                            "progress": 100,
                            "status": "completed",
                            "estado": "procesado"
                        }
                else:
                    # Failed
                    job_info = {
                        "request_id": job.request_id,
                        "progress": job.progress,
                        "status": job.status,
                        "error": job.error
                    }
                
                jobs_info.append(job_info)
                if job.status in status_counts:
                    status_counts[job.status] += 1
            else:
                # Job no encontrado en worker_manager
                # Verificar si está en archivos_asociados (completado)
                req_id_str = str(req_id)
                request_ids_completados_str = {str(rid) for rid in request_ids_completados}
                
                if req_id_str in request_ids_completados_str:
                    # Está en archivos_asociados = completado
                    # Obtener información completa del archivo
                    archivo_info = _get_archivo_info_from_json(req_id)
                    if archivo_info:
                        jobs_info.append(archivo_info)
                    else:
                        # Fallback
                        jobs_info.append({
                            "request_id": req_id,
                            "progress": 100,
                            "status": "completed",
                            "estado": "procesado",
                            "filename": "unknown"
                        })
                    status_counts["completed"] += 1
                else:
                    # No está ni en worker_manager ni en archivos_asociados
                    jobs_info.append({
                        "request_id": req_id,
                        "status": "not_found",
                        "progress": 0
                    })
        
        # Calcular estado del periodo dinámicamente (misma lógica que get_periodo_detail)
        upload_manager = get_upload_manager()
        uploaded_files = upload_manager.list_uploaded_files(processed=None)
        
        # Contar archivos del periodo
        archivos_periodo = []
        for uf in uploaded_files:
            file_metadata = uf.get("metadata", {})
            if file_metadata.get("periodo_id") == periodo_id:
                archivos_periodo.append(uf)
        
        # Contar estados de archivos
        archivos_procesados_count = sum(1 for a in archivos_periodo if a.get("processed", False))
        archivos_pendientes_count = sum(1 for a in archivos_periodo if not a.get("processed", False))
        total_archivos = len(archivos_periodo)
        
        # Verificar jobs activos
        jobs_activos_periodo = worker_manager.get_jobs_by_periodo_id(periodo_id)
        tiene_jobs_activos = any(job.status in ["queued", "processing"] for job in jobs_activos_periodo)
        
        # Verificar si hay archivos "subiendo" (subidos pero sin job creado aún)
        # Un archivo está "subiendo" si:
        # - Está subido (en uploaded_files)
        # - No tiene job activo asociado
        # - No está procesado
        archivos_subiendo = 0
        file_ids_subidos = {uf.get("file_id") for uf in uploaded_files 
                           if uf.get("metadata", {}).get("periodo_id") == periodo_id}
        file_ids_con_job = {job.file_id for job in jobs_activos_periodo}
        
        for file_id in file_ids_subidos:
            # Si el archivo está subido pero no tiene job, está "subiendo"
            if file_id not in file_ids_con_job:
                metadata_file = upload_manager.get_uploaded_metadata(file_id)
                if metadata_file and not metadata_file.get("processed", False):
                    archivos_subiendo += 1
        
        # Verificar primero si el periodo está "cerrado" en la base de datos
        estado_guardado = periodo_data.get("estado", "")
        if estado_guardado == "cerrado":
            periodo_estado = "cerrado"
        else:
            # Calcular estado del periodo según prioridad (misma lógica que get_periodo_detail)
            periodo_estado = "vacio"
            if total_archivos == 0:
                periodo_estado = "vacio"
            elif archivos_subiendo > 0:
                # Hay archivos recién subidos (estado "subiendo")
                periodo_estado = "subiendo"
            elif tiene_jobs_activos or status_counts["processing"] > 0 or status_counts["queued"] > 0:
                periodo_estado = "procesando"
            elif archivos_pendientes_count > 0:
                periodo_estado = "pendiente"
            elif archivos_procesados_count == total_archivos and archivos_procesados_count > 0:
                periodo_estado = "procesado"
            else:
                periodo_estado = "pendiente" if total_archivos > 0 else "vacio"
        
        return {
            "periodo_id": periodo_id,
            "total_jobs": len(all_request_ids),
            "completed": status_counts["completed"],
            "processing": status_counts["processing"],
            "queued": status_counts["queued"],
            "failed": status_counts["failed"],
            "jobs": jobs_info,
            "periodo_estado": periodo_estado,  # Estado del periodo: vacio/pendiente/procesado/procesando
            "total_archivos": total_archivos,
            "archivos_pendientes": archivos_pendientes_count,
            "archivos_procesados": archivos_procesados_count
        }
    
    # Enviar estado inicial
    initial_state = get_consolidated_status()
    initial_event = f"data: {json.dumps(initial_state)}\n\n"
    yield initial_event
    # Ceder control al event loop para forzar envío inmediato
    await asyncio.sleep(0)
    last_jobs_state = {job["request_id"]: (job.get("status") or job.get("estado", "unknown"), job.get("progress", 0)) for job in initial_state["jobs"]}
    
    # Monitorear cambios
    while True:
        # Verificar timeout
        if time.time() - start_time > timeout_seconds:
            timeout_data = {
                "periodo_id": periodo_id,
                "status": "timeout",
                "error": "Timeout: conexión SSE cerrada después de 30 minutos",
                "message": "El procesamiento puede continuar, pero la conexión SSE ha expirado"
            }
            timeout_event = f"data: {json.dumps(timeout_data)}\n\n"
            yield timeout_event
            await asyncio.sleep(0)
            break
        
        # Actualizar lista de request_ids (pueden agregarse nuevos jobs)
        # Combinar jobs completados + jobs activos
        current_request_ids_completados = set(periodo_manager.get_archivos_from_periodo(periodo_id))
        current_jobs_activos = worker_manager.get_jobs_by_periodo_id(periodo_id)
        current_request_ids_activos = {job.request_id for job in current_jobs_activos}
        current_request_ids = list(current_request_ids_completados | current_request_ids_activos)
        
        if set(current_request_ids) != set(request_ids):
            request_ids = current_request_ids
        
        # Obtener estado actual
        current_state = get_consolidated_status()
        
        # Verificar si hay cambios (esto debe hacerse PRIMERO para enviar eventos durante procesamiento)
        has_changes = False
        
        # Verificar cambios en el estado del periodo (vacio -> subiendo -> pendiente -> procesando -> procesado)
        current_periodo_estado = current_state.get("periodo_estado", "vacio")
        if current_periodo_estado != last_periodo_estado:
            has_changes = True
            last_periodo_estado = current_periodo_estado
        
        # Verificar cambios en total de archivos (indica que se subió un archivo nuevo)
        current_total_archivos = current_state.get("total_archivos", 0)
        if current_total_archivos != last_total_archivos:
            has_changes = True
            last_total_archivos = current_total_archivos
        
        # Verificar cambios en jobs
        for job in current_state["jobs"]:
            req_id = job["request_id"]
            # Obtener status (puede estar en "status" o "estado")
            current_status = job.get("status") or job.get("estado", "unknown")
            current_progress = job.get("progress", 0)
            
            if req_id not in last_jobs_state:
                has_changes = True
                break
            
            last_status, last_progress = last_jobs_state[req_id]
            
            # Detectar cambios de estado o progreso
            if current_status != last_status or current_progress != last_progress:
                has_changes = True
                break
            
            # Si cambió de processing/queued a completed, forzar actualización inmediata
            if last_status in ["queued", "processing"] and current_status in ["completed", "procesado"]:
                has_changes = True
                break
        
        # También verificar si cambió el total de jobs
        if len(current_state["jobs"]) != len(last_jobs_state):
            has_changes = True
        
        # Enviar actualización si hay cambios o cada 1 segundo (heartbeat más frecuente para detectar cambios rápidos)
        # IMPORTANTE: Esto envía eventos DURANTE el procesamiento y cuando cambia el estado del periodo
        elapsed_time = time.time() - start_time
        heartbeat_trigger = elapsed_time % 1 < poll_interval  # Heartbeat cada 1 segundo
        
        if has_changes or heartbeat_trigger:
            event_data = f"data: {json.dumps(current_state)}\n\n"
            yield event_data
            # Ceder control al event loop para forzar envío inmediato (evitar buffering)
            await asyncio.sleep(0)
            # Actualizar estado guardado (usar status o estado según esté disponible)
            last_jobs_state = {}
            for job in current_state["jobs"]:
                req_id = job["request_id"]
                status = job.get("status") or job.get("estado", "unknown")
                progress = job.get("progress", 0)
                last_jobs_state[req_id] = (status, progress)
        
        # Verificar si todos los jobs terminaron DESPUÉS de enviar la actualización
        # Esto asegura que se envíen todos los eventos de progreso antes del evento final
        all_finished = False
        if current_state["total_jobs"] > 0:
            # Normalizar status de cada job para verificación
            job_statuses = []
            for job in current_state["jobs"]:
                # Obtener status (puede estar en "status" o "estado")
                job_status = job.get("status") or job.get("estado", "unknown")
                # Normalizar: "procesado" -> "completed" para consistencia
                if job_status == "procesado":
                    job_status = "completed"
                job_statuses.append(job_status)
            
            # Verificar si todos terminaron (completed, failed, o not_found)
            all_finished = all(
                status in ["completed", "failed", "not_found"]
                for status in job_statuses
            )
        
        # Si todos terminaron, enviar evento final y cerrar conexión
        if all_finished:
            final_state = current_state.copy()
            final_state["finished"] = True
            final_state["message"] = f"Todos los jobs del periodo han terminado: {current_state['completed']} completados, {current_state['failed']} fallidos"
            # Asegurar que el estado del periodo sea "completed" en el objeto
            final_state["periodo_status"] = "completed"
            final_event = f"data: {json.dumps(final_state)}\n\n"
            yield final_event
            # Ceder control al event loop para forzar envío inmediato
            await asyncio.sleep(0)
            break
        
        await asyncio.sleep(poll_interval)


@app.post("/api/v1/periodos/{periodo_id}/process-all", response_model=BatchProcessResponse, tags=["Periodos"])
async def process_all_periodo_files(
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Procesa automáticamente todos los archivos pendientes de un periodo.
    
    Este endpoint detecta automáticamente los archivos subidos que:
    - Coinciden con el año y mes del periodo
    - No están procesados aún
    - Tienen email autorizado
    - No están ya asociados al periodo
    
    Los archivos se procesan en batch (hasta 3 workers simultáneos).
    
    Args:
        periodo_id: ID del periodo (ej: "2025-11-onshore")
        
    Returns:
        BatchProcessResponse con lista de jobs creados (cada uno con su request_id)
    """
    try:
        periodo_manager = get_periodo_manager()
        upload_manager = get_upload_manager()
        worker_manager = get_worker_manager()
        
        # 1. Obtener información del periodo
        periodo_data = periodo_manager.get_periodo(periodo_id)
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo '{periodo_id}' no encontrado"
            )
        
        # Extraer año y mes del periodo
        # Formato periodo: "MM/AAAA" (ej: "11/2025")
        periodo_str = periodo_data.get("periodo", "")
        if "/" not in periodo_str:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Formato de periodo inválido: '{periodo_str}'. Debe ser 'MM/AAAA'"
            )
        
        mes_str, anio_str = periodo_str.split("/")
        try:
            periodo_year = int(anio_str)
            periodo_month_num = int(mes_str)
        except ValueError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"No se pudo parsear año/mes del periodo: '{periodo_str}'"
            )
        
        # Normalizar mes a nombre (ej: "11" -> "Noviembre")
        periodo_month_normalized = _normalize_month(mes_str)
        
        # Obtener archivos ya asociados al periodo (para evitar duplicados)
        archivos_asociados = periodo_manager.get_archivos_from_periodo(periodo_id)
        archivos_asociados_set = set(archivos_asociados)
        
        # 2. Buscar archivos subidos que coincidan con el periodo
        all_uploaded_files = upload_manager.list_uploaded_files(processed=False)
        
        # Filtrar archivos que:
        # - Coincidan con año y mes del periodo
        # - No estén procesados
        # - Tengan email autorizado
        # - No estén ya asociados al periodo
        pending_file_ids = []
        errores_validacion = []
        
        for file_data in all_uploaded_files:
            file_id = file_data.get("file_id")
            if not file_id:
                continue
            
            metadata = file_data.get("metadata", {})
            file_year = metadata.get("year")
            file_month = metadata.get("month")
            file_email = metadata.get("email")
            
            # Verificar año y mes
            if file_year != periodo_year:
                continue
            
            # Normalizar mes del archivo para comparar
            file_month_normalized = _normalize_month(str(file_month)) if file_month else None
            if file_month_normalized != periodo_month_normalized:
                continue
            
            # Verificar que no esté procesado
            if file_data.get("processed", False):
                continue
            
            # Verificar que no esté ya asociado al periodo
            # (buscamos por request_id en archivos_asociados)
            request_id_existente = file_data.get("request_id")
            if request_id_existente and request_id_existente in archivos_asociados_set:
                continue
            
            # Archivo válido para procesar
            pending_file_ids.append(file_id)
        
        if not pending_file_ids:
            # No hay archivos pendientes
            return BatchProcessResponse(
                success=True,
                total=0,
                procesados=0,
                jobs=[],
                errores=None,
                message=f"No se encontraron archivos pendientes para el periodo {periodo_id}"
            )
        
        # 3. Procesar archivos en batch
        logger.info(f"Procesando {len(pending_file_ids)} archivos pendientes del periodo {periodo_id}")
        
        jobs_creados = []
        errores_procesamiento = []
        
        for file_id in pending_file_ids:
            try:
                # Obtener PDF y metadata
                pdf_path = upload_manager.get_uploaded_pdf_path(file_id)
                metadata = upload_manager.get_uploaded_metadata(file_id)
                
                if not pdf_path or not metadata:
                    errores_procesamiento.append({
                        "file_id": file_id,
                        "error": f"Error obteniendo archivo o metadata para file_id '{file_id}'"
                    })
                    continue
                
                # Obtener datos del PDF subido
                pdf_filename = metadata["filename"]
                email = metadata["metadata"]["email"]
                year = metadata["metadata"]["year"]
                normalized_month = metadata["metadata"]["month"]
                
                # Usar file_id como request_id
                request_id = file_id
                
                logger.info(f"[{request_id}] Agregando PDF a cola de procesamiento (periodo: {periodo_id}) - file_id: {file_id}")
                logger.info(f"[{request_id}] Email: {email}, Año: {year}, Mes: {normalized_month}")
                
                # Crear job con periodo_id
                job = ProcessingJob(
                    request_id=request_id,
                    file_id=file_id,
                    pdf_path=Path(pdf_path),
                    metadata={
                        "filename": pdf_filename,
                        "email": email,
                        "year": year,
                        "month": normalized_month
                    },
                    save_files=True,  # Siempre guardar archivos
                    output_folder=f"api/{truncate_request_id_for_folder(request_id)}",  # Carpeta específica por request_id (truncado a 30 chars)
                    periodo_id=periodo_id  # Asociar automáticamente al periodo
                )
                
                # Agregar job a la cola
                worker_manager.add_job(job)
                
                # Determinar estado inicial
                initial_status = "queued"
                initial_message = "Esperando en cola de procesamiento..."
                
                # Si hay workers disponibles, el job puede empezar inmediatamente
                if worker_manager.get_active_jobs_count() < worker_manager.max_workers:
                    initial_status = "processing"
                    initial_message = "Iniciando procesamiento..."
                    job.status = "processing"
                    job.message = "Iniciando procesamiento..."
                
                # Agregar a la lista de jobs creados
                jobs_creados.append(BatchJobInfo(
                    file_id=file_id,
                    request_id=request_id,
                    status=initial_status,
                    message=initial_message
                ))
                
            except Exception as e:
                logger.error(f"Error procesando file_id '{file_id}' en process-all: {e}")
                errores_procesamiento.append({
                    "file_id": file_id,
                    "error": f"Error inesperado: {str(e)}"
                })
        
        # Combinar errores de validación y procesamiento
        todos_errores = errores_validacion + errores_procesamiento
        
        logger.info(f"Process-all completado para periodo {periodo_id}: {len(jobs_creados)} jobs creados, {len(todos_errores)} errores")
        
        return BatchProcessResponse(
            success=True,
            total=len(pending_file_ids),
            procesados=len(jobs_creados),
            jobs=jobs_creados,
            errores=todos_errores if todos_errores else None
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error en process-all para periodo {periodo_id}: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error procesando archivos del periodo: {str(e)}"
        )


@app.post("/api/v1/periodos/{periodo_id}/process-selected", response_model=BatchProcessResponse, tags=["Periodos"])
async def process_selected_periodo_files(
    periodo_id: str, 
    request: ProcessSelectedRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Procesa solo los archivos seleccionados (por file_id/request_id) de un periodo.
    
    Este endpoint procesa únicamente los archivos cuyos file_ids se proporcionan en la lista.
    Verifica que cada archivo tenga el periodo_id correcto en su metadata.
    
    Los archivos se procesan en batch (hasta 3 workers simultáneos).
    
    Args:
        periodo_id: ID del periodo (ej: "2025-11-onshore")
        request: ProcessSelectedRequest con lista de file_ids a procesar
        
    Returns:
        BatchProcessResponse con lista de jobs creados (cada uno con su request_id)
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        periodo_manager = get_periodo_manager()
        upload_manager = get_upload_manager()
        worker_manager = get_worker_manager()
        
        logger.info(f"[process-selected] Procesando {len(request.file_ids)} archivos para periodo {periodo_id}")
        
        # 1. Validar que el periodo existe
        periodo_data = periodo_manager.get_periodo(periodo_id)
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo '{periodo_id}' no encontrado"
            )
        
        # 2. Obtener archivos asociados al periodo (para mostrar en errores)
        archivos_asociados = periodo_manager.get_archivos_from_periodo(periodo_id)
        archivos_info = []
        for request_id in archivos_asociados:
            try:
                metadata = upload_manager.get_uploaded_metadata(request_id)
                if metadata:
                    archivos_info.append({
                        "file_id": request_id,
                        "filename": metadata.get("filename", "unknown")
                    })
            except Exception:
                archivos_info.append({
                    "file_id": request_id,
                    "filename": "unknown"
                })
        
        # 3. Validar y procesar los file_ids proporcionados
        if not request.file_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="La lista de file_ids no puede estar vacía"
            )
        
        jobs_creados = []
        errores_procesamiento = []
        
        for file_id in request.file_ids:
            try:
                # Validar que el file_id existe en uploads
                if not upload_manager.file_exists(file_id):
                    errores_procesamiento.append({
                        "file_id": file_id,
                        "error": f"Archivo con file_id '{file_id}' no encontrado en uploads"
                    })
                    continue
                
                # Obtener metadata del archivo
                metadata = upload_manager.get_uploaded_metadata(file_id)
                if not metadata:
                    errores_procesamiento.append({
                        "file_id": file_id,
                        "error": f"No se pudo obtener metadata para file_id '{file_id}'"
                    })
                    continue
                
                # Verificar que el periodo_id en metadata coincide con el periodo solicitado
                file_metadata = metadata.get("metadata", {})
                file_periodo_id = file_metadata.get("periodo_id")
                
                if file_periodo_id != periodo_id:
                    errores_procesamiento.append({
                        "file_id": file_id,
                        "error": f"Archivo pertenece al periodo '{file_periodo_id}', no a '{periodo_id}'"
                    })
                    continue
                
                # Obtener PDF path
                pdf_path = upload_manager.get_uploaded_pdf_path(file_id)
                if not pdf_path:
                    errores_procesamiento.append({
                        "file_id": file_id,
                        "error": f"No se pudo obtener ruta del PDF para file_id '{file_id}'"
                    })
                    continue
                
                # Si el archivo no está en archivos_asociados, agregarlo al periodo
                if file_id not in archivos_asociados:
                    logger.info(f"[process-selected] Agregando archivo {file_id} al periodo {periodo_id}")
                    periodo_manager.add_archivo_to_periodo(periodo_id, file_id)
                
                # Obtener datos del PDF
                pdf_filename = metadata.get("filename", "unknown")
                email = file_metadata.get("email", "unknown")
                year = file_metadata.get("year")
                normalized_month = file_metadata.get("month")
                
                # Usar file_id como request_id (request_id = file_id)
                request_id = file_id
                
                logger.info(f"[process-selected] [{request_id}] Procesando archivo: {pdf_filename}")
                
                # Crear job con periodo_id
                job = ProcessingJob(
                    request_id=request_id,
                    file_id=file_id,
                    pdf_path=Path(pdf_path),
                    metadata={
                        "filename": pdf_filename,
                        "email": email,
                        "year": year,
                        "month": normalized_month
                    },
                    save_files=True,
                    output_folder=f"api/{truncate_request_id_for_folder(request_id)}",  # Carpeta específica por request_id (truncado a 30 chars)
                    periodo_id=periodo_id
                )
                
                # Agregar job a la cola
                worker_manager.add_job(job)
                
                # Determinar estado inicial
                initial_status = "queued"
                initial_message = "Esperando en cola de procesamiento..."
                
                if worker_manager.get_active_jobs_count() < worker_manager.max_workers:
                    initial_status = "processing"
                    initial_message = "Iniciando procesamiento..."
                    job.status = "processing"
                    job.message = "Iniciando procesamiento..."
                
                # Agregar a la lista de jobs creados
                jobs_creados.append(BatchJobInfo(
                    file_id=file_id,
                    request_id=request_id,
                    status=initial_status,
                    message=initial_message
                ))
                
            except Exception as e:
                logger.exception(f"[process-selected] Error procesando file_id '{file_id}': {e}")
                errores_procesamiento.append({
                    "file_id": file_id,
                    "error": f"Error inesperado: {str(e)}"
                })
        
        # Crear mensaje con información de archivos disponibles si hay errores
        message = None
        if not jobs_creados and archivos_info:
            archivos_list = "\n".join([f"  - {a['file_id']}: {a['filename']}" for a in archivos_info[:10]])
            message = f"No se pudo procesar ningún archivo. Archivos asociados al periodo ({len(archivos_info)} total):\n{archivos_list}"
            if len(archivos_info) > 10:
                message += f"\n  ... y {len(archivos_info) - 10} más"
        elif errores_procesamiento:
            message = f"Procesados {len(jobs_creados)} archivos. {len(errores_procesamiento)} archivos con errores."
        
        return BatchProcessResponse(
            success=len(jobs_creados) > 0,
            total=len(request.file_ids),
            procesados=len(jobs_creados),
            jobs=jobs_creados,
            errores=errores_procesamiento if errores_procesamiento else None,
            message=message
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"[process-selected] Error inesperado: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error procesando archivos seleccionados: {str(e)}"
        )


@app.get("/api/v1/periodos/{periodo_id}/process-status-stream", tags=["Periodos"])
async def stream_periodo_status(
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Stream de estado de procesamiento de todos los jobs de un periodo usando Server-Sent Events (SSE).
    
    Este endpoint mantiene una conexión abierta y envía actualizaciones en tiempo real
    del estado de todos los archivos procesados en un periodo. Ideal para monitorear
    procesamiento batch completo.
    
    El stream se cierra automáticamente cuando todos los jobs terminan o después de 30 minutos.
    
    Args:
        periodo_id: ID del periodo a monitorear (ej: "2025-11-onshore")
        
    Returns:
        StreamingResponse con eventos SSE (text/event-stream)
        
    Ejemplo de uso en frontend:
        const eventSource = new EventSource('/api/v1/periodos/2025-11-onshore/process-status-stream');
        eventSource.onmessage = (event) => {
            const data = JSON.parse(event.data);
            console.log('Total:', data.total_jobs, 'Completados:', data.completed);
            data.jobs.forEach(job => {
                console.log('Job:', job.request_id, 'Estado:', job.status);
            });
        };
    """
    return StreamingResponse(
        _stream_periodo_status(periodo_id),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache, no-transform",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",  # Deshabilitar buffering en nginx
            "Content-Type": "text/event-stream; charset=utf-8",
            "X-Content-Type-Options": "nosniff"
        }
    )


@app.get("/api/v1/uploaded-files", response_model=UploadedFilesResponse, tags=["Files"])
async def get_uploaded_files():
    """
    Obtiene lista de archivos subidos (pendientes y procesados).
    Solo muestra archivos de correos autorizados.
    
    Returns:
        Lista de archivos subidos (pendientes y procesados, solo correos autorizados)
    """
    upload_manager = get_upload_manager()
    # Mostrar tanto pendientes como procesados
    files = upload_manager.list_uploaded_files(processed=None)
    
    # Filtrar solo archivos de correos autorizados
    file_info_list = [
        UploadedFileInfo(
            file_id=f["file_id"],
            filename=f["filename"],
            uploaded_at=f["uploaded_at"],
            file_size_bytes=f["file_size_bytes"],
            metadata=f["metadata"],
            processed=f.get("processed", False)
        )
        for f in files
    ]
    
    return UploadedFilesResponse(
        success=True,
        total=len(file_info_list),
        files=file_info_list
    )


@app.get("/api/v1/process-status/{request_id}", response_model=ProcessStatusResponse, tags=["Processing"])
async def get_process_status(request_id: str):
    """
    Consulta el estado de un procesamiento de PDF.
    
    Args:
        request_id: ID del request de procesamiento (obtenido de POST /api/v1/process-pdf)
        
    Returns:
        ProcessStatusResponse con estado actual del procesamiento
    """
    worker_manager = get_worker_manager()
    job = worker_manager.get_job_status(request_id)
    
    if not job:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"Request ID '{request_id}' no encontrado. Verifica que el ID sea correcto."}
        )
    
    return ProcessStatusResponse(
            success=True,
        request_id=job.request_id,
        status=job.status,
        progress=job.progress,
        message=job.message,
        pages_processed=job.pages_processed,
        processing_time=job.processing_time,
        download_url=job.download_url,
        excel_download_url=job.excel_download_url,
        error=job.error
    )


@app.get("/api/v1/uploaded-files", response_model=UploadedFilesResponse, tags=["Files"])
async def get_uploaded_files():
    """
    Obtiene lista de archivos subidos (pendientes y procesados).
    Solo muestra archivos de correos autorizados.
    
    Returns:
        Lista de archivos subidos (pendientes y procesados, solo correos autorizados)
    """
    upload_manager = get_upload_manager()
    # Mostrar tanto pendientes como procesados
    files = upload_manager.list_uploaded_files(processed=None)
    
    # Filtrar solo archivos de correos autorizados
    file_info_list = [
        UploadedFileInfo(
            file_id=f["file_id"],
            filename=f["filename"],
            uploaded_at=f["uploaded_at"],
            file_size_bytes=f["file_size_bytes"],
            metadata=f["metadata"],
            processed=f.get("processed", False)
        )
        for f in files
    ]
    
    return UploadedFilesResponse(
        success=True,
        total=len(file_info_list),
        files=file_info_list
    )


@app.get("/api/v1/processed-files", response_model=ProcessedFilesResponse, tags=["Files"])
async def get_processed_files(limit: int = 10):
    """
    Obtiene lista de archivos que han sido procesados con sus enlaces de descarga.
    Solo muestra archivos de correos autorizados.
    
    Incluye:
    - Archivos procesados desde upload-pdf (con file_id)
    - Archivos procesados directamente (sin file_id)
    
    Args:
        limit: Número máximo de archivos a retornar (default: 10, los más recientes)
    
    Returns:
        Lista de archivos procesados con download_url (limitada a los más recientes, solo correos autorizados)
    """
    upload_manager = get_upload_manager()
    processed_tracker = get_processed_tracker()
    archive_manager = get_archive_manager()
    
    # Archivos procesados desde upload-pdf
    uploaded_files = upload_manager.list_uploaded_files(processed=True)
    
    # Archivos procesados directamente (sin upload previo)
    direct_files = processed_tracker.get_processed_files()
    
    # Buscar archivos Excel en la carpeta pública y mapearlos por request_id
    public_folder = archive_manager.public_folder
    excel_files = {}  # Mapa: request_id -> nombre_archivo_excel
    if public_folder.exists():
        # Buscar todos los archivos Excel
        for excel_file in public_folder.glob("*.xlsx"):
            # Formato: {pdf_name}_consolidado_{timestamp}_{request_id[:8]}.xlsx
            # Extraer request_id del nombre del archivo
            excel_name = excel_file.stem  # Sin extensión
            if "_consolidado_" in excel_name:
                # El formato es: ..._consolidado_{timestamp}_{request_id[:8]}
                # Extraer el request_id[:8] (último segmento antes de .xlsx)
                parts = excel_name.split("_")
                if len(parts) >= 2:
                    request_id_prefix = parts[-1]  # Último segmento es request_id[:8]
                    
                    # Buscar en los JSONs estructurados para encontrar el request_id completo
                    try:
                        file_manager = get_file_manager()
                        base_output = file_manager.get_output_folder() or "./output"
                        api_folder = Path(base_output) / "api"
                        
                        if api_folder.exists() and request_id_prefix:
                            # Buscar en todas las carpetas api/{request_id}/structured/
                            for request_folder in api_folder.iterdir():
                                if request_folder.is_dir():
                                    structured_folder = request_folder / "structured"
                                    if structured_folder.exists():
                                        # Buscar JSONs en esta carpeta específica
                                        for json_file in structured_folder.glob("*_structured.json"):
                                            try:
                                                with open(json_file, 'r', encoding='utf-8') as jf:
                                                    json_data = json.load(jf)
                                                metadata = json_data.get("metadata", {})
                                                json_request_id = metadata.get("request_id", "")
                                                if json_request_id and json_request_id[:8] == request_id_prefix:
                                                    # Encontramos el request_id completo
                                                    excel_files[json_request_id] = excel_file.name
                                                    break
                                            except Exception:
                                                continue
                    except Exception:
                        pass
    
    file_info_list = []
    
    # Agregar archivos de upload-pdf
    for f in uploaded_files:
        request_id = f.get("request_id")
        excel_url = None
        # Buscar Excel por request_id en el mapa
        if request_id and request_id in excel_files:
            excel_url = f"/public/{excel_files[request_id]}"
        # Si no se encontró en el mapa, intentar leer de la metadata
        elif not excel_url:
            excel_url = f.get("excel_download_url")
        
        file_info_list.append(
            UploadedFileInfo(
                file_id=f["file_id"],
                filename=f["filename"],
                uploaded_at=f["uploaded_at"],
                file_size_bytes=f["file_size_bytes"],
                metadata=f["metadata"],
                processed=True,
                processed_at=f.get("processed_at"),
                download_url=f.get("download_url"),
                request_id=request_id,
                excel_download_url=excel_url
            )
        )
    
    # Agregar archivos procesados directamente
    for f in direct_files:
        request_id = f.get("request_id")
        excel_url = None
        # Buscar Excel por request_id en el mapa
        if request_id and request_id in excel_files:
            excel_url = f"/public/{excel_files[request_id]}"
        # Si no se encontró en el mapa, intentar leer de la metadata
        elif not excel_url:
            excel_url = f.get("excel_download_url")
        
        file_info_list.append(
            UploadedFileInfo(
                file_id=f.get("request_id", "unknown"),  # Usar request_id como file_id
                filename=f["filename"],
                uploaded_at=f.get("processed_at", ""),  # Usar processed_at como uploaded_at
                file_size_bytes=0,  # No tenemos info de tamaño
                metadata=f["metadata"],
                processed=True,
                processed_at=f.get("processed_at"),
                download_url=f.get("download_url"),
                request_id=request_id,
                excel_download_url=excel_url
            )
        )
    
    # Ordenar por fecha de procesamiento (más reciente primero)
    file_info_list.sort(key=lambda x: x.processed_at or "", reverse=True)
    
    # Limitar a los N más recientes
    total_files = len(file_info_list)
    file_info_list = file_info_list[:limit]
    
    return ProcessedFilesResponse(
        success=True,
        total=total_files,  # Total real (antes de limitar)
        files=file_info_list  # Lista limitada
    )


# ===== Learning System Endpoints =====

@app.get("/api/v1/learning/errors", response_model=ErrorsResponse, tags=["Learning"])
async def get_learning_errors(limit: int = 100):
    """
    Obtiene lista de errores registrados por el sistema de aprendizaje.
    
    Args:
        limit: Número máximo de errores a retornar (default: 100)
        
    Returns:
        Lista de errores registrados
    """
    error_tracker, _, _ = get_learning_system()
    
    if error_tracker is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Sistema de aprendizaje no está activado. Activa 'learning.enabled: true' en config.json"
        )
    
    try:
        errors = error_tracker.get_recent_errors(limit=limit)
        
        error_list = [
            ErrorInfo(
                error_id=err.get("error_id", ""),
                timestamp=err.get("timestamp", ""),
                pdf_name=err.get("pdf_name", ""),
                page_number=err.get("page_number", 0),
                error_type=err.get("error_type", ""),
                error_message=err.get("error_message", ""),
                field_name=err.get("context", {}).get("field_name"),
                context=err.get("context", {})
            )
            for err in errors
        ]
        
        return ErrorsResponse(
            success=True,
            total=len(error_list),
            errors=error_list
        )
    
    except Exception as e:
        logger.exception(f"Error obteniendo errores: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener errores: {str(e)}"
        )


@app.get("/api/v1/learning/errors/summary", response_model=ErrorsSummaryResponse, tags=["Learning"])
async def get_learning_errors_summary():
    """
    Obtiene un resumen de errores registrados.
    
    Returns:
        Resumen con estadísticas de errores
    """
    error_tracker, _, _ = get_learning_system()
    
    if error_tracker is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Sistema de aprendizaje no está activado. Activa 'learning.enabled: true' en config.json"
        )
    
    try:
        summary = error_tracker.get_errors_summary()
        
        error_list = [
            ErrorInfo(
                error_id=err.get("error_id", ""),
                timestamp=err.get("timestamp", ""),
                pdf_name=err.get("pdf_name", ""),
                page_number=err.get("page_number", 0),
                error_type=err.get("error_type", ""),
                error_message=err.get("error_message", ""),
                field_name=err.get("context", {}).get("field_name"),
                context=err.get("context", {})
            )
            for err in summary.get("recent_errors", [])
        ]
        
        return ErrorsSummaryResponse(
            success=True,
            total_errors=summary.get("total_errors", 0),
            error_types=summary.get("error_types", {}),
            most_common_fields=summary.get("most_common_fields", {}),
            recent_errors=error_list
        )
    
    except Exception as e:
        logger.exception(f"Error obteniendo resumen de errores: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener resumen: {str(e)}"
        )


@app.post("/api/v1/learning/analyze", response_model=AnalysisResponse, tags=["Learning"])
async def analyze_learning_errors(limit: int = Query(default=20, ge=1, le=100)):
    """
    Analiza errores registrados usando Gemini para generar sugerencias de mejora.
    
    Este endpoint analiza los errores acumulados y genera sugerencias automáticas
    usando Gemini AI. Las sugerencias se guardan en learning/suggestions/.
    
    Args:
        limit: Número máximo de errores a analizar (default: 20, máximo recomendado: 50)
        
    Returns:
        Análisis de errores con patrones y sugerencias
    """
    error_tracker, _, learning_service = get_learning_system()
    
    if error_tracker is None or learning_service is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Sistema de aprendizaje no está activado. Activa 'learning.enabled: true' en config.json"
        )
    
    try:
        # Obtener errores
        errors = error_tracker.get_recent_errors(limit=limit)
        
        if not errors:
            return AnalysisResponse(
                success=True,
                analysis={"message": "No hay errores para analizar"},
                patterns=[],
                suggestions=[],
                analyzed_at=datetime.now().isoformat(),
                total_errors_analyzed=0
            )
        
        # Analizar con Gemini
        analysis = learning_service.analyze_with_gemini(errors, limit=limit)
        
        # También analizar sin Gemini para obtener patrones básicos
        basic_analysis = learning_service.analyze_errors(errors)
        
        return AnalysisResponse(
            success=True,
            analysis=analysis,
            patterns=basic_analysis.get("patterns", []),
            suggestions=basic_analysis.get("suggestions", []),
            analyzed_at=analysis.get("analyzed_at", datetime.now().isoformat()),
            total_errors_analyzed=len(errors)
        )
    
    except Exception as e:
        logger.exception(f"Error analizando errores: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al analizar errores: {str(e)}"
        )


@app.get("/api/v1/learning/prompts", response_model=PromptsResponse, tags=["Learning"])
async def get_learning_prompts(history_limit: int = 10):
    """
    Obtiene información sobre las versiones de prompts.
    
    Args:
        history_limit: Número máximo de versiones históricas a retornar (default: 10)
        
    Returns:
        Información del prompt actual e historial
    """
    _, prompt_manager, _ = get_learning_system()
    
    if prompt_manager is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Sistema de aprendizaje no está activado. Activa 'learning.enabled: true' en config.json"
        )
    
    try:
        current_version_info = prompt_manager.get_current_version_info()
        history = prompt_manager.get_history(limit=history_limit)
        
        current_version = PromptVersionInfo(
            version=current_version_info.get("version", 1),
            created_at=current_version_info.get("created_at", ""),
            description=current_version_info.get("description", ""),
            improvements=current_version_info.get("improvements", []),
            source=None
        )
        
        history_list = [
            PromptVersionInfo(
                version=h.get("version", 1),
                created_at=h.get("created_at", ""),
                description=h.get("description", ""),
                improvements=h.get("improvements", []),
                source=h.get("source")
            )
            for h in history
        ]
        
        return PromptsResponse(
            success=True,
            current_version=current_version,
            history=history_list
        )
    
    except Exception as e:
        logger.exception(f"Error obteniendo prompts: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener prompts: {str(e)}"
        )


@app.post("/api/v1/learning/prompts/apply", response_model=ApplyPromptResponse, tags=["Learning"])
async def apply_new_prompt(
    new_prompt: str = Form(..., description="Nuevo prompt a aplicar"),
    description: str = Form(..., description="Descripción de los cambios"),
    improvements: str = Form(default="", description="Lista de mejoras separadas por comas")
):
    """
    Aplica una nueva versión del prompt.
    
    Args:
        new_prompt: Nuevo prompt a aplicar
        description: Descripción de los cambios
        improvements: Lista de mejoras separadas por comas
        
    Returns:
        Confirmación de la nueva versión aplicada
    """
    _, prompt_manager, _ = get_learning_system()
    
    if prompt_manager is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Sistema de aprendizaje no está activado. Activa 'learning.enabled: true' en config.json"
        )
    
    try:
        improvements_list = [imp.strip() for imp in improvements.split(",") if imp.strip()] if improvements else []
        
        new_version = prompt_manager.save_new_version(
            new_prompt=new_prompt,
            description=description,
            improvements=improvements_list,
            source="api"
        )
        
        return ApplyPromptResponse(
            success=True,
            new_version=new_version,
            message=f"Prompt versión {new_version} aplicado exitosamente"
        )
    
    except Exception as e:
        logger.exception(f"Error aplicando nuevo prompt: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al aplicar prompt: {str(e)}"
        )


@app.get("/api/v1/learning/suggestions", tags=["Learning"])
async def get_learning_suggestions():
    """
    Obtiene sugerencias de mejora generadas por el sistema.
    
    Returns:
        Lista de archivos de sugerencias disponibles
    """
    _, _, learning_service = get_learning_system()
    
    if learning_service is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Sistema de aprendizaje no está activado. Activa 'learning.enabled: true' en config.json"
        )
    
    try:
        suggestions_folder = learning_service.suggestions_folder
        
        if not suggestions_folder.exists():
            return {
                "success": True,
                "total": 0,
                "suggestions": []
            }
        
        suggestion_files = sorted(suggestions_folder.glob("analysis_*.json"), reverse=True)
        
        suggestions_list = []
        for sf in suggestion_files[:10]:  # Últimas 10 sugerencias
            try:
                import json
                with open(sf, 'r', encoding='utf-8') as f:
                    suggestion_data = json.load(f)
                    suggestions_list.append({
                        "file": sf.name,
                        "analyzed_at": suggestion_data.get("analyzed_at", ""),
                        "errors_analyzed": suggestion_data.get("errors_analyzed", 0),
                        "analysis": suggestion_data.get("analysis", {})
                    })
            except Exception:
                continue
        
        return {
            "success": True,
            "total": len(suggestions_list),
            "suggestions": suggestions_list
        }
    
    except Exception as e:
        logger.exception(f"Error obteniendo sugerencias: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener sugerencias: {str(e)}"
        )


# Función movida a excel_generator.py para evitar problemas de imports
from .excel_generator import generate_excel_for_request as _generate_excel_for_request


@app.get("/api/v1/export-zip/{request_id}", tags=["Export"])
async def export_zip(
    request_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Redirige a la descarga del archivo ZIP para un request_id.
    
    Busca el zip_filename en la metadata y redirige a /public/{zip_filename}.
    
    Args:
        request_id: ID del procesamiento (obtener de la respuesta de /api/v1/process-pdf)
        
    Returns:
        Redirección a /public/{zip_filename} o error 404 si no se encuentra
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    from fastapi.responses import RedirectResponse
    
    upload_manager = get_upload_manager()
    processed_tracker = get_processed_tracker()
    archive_manager = get_archive_manager()
    
    # Buscar zip_filename en metadata
    zip_filename = None
    email_found = None
    
    # Buscar en archivos procesados desde upload-pdf
    uploaded_files = upload_manager.list_uploaded_files(processed=True)
    for f in uploaded_files:
        if f.get("request_id") == request_id:
            zip_filename = f.get("zip_filename")
            email_found = f.get("metadata", {}).get("email", "")
            break
    
    # Si no se encontró, buscar en archivos procesados directamente
    if not zip_filename:
        direct_files = processed_tracker.get_processed_files()
        for f in direct_files:
            if f.get("request_id") == request_id:
                zip_filename = f.get("zip_filename")
                email_found = f.get("metadata", {}).get("email", "")
                break
    
    # Si no se encontró el zip_filename
    if not zip_filename:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"ZIP no encontrado para request_id '{request_id}'. El archivo puede no haber sido procesado aún."}
        )
    
    # Verificar que el archivo existe físicamente
    zip_path = archive_manager.public_folder / zip_filename
    if not zip_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"Archivo ZIP '{zip_filename}' no existe físicamente para request_id '{request_id}'"}
        )
    
    # Redirigir a /public/{zip_filename}
    return RedirectResponse(url=f"/public/{zip_filename}", status_code=302)


@app.post("/api/v1/export-bulk", tags=["Export"])
async def export_bulk_files(request: BulkExportRequest):
    """
    Exporta múltiples archivos seleccionados mediante checkbox.
    
    Recibe un array de request_ids y genera un ZIP maestro que contiene:
    - Todos los ZIPs individuales de cada request_id seleccionado
    - Todos los Excels individuales de cada request_id seleccionado (si existen)
    
    Args:
        request: BulkExportRequest con array de request_ids
        
    Returns:
        Redirección a /public/{timestamp}_bulk_export.zip
    """
    from fastapi.responses import RedirectResponse
    import zipfile
    
    upload_manager = get_upload_manager()
    processed_tracker = get_processed_tracker()
    archive_manager = get_archive_manager()
    
    request_ids = request.request_ids
    
    if not request_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El array de request_ids no puede estar vacío"
        )
    
    # Buscar zip_filename y excel_filename para cada request_id
    archivos_para_exportar = []
    
    # Buscar en archivos procesados desde upload-pdf
    uploaded_files = upload_manager.list_uploaded_files(processed=True)
    for f in uploaded_files:
        request_id = f.get("request_id")
        if request_id and request_id in request_ids:
            zip_filename = f.get("zip_filename")
            excel_filename = f.get("excel_filename")
            filename = f.get("filename", "unknown")
            
            if zip_filename or excel_filename:
                archivos_para_exportar.append({
                    "request_id": request_id,
                    "filename": filename,
                    "zip_filename": zip_filename,
                    "excel_filename": excel_filename
                })
    
    # Buscar también en archivos procesados directamente (por si acaso)
    direct_files = processed_tracker.get_processed_files()
    for f in direct_files:
        request_id = f.get("request_id")
        if request_id and request_id in request_ids:
            # Verificar que no esté ya en la lista
            if not any(a["request_id"] == request_id for a in archivos_para_exportar):
                zip_filename = f.get("zip_filename")
                excel_filename = f.get("excel_filename")
                filename = f.get("filename", "unknown")
                
                if zip_filename or excel_filename:
                    archivos_para_exportar.append({
                        "request_id": request_id,
                        "filename": filename,
                        "zip_filename": zip_filename,
                        "excel_filename": excel_filename
                    })
    
    if not archivos_para_exportar:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontraron archivos ZIP/Excel para los request_ids proporcionados"
        )
    
    # Crear ZIP maestro con todos los archivos
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_maestro_filename = f"{timestamp}_bulk_export.zip"
    zip_maestro_path = archive_manager.public_folder / zip_maestro_filename
    
    archivos_agregados = 0
    archivos_no_encontrados = []
    
    try:
        with zipfile.ZipFile(zip_maestro_path, 'w', zipfile.ZIP_DEFLATED) as zip_maestro:
            for archivo_info in archivos_para_exportar:
                # Agregar ZIP si existe
                if archivo_info.get("zip_filename"):
                    zip_filename = archivo_info["zip_filename"]
                    zip_path = archive_manager.public_folder / zip_filename
                    
                    if zip_path.exists():
                        # Agregar al ZIP maestro (en la raíz, sin subcarpetas)
                        zip_maestro.write(zip_path, zip_filename)
                        archivos_agregados += 1
                        logger.info(f"[Bulk Export] Agregado ZIP: {zip_filename}")
                    else:
                        archivos_no_encontrados.append(f"ZIP: {zip_filename}")
                
                # Agregar Excel si existe
                if archivo_info.get("excel_filename"):
                    excel_filename = archivo_info["excel_filename"]
                    excel_path = archive_manager.public_folder / excel_filename
                    
                    if excel_path.exists():
                        # Agregar al ZIP maestro (en la raíz, sin subcarpetas)
                        zip_maestro.write(excel_path, excel_filename)
                        archivos_agregados += 1
                        logger.info(f"[Bulk Export] Agregado Excel: {excel_filename}")
                    else:
                        archivos_no_encontrados.append(f"Excel: {excel_filename}")
        
        if archivos_agregados == 0:
            # Si no se agregó ningún archivo, eliminar el ZIP vacío
            zip_maestro_path.unlink()
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No se encontraron archivos físicos para exportar. Archivos no encontrados: {', '.join(archivos_no_encontrados)}"
            )
        
        logger.info(f"[Bulk Export] ZIP maestro creado: {zip_maestro_filename} ({archivos_agregados} archivos)")
        
        if archivos_no_encontrados:
            logger.warning(f"[Bulk Export] Algunos archivos no se encontraron: {', '.join(archivos_no_encontrados)}")
        
        # Redirigir a la descarga del ZIP maestro
        return RedirectResponse(url=f"/public/{zip_maestro_filename}", status_code=302)
    
    except Exception as e:
        logger.exception(f"Error en bulk export: {e}")
        # Limpiar ZIP maestro si se creó parcialmente
        if zip_maestro_path.exists():
            try:
                zip_maestro_path.unlink()
            except Exception:
                pass
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creando exportación bulk: {str(e)}"
        )


@app.get("/api/v1/export-excel/{request_id}", tags=["Export"])
async def export_structured_to_excel(
    request_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Redirige a la descarga del Excel consolidado para un request_id.
    
    Busca el excel_filename en la metadata (igual que el ZIP) y redirige a /public/{excel_filename}.
    Si no existe, intenta generarlo.
    
    Args:
        request_id: ID del procesamiento (obtener de la respuesta de /api/v1/process-pdf)
        
    Returns:
        Redirección a /public/{excel_filename} o archivo Excel para descarga directa
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    from fastapi.responses import RedirectResponse
    
    upload_manager = get_upload_manager()
    processed_tracker = get_processed_tracker()
    archive_manager = get_archive_manager()
    
    # Buscar excel_filename en metadata (igual que se busca zip_filename para el ZIP)
    excel_filename = None
    email_found = None
    pdf_name = None
    
    # Buscar en archivos procesados desde upload-pdf
    uploaded_files = upload_manager.list_uploaded_files(processed=True)
    for f in uploaded_files:
        if f.get("request_id") == request_id:
            excel_filename = f.get("excel_filename")
            email_found = f.get("metadata", {}).get("email", "")
            filename = f.get("filename", "")
            if filename:
                pdf_name = Path(filename).stem
            break
    
    # Si no se encontró, buscar en archivos procesados directamente
    if not excel_filename:
        direct_files = processed_tracker.get_processed_files()
        for f in direct_files:
            if f.get("request_id") == request_id:
                excel_filename = f.get("excel_filename")
                email_found = f.get("metadata", {}).get("email", "")
                filename = f.get("filename", "")
                if filename:
                    pdf_name = Path(filename).stem
                break
    
    # Si se encontró el excel_filename, redirigir a /public/{excel_filename}
    if excel_filename:
        excel_path = archive_manager.public_folder / excel_filename
        if excel_path.exists():
            # Redirigir a /public/{excel_filename} (igual que el ZIP)
            return RedirectResponse(url=f"/public/{excel_filename}", status_code=302)
        else:
            # El Excel no existe físicamente, intentar generarlo
            logger.warning(f"[{request_id}] Excel en metadata no existe físicamente: {excel_filename}")
    
    # Si no se encontró excel_filename o no existe físicamente, intentar generarlo
    if not pdf_name:
        pdf_name = "unknown"
    
    # Generar Excel usando la función helper (generará Excel vacío si no hay JSONs)
    file_manager = get_file_manager()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename, excel_download_url = await _generate_excel_for_request(
        request_id=request_id,
        pdf_name=pdf_name,
        timestamp=timestamp,
        archive_manager=archive_manager,
        file_manager=file_manager
    )
    
    if not excel_filename:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail={"error": f"Error generando Excel para request_id '{request_id}'"}
        )
    
    # Redirigir a /public/{excel_filename}
    return RedirectResponse(url=f"/public/{excel_filename}", status_code=302)


@app.get("/public/{filename}", tags=["Public"])
async def serve_public_file(filename: str):
    """
    Endpoint para servir archivos públicos (zips y excels) para descarga.
    
    Args:
        filename: Nombre del archivo a descargar
        
    Returns:
        Archivo para descarga
    """
    from fastapi.responses import FileResponse
    
    archive_manager = get_archive_manager()
    file_path = archive_manager.public_folder / filename
    
    if not file_path.exists():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail={"error": f"Archivo no encontrado: {filename}"}
        )
    
    # Verificar que sea un archivo zip o excel
    if not (filename.lower().endswith('.zip') or filename.lower().endswith('.xlsx')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={"error": "Solo se permiten archivos .zip o .xlsx"}
        )
    
    # Buscar el correo asociado a este archivo (ZIP o Excel)
    upload_manager = get_upload_manager()
    processed_tracker = get_processed_tracker()
    file_manager = get_file_manager()
    
    email_found = None
    
    # Si es un Excel, extraer request_id del nombre del archivo
    # Formato: {pdf_name}_consolidado_{timestamp}_{request_id[:8]}.xlsx
    if filename.lower().endswith('.xlsx'):
        # Intentar extraer request_id del nombre
        parts = filename.replace('.xlsx', '').split('_')
        if len(parts) >= 3:
            # El request_id está en los últimos caracteres (8 caracteres)
            potential_request_id_short = parts[-1]
            # Buscar request_id en los JSONs estructurados (buscar en todas las carpetas api/{request_id}/structured/)
            base_output = file_manager.get_output_folder() or "./output"
            api_folder = Path(base_output) / "api"
            
            if api_folder.exists():
                # Buscar en todas las carpetas api/{request_id}/structured/
                for request_folder in api_folder.iterdir():
                    if request_folder.is_dir():
                        structured_folder = request_folder / "structured"
                        if structured_folder.exists():
                            all_json_files = structured_folder.glob("*_structured.json")
                            for json_file in all_json_files:
                                try:
                                    with open(json_file, 'r', encoding='utf-8') as f:
                                        json_data = json.load(f)
                                    metadata = json_data.get("metadata", {})
                                    request_id = metadata.get("request_id", "")
                                    if request_id and request_id[:8] == potential_request_id_short:
                                        email_found = metadata.get("email", "")
                                        break
                                except Exception:
                                    continue
                        if email_found:
                            break
    
    # Buscar en archivos procesados desde upload-pdf (para ZIPs y Excel)
    if not email_found:
        uploaded_files = upload_manager.list_uploaded_files(processed=True)
        for f in uploaded_files:
            # Verificar si es el ZIP o el Excel
            if (f.get("zip_filename") == filename or 
                f.get("excel_filename") == filename or 
                f.get("download_url", "").endswith(filename) or
                f.get("excel_download_url", "").endswith(filename)):
                email_found = f.get("metadata", {}).get("email", "")
                break
    
    # Si no se encontró, buscar en archivos procesados directamente
    if not email_found:
        direct_files = processed_tracker.get_processed_files()
        for f in direct_files:
            # Verificar si es el ZIP o el Excel
            if (f.get("zip_filename") == filename or 
                f.get("excel_filename") == filename or 
                f.get("download_url", "").endswith(filename) or
                f.get("excel_download_url", "").endswith(filename)):
                email_found = f.get("metadata", {}).get("email", "")
                break
    
    # Determinar media type según extensión
    if filename.lower().endswith('.xlsx'):
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        media_type = "application/zip"
    
    return FileResponse(
        path=str(file_path),
        filename=filename,
        media_type=media_type
    )


# ===== Dashboard Endpoints =====

def _load_dashboard_mock_data() -> Dict[str, Any]:
    """
    Carga los datos mockeados del dashboard desde JSON.
    
    Returns:
        Diccionario con todos los datos mockeados
    """
    try:
        config_path = Path(__file__).parent.parent.parent.parent / "config" / "dashboard_mock_data.json"
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        else:
            logger.warning(f"Archivo de mock data no encontrado: {config_path}. Usando valores por defecto.")
            return {}
    except Exception as e:
        logger.error(f"Error cargando dashboard mock data: {e}")
        return {}


@app.get("/api/v1/dashboard/stats", response_model=DashboardStatsResponse, tags=["Dashboard"])
async def get_dashboard_stats(
    fecha_inicio: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_fin: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    moneda: str = Query("USD", description="Moneda (USD, PEN, EUR)"),
    tipo_documento: Optional[str] = Query(None, description="Tipo de documento"),
    departamento: Optional[str] = Query(None, description="Departamento"),
    disciplina: Optional[str] = Query(None, description="Disciplina"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Obtiene estadísticas globales del dashboard.
    Requiere autenticación.
    
    Args:
        fecha_inicio: Fecha de inicio para filtrar
        fecha_fin: Fecha de fin para filtrar
        moneda: Moneda para los cálculos
        tipo_documento: Filtrar por tipo de documento
        departamento: Filtrar por departamento
        disciplina: Filtrar por disciplina
        
    Returns:
        Estadísticas globales (Monto Total, Total Horas)
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        # ============================================================
        # NOTA: Actualmente lee de JSONs estructurados (temporal)
        # Cuando tengas conexión a SQL Server, cambiar a leer de BD:
        # 
        # SELECT 
        #   SUM(nPrecioTotal) as monto_total,
        #   SUM(nTotalHoras) as total_horas
        # FROM MCOMPROBANTE c
        # LEFT JOIN MJORNADA j ON j.iMHoja = c.iMHoja
        # WHERE ... (aplicar filtros de fecha, moneda, etc.)
        # ============================================================
        
        # ============================================================
        # DATOS MOCKEADOS PARA PRUEBAS (TEMPORAL)
        # TODO: Reemplazar con lectura real de JSONs o SQL Server
        # ============================================================
        
        # Intentar leer de JSONs reales primero
        file_manager = get_file_manager()
        base_output = file_manager.get_output_folder() or "./output"
        api_folder = Path(base_output) / "api"
        
        monto_total = 0.0
        total_horas = 0.0
        has_real_data = False
        
        # TEMPORAL: Leer todos los JSONs estructurados desde todas las carpetas api/{request_id}/structured/
        # TODO: Migrar a SQL Server cuando tengas conexión a BD
        if api_folder.exists():
            # Buscar en todas las carpetas api/{request_id}/structured/
            for request_folder in api_folder.iterdir():
                if request_folder.is_dir():
                    structured_folder = request_folder / "structured"
                    if structured_folder.exists():
                        for json_file in structured_folder.glob("*_structured.json"):
                            try:
                                with open(json_file, 'r', encoding='utf-8') as f:
                                    json_data = json.load(f)
                                
                                has_real_data = True
                                
                                # Aplicar filtros aquí si es necesario
                                metadata = json_data.get("metadata", {})
                                
                                # Filtrar por fecha si se proporciona
                                if fecha_inicio or fecha_fin:
                                    processed_at = metadata.get("processed_at", "")
                                    if processed_at:
                                        # TODO: Implementar filtro de fechas
                                        # Cuando tengas BD: WHERE fEmision BETWEEN fecha_inicio AND fecha_fin
                                        pass
                                
                                # Extraer montos y horas (las tablas ahora están en el nivel raíz)
                                # TODO: Cuando tengas BD, esto vendrá de:
                                # - MCOMPROBANTE.nPrecioTotal
                                # - MJORNADA.nTotalHoras
                                
                                # Buscar en mcomprobante
                                comprobantes = json_data.get("mcomprobante", [])
                                for comp in comprobantes:
                                    if isinstance(comp, dict):
                                        precio_total = comp.get("nPrecioTotal", 0)
                                        if precio_total:
                                            monto_total += float(precio_total)
                                
                                # Buscar en mjornada
                                jornadas = json_data.get("mjornada", [])
                                for jornada in jornadas:
                                    if isinstance(jornada, dict):
                                        horas = jornada.get("nTotalHoras", 0)
                                        if horas:
                                            total_horas += float(horas)
                            except Exception as e:
                                logger.warning(f"Error leyendo {json_file}: {e}")
                    continue
        
        # Si no hay datos reales, usar datos mockeados para pruebas
        if not has_real_data:
            mock_data = _load_dashboard_mock_data()
            stats_mock = mock_data.get("stats", {})
            monto_total = stats_mock.get("monto_total_global", 2450000.75)
            total_horas = stats_mock.get("total_horas_global", 1875.5)
        
        return DashboardStatsResponse(
            success=True,
            monto_total_global=round(monto_total, 2),
            total_horas_global=round(total_horas, 2),
            currency=moneda.upper()  # Asegurar formato consistente (USD, PEN, EUR)
        )
    
    except Exception as e:
        logger.exception(f"Error obteniendo estadísticas: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener estadísticas: {str(e)}"
        )


@app.get("/api/v1/dashboard/analytics", response_model=DashboardAnalyticsResponse, tags=["Dashboard"])
async def get_dashboard_analytics(
    fecha_inicio: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_fin: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    moneda: str = Query("USD", description="Moneda (USD, PEN, EUR)"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Obtiene análisis Off-Shore y On-Shore.
    Requiere autenticación.
    
    Returns:
        Análisis con totales, distribución por departamento y top 5 disciplinas
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        # ============================================================
        # DATOS MOCKEADOS PARA PRUEBAS (TEMPORAL)
        # TODO: Reemplazar con lectura real de JSONs o SQL Server
        # ============================================================
        
        # Cargar datos mockeados desde JSON
        mock_data = _load_dashboard_mock_data()
        analytics_mock = mock_data.get("analytics", {})
        
        # Datos mockeados para Offshore
        offshore_mock = analytics_mock.get("offshore", {})
        offshore_dept_values = offshore_mock.get("departamentos", {
            "Engineering": 850000.00,
            "Procurement": 320000.00,
            "Construction": 180000.00,
            "Project Management": 95000.00,
            "Quality Control": 45000.00,
            "Health & Safety": 35000.00,
            "Environmental": 28000.00,
            "Logistics": 22000.00,
            "Other Services": 100000.50
        })
        
        # Calcular total de departamentos y convertir a porcentajes
        offshore_dept_total = sum(offshore_dept_values.values())
        offshore_departamentos = [
            DepartamentoItem(
                label=label,
                value=round((valor / offshore_dept_total) * 100, 2)
            )
            for label, valor in offshore_dept_values.items()
        ]
        # Ordenar de mayor a menor por porcentaje
        offshore_departamentos.sort(key=lambda x: x.value, reverse=True)
        
        # Valores absolutos de disciplinas
        offshore_disc_values = offshore_mock.get("disciplinas", {
            "Procurement": 1800.00,
            "Engineering": 1450.00,
            "Construction": 1200.00,
            "Project Management": 950.00,
            "Quality Control": 680.00
        })
        
        # Calcular total de disciplinas y convertir a porcentajes
        offshore_disc_total = sum(offshore_disc_values.values())
        offshore_disciplinas = [
            DisciplinaItem(
                label=label,
                value=round((valor / offshore_disc_total) * 100, 2)
            )
            for label, valor in offshore_disc_values.items()
        ]
        # Ordenar de mayor a menor por porcentaje
        offshore_disciplinas.sort(key=lambda x: x.value, reverse=True)
        
        offshore_item = AnalyticsItem(
            total_gasto=offshore_mock.get("total_gasto", 1450000.50),
            total_horas=offshore_mock.get("total_horas", 1125.25),
            total_disciplinas=offshore_mock.get("total_disciplinas", 12),
            distribucion_departamento=offshore_departamentos,
            top_5_disciplinas=offshore_disciplinas
        )
        
        # Datos mockeados para Onshore
        onshore_mock = analytics_mock.get("onshore", {})
        onshore_dept_values = onshore_mock.get("departamentos", {
            "Engineering": 450000.00,
            "Operations": 280000.00,
            "Maintenance": 150000.00,
            "Safety": 85000.00,
            "Environmental": 45000.00,
            "Human Resources": 35000.00,
            "Finance": 28000.00,
            "IT Services": 22000.00,
            "Other Services": 120000.25
        })
        
        # Calcular total de departamentos y convertir a porcentajes
        onshore_dept_total = sum(onshore_dept_values.values())
        onshore_departamentos = [
            DepartamentoItem(
                label=label,
                value=round((valor / onshore_dept_total) * 100, 2)
            )
            for label, valor in onshore_dept_values.items()
        ]
        # Ordenar de mayor a menor por porcentaje
        onshore_departamentos.sort(key=lambda x: x.value, reverse=True)
        
        # Valores absolutos de disciplinas
        onshore_disc_values = onshore_mock.get("disciplinas", {
            "Engineering": 1250.00,
            "Operations": 980.00,
            "Maintenance": 720.00,
            "Safety": 550.00,
            "Environmental": 420.00
        })
        
        # Calcular total de disciplinas y convertir a porcentajes
        onshore_disc_total = sum(onshore_disc_values.values())
        onshore_disciplinas = [
            DisciplinaItem(
                label=label,
                value=round((valor / onshore_disc_total) * 100, 2)
            )
            for label, valor in onshore_disc_values.items()
        ]
        # Ordenar de mayor a menor por porcentaje
        onshore_disciplinas.sort(key=lambda x: x.value, reverse=True)
        
        onshore_item = AnalyticsItem(
            total_gasto=onshore_mock.get("total_gasto", 1000000.25),
            total_horas=onshore_mock.get("total_horas", 750.25),
            total_disciplinas=onshore_mock.get("total_disciplinas", 10),
            distribucion_departamento=onshore_departamentos,
            top_5_disciplinas=onshore_disciplinas
        )
        
        return DashboardAnalyticsResponse(
            success=True,
            offshore=offshore_item,
            onshore=onshore_item
        )
    
    except Exception as e:
        logger.exception(f"Error obteniendo analytics: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener analytics: {str(e)}"
        )


@app.get("/api/v1/dashboard/rejected-concepts", response_model=RejectedConceptsResponse, tags=["Dashboard"])
async def get_rejected_concepts(
    fecha_inicio: Optional[str] = Query(None, description="Fecha inicio (YYYY-MM-DD)"),
    fecha_fin: Optional[str] = Query(None, description="Fecha fin (YYYY-MM-DD)"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Obtiene lista de conceptos rechazados.
    Requiere autenticación.
    
    Returns:
        Lista de conceptos rechazados con cantidad, monto y porcentaje
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    try:
        # ============================================================
        # DATOS MOCKEADOS PARA PRUEBAS (TEMPORAL)
        # TODO: Reemplazar con lectura real de JSONs o SQL Server
        # ============================================================
        
        # Cargar datos mockeados desde JSON
        mock_data = _load_dashboard_mock_data()
        rejected_concepts_data = mock_data.get("rejected_concepts", [
            {
                "concepto": "Materiales no especificados",
                "cantidad_total": 15,
                "monto_total": 45000.00,
                "porcentaje_total": 36.0
            },
            {
                "concepto": "Servicios sin factura",
                "cantidad_total": 8,
                "monto_total": 32000.00,
                "porcentaje_total": 25.6
            },
            {
                "concepto": "Conceptos duplicados",
                "cantidad_total": 12,
                "monto_total": 28000.00,
                "porcentaje_total": 22.4
            },
            {
                "concepto": "Documentación incompleta",
                "cantidad_total": 6,
                "monto_total": 15000.00,
                "porcentaje_total": 12.0
            },
            {
                "concepto": "Fechas fuera de rango",
                "cantidad_total": 4,
                "monto_total": 5000.00,
                "porcentaje_total": 4.0
            }
        ])
        
        concepts_mock = [
            RejectedConcept(**concept) for concept in rejected_concepts_data
        ]
        
        return RejectedConceptsResponse(
            success=True,
            total=len(concepts_mock),
            concepts=concepts_mock
        )
    
    except Exception as e:
        logger.exception(f"Error obteniendo conceptos rechazados: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener conceptos rechazados: {str(e)}"
        )


# ===== Periodos Endpoints =====

@app.post("/api/v1/periodos", response_model=PeriodoResponse, tags=["Periodos"])
@limiter.limit("10/minute")  # Máximo 10 requests por minuto por IP
async def create_periodo(
    request: Request, 
    periodo_data: CreatePeriodoRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Crea un nuevo periodo.
    
    Args:
        request: Request de FastAPI (para rate limiting)
        periodo_data: Datos del periodo a crear (periodo: "MM/AAAA", tipo: "onshore"|"offshore")
        
    Returns:
        Periodo creado
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        periodo_manager = get_periodo_manager()
        
        # Validación adicional: asegurar que periodo no esté vacío (ya validado por Pydantic, pero por seguridad)
        if not periodo_data.periodo or not periodo_data.periodo.strip():
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="El campo 'Periodo' es obligatorio y no puede estar vacío"
            )
        
        # Validar tipo (ya validado por Pydantic, pero por seguridad)
        if periodo_data.tipo.lower() not in ["onshore", "offshore"]:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Tipo debe ser 'onshore' o 'offshore'"
            )
        
        periodo_data_result = periodo_manager.create_periodo(periodo_data.periodo, periodo_data.tipo)
        
        # Función helper para formatear fechas
        def format_date(date_str: Optional[str]) -> Optional[str]:
            """Formatea fecha ISO a formato DD/MM/YYYY, HH:MM"""
            if not date_str:
                return None
            try:
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return dt.strftime("%d/%m/%Y, %H:%M")
            except Exception:
                return date_str  # Si falla, retornar original
        
        periodo_info = PeriodoInfo(
            periodo_id=periodo_data_result["periodo_id"],
            periodo=periodo_data_result["periodo"],
            tipo=periodo_data_result["tipo"],
            estado=periodo_data_result["estado"],
            registros=periodo_data_result["registros"],
            ultimo_procesamiento=format_date(periodo_data_result.get("ultimo_procesamiento")),
            created_at=format_date(periodo_data_result.get("created_at"))
        )
        
        return PeriodoResponse(success=True, periodo=periodo_info)
    
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )
    except Exception as e:
        logger.exception(f"Error creando periodo: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al crear periodo: {str(e)}"
        )


@app.get("/api/v1/periodos", response_model=PeriodosListResponse, tags=["Periodos"])
@limiter.limit("30/minute")  # Máximo 30 requests por minuto por IP
async def list_periodos(
    request: Request,
    tipo: Optional[str] = Query(None, description="Filtrar por tipo (onshore/offshore)"),
    estado: Optional[str] = Query(None, description="Filtrar por estado"),
    search: Optional[str] = Query(None, description="Buscar en periodo"),
    limit: int = Query(15, ge=1, le=100, description="Límite de resultados por página"),
    page: int = Query(1, ge=1, description="Número de página (empieza desde 1)"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Lista periodos con filtros opcionales y paginación.
    Requiere autenticación.
    
    Args:
        tipo: Filtrar por tipo (onshore/offshore)
        estado: Filtrar por estado
        search: Buscar en periodo
        limit: Límite de resultados por página (default: 15)
        page: Número de página, empieza desde 1 (default: 1)
    
    Returns:
        Lista de periodos con información de paginación
    """
    try:
        import math
        
        periodo_manager = get_periodo_manager()
        periodos_data = periodo_manager.list_periodos(tipo=tipo, estado=estado, search=search)
        
        # Calcular paginación (page empieza desde 1)
        total_periodos = len(periodos_data)
        offset = (page - 1) * limit
        paginas = math.ceil(total_periodos / limit) if total_periodos > 0 else 1
        
        # Aplicar paginación
        periodos_data = periodos_data[offset:offset + limit]
        
        # Función helper para formatear fechas
        def format_date(date_str: Optional[str]) -> Optional[str]:
            """Formatea fecha ISO a formato DD/MM/YYYY, HH:MM"""
            if not date_str:
                return None
            try:
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return dt.strftime("%d/%m/%Y, %H:%M")
            except Exception:
                return date_str  # Si falla, retornar original
        
        # Calcular estado dinámicamente para cada periodo
        upload_manager = get_upload_manager()
        uploaded_files = upload_manager.list_uploaded_files(processed=False)
        
        # Crear diccionario de periodo_id -> cantidad de archivos pendientes
        periodo_pendientes = {}
        for uploaded_file in uploaded_files:
            file_metadata = uploaded_file.get("metadata", {})
            file_periodo_id = file_metadata.get("periodo_id")
            if file_periodo_id:
                periodo_pendientes[file_periodo_id] = periodo_pendientes.get(file_periodo_id, 0) + 1
        
        # Obtener jobs activos por periodo para determinar estado "procesando"
        worker_manager = get_worker_manager()
        periodos_con_jobs_activos = set()
        for periodo_id_temp in [p["periodo_id"] for p in periodos_data]:
            jobs_activos_temp = worker_manager.get_jobs_by_periodo_id(periodo_id_temp)
            if any(job.status in ["queued", "processing"] for job in jobs_activos_temp):
                periodos_con_jobs_activos.add(periodo_id_temp)
        
        periodos = []
        for p in periodos_data:
            periodo_id = p["periodo_id"]
            # Verificar primero si el periodo está "cerrado" en la base de datos
            estado_guardado = p.get("estado", "")
            
            # Contar archivos procesados
            archivos_procesados = len(p.get("archivos_asociados", []))
            # Contar archivos pendientes
            archivos_pendientes = periodo_pendientes.get(periodo_id, 0)
            total_archivos = archivos_procesados + archivos_pendientes
            
            # Calcular registros dinámicamente: total de archivos (procesados + pendientes)
            # Esto muestra la cantidad real de archivos asociados al periodo
            registros_calculados = total_archivos
            
            # Si el periodo está "cerrado", usar "cerrado" directamente sin calcular
            if estado_guardado == "cerrado":
                estado_calculado = "cerrado"
            else:
                # Calcular estado según los 4 estados posibles
                estado_calculado = "vacio"
                if total_archivos == 0:
                    estado_calculado = "vacio"
                elif periodo_id in periodos_con_jobs_activos:
                    # Hay jobs activos (queued/processing)
                    estado_calculado = "procesando"
                elif archivos_procesados == total_archivos and archivos_procesados > 0:
                    # Todos los archivos están procesados
                    estado_calculado = "procesado"
                elif archivos_pendientes > 0:
                    # Hay archivos subidos pero no procesados
                    estado_calculado = "pendiente"
                else:
                    # Fallback
                    estado_calculado = "pendiente"
            
            periodos.append(
                PeriodoInfo(
                    periodo_id=periodo_id,
                    periodo=p["periodo"],
                    tipo=p["tipo"],
                    estado=estado_calculado,
                    registros=registros_calculados,  # Usar registros calculados dinámicamente
                    ultimo_procesamiento=format_date(p.get("ultimo_procesamiento")),
                    created_at=format_date(p.get("created_at"))
                )
            )
        
        return PeriodosListResponse(
            success=True,
            totalPeriodos=total_periodos,
            paginas=paginas,
            periodos=periodos
        )
    
    except Exception as e:
        logger.exception(f"Error listando periodos: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al listar periodos: {str(e)}"
        )


@app.get("/api/v1/periodos/{periodo_id}", response_model=PeriodoDetailResponse, tags=["Periodos"])
@limiter.limit("30/minute")  # Máximo 30 requests por minuto por IP
async def get_periodo_detail(
    request: Request, 
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Obtiene el detalle completo de un periodo incluyendo sus archivos.
    Requiere autenticación.
    
    Args:
        periodo_id: ID del periodo
        
    Returns:
        Detalle del periodo con lista de archivos
    """
    try:
        periodo_manager = get_periodo_manager()
        periodo_data = periodo_manager.get_periodo(periodo_id)
        
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo {periodo_id} no encontrado"
            )
        
        # Obtener archivos asociados
        request_ids = periodo_manager.get_archivos_from_periodo(periodo_id)
        archivos = []
        
        # Buscar información de cada archivo
        # Primero en JSONs estructurados, luego en processed_tracking.json
        file_manager = get_file_manager()
        base_output = file_manager.get_output_folder() or "./output"
        processed_tracker = get_processed_tracker()
        upload_manager = get_upload_manager()
        
        for request_id in request_ids:
            archivo_info = None
            
            # Obtener metadata del archivo subido para file_size_bytes y uploaded_at
            uploaded_file_metadata = None
            try:
                uploaded_file_metadata = upload_manager.get_uploaded_metadata(request_id)
            except Exception:
                pass
            
            # 1. Buscar en JSONs estructurados en la carpeta específica por request_id
            # Para batch jobs, extraer el request_id maestro (antes de _batch_)
            request_id_to_search = request_id
            if "_batch_" in request_id:
                request_id_to_search = request_id.split("_batch_")[0]
            
            # Truncar request_id para buscar en la carpeta (mismo truncamiento que al crear)
            request_id_folder = truncate_request_id_for_folder(request_id_to_search)
            
            structured_folder = Path(base_output) / "api" / request_id_folder / "structured"
            if structured_folder.exists():
                # Buscar todos los JSONs en esta carpeta específica
                json_files = list(structured_folder.glob("*_structured.json"))
                if json_files:
                    # Usar el primer JSON encontrado (todos tienen la misma metadata)
                    json_file = json_files[0]
                    try:
                        with open(json_file, 'r', encoding='utf-8') as f:
                            json_data = json.load(f)
                        metadata = json_data.get("metadata", {})
                        # Verificar que el request_id coincide (puede ser el maestro o un batch)
                        json_request_id = metadata.get("request_id", "")
                        if json_request_id == request_id or json_request_id.startswith(request_id + "_batch_"):
                            # Extraer información del archivo
                            # Las tablas ahora están en el nivel raíz, no en additional_data
                            # Extraer job_no, type, etc. del nivel raíz si están disponibles
                            mresumen = json_data.get("mresumen", [])
                            mcomprobante = json_data.get("mcomprobante", [])
                            
                            job_no = None
                            source_ref = None
                            entered_curr = None
                            entered_amount = None
                            total_usd = None
                            fecha_valoracion = None
                            
                            # Intentar extraer de mresumen o mcomprobante
                            if mresumen:
                                first_item = mresumen[0]
                                job_no = first_item.get("job_no")
                                source_ref = first_item.get("source_reference")
                                entered_curr = first_item.get("entered_curr")
                                entered_amount = first_item.get("entered_amount")
                                total_usd = first_item.get("total_usd")
                                fecha_valoracion = first_item.get("fecha_valoracion")
                            elif mcomprobante:
                                first_item = mcomprobante[0]
                                job_no = first_item.get("job_no")
                                source_ref = first_item.get("source_reference")
                                entered_curr = first_item.get("entered_curr")
                                entered_amount = first_item.get("entered_amount")
                                total_usd = first_item.get("total_usd")
                                fecha_valoracion = first_item.get("fecha_valoracion")
                            
                            # Obtener file_size_bytes y uploaded_at del metadata del archivo subido
                            file_size_bytes = None
                            uploaded_at = None
                            if uploaded_file_metadata:
                                file_size_bytes = uploaded_file_metadata.get("file_size_bytes")
                                uploaded_at = uploaded_file_metadata.get("uploaded_at")
                            
                            archivo_info = PeriodoArchivoInfo(
                                archivo_id=request_id[:8],
                                request_id=request_id,
                                filename=metadata.get("filename", "unknown"),
                                estado="procesado",
                                job_no=job_no,
                                type=metadata.get("document_type"),
                                source_reference=source_ref,
                                source_ref_id=source_ref,
                                entered_curr=entered_curr,
                                entered_amount=entered_amount,
                                total_usd=total_usd,
                                fecha_valoracion=fecha_valoracion,
                                processed_at=metadata.get("processed_at"),
                                file_size_bytes=file_size_bytes,
                                uploaded_at=uploaded_at
                            )
                            break
                    except Exception:
                        continue
            
            # 2. Si no se encontró en JSONs estructurados, buscar en processed_tracking.json
            if not archivo_info:
                try:
                    tracking_file = Path("./processed_tracking.json")
                    if tracking_file.exists():
                        with open(tracking_file, 'r', encoding='utf-8') as f:
                            tracking_data = json.load(f)
                        
                        if request_id in tracking_data:
                            file_data = tracking_data[request_id]
                            # Obtener file_size_bytes y uploaded_at del metadata del archivo subido
                            file_size_bytes = None
                            uploaded_at = None
                            if uploaded_file_metadata:
                                file_size_bytes = uploaded_file_metadata.get("file_size_bytes")
                                uploaded_at = uploaded_file_metadata.get("uploaded_at")
                            
                            archivo_info = PeriodoArchivoInfo(
                                archivo_id=request_id[:8],
                                request_id=request_id,
                                filename=file_data.get("filename", "unknown"),
                                estado="procesado",
                                job_no=None,
                                type=None,
                                source_reference=None,
                                source_ref_id=None,
                                entered_curr=None,
                                entered_amount=None,
                                total_usd=None,
                                fecha_valoracion=None,
                                processed_at=file_data.get("processed_at"),
                                file_size_bytes=file_size_bytes,
                                uploaded_at=uploaded_at
                            )
                except Exception:
                    pass
            
            if archivo_info:
                archivos.append(archivo_info)
        
        # Agregar archivos subidos (pendientes y procesados) que tengan este periodo_id en metadata
        upload_manager = get_upload_manager()
        # Mostrar tanto pendientes como procesados
        uploaded_files = upload_manager.list_uploaded_files(processed=None)
        
        # Obtener request_ids ya incluidos para evitar duplicados
        request_ids_incluidos = {archivo.request_id for archivo in archivos}
        
        for uploaded_file in uploaded_files:
            file_metadata = uploaded_file.get("metadata", {})
            file_periodo_id = file_metadata.get("periodo_id")
            
            # Incluir si el periodo_id coincide (tanto pendientes como procesados)
            if file_periodo_id == periodo_id:
                file_id = uploaded_file.get("file_id")
                filename = uploaded_file.get("filename", "unknown")
                is_processed = uploaded_file.get("processed", False)
                
                # Verificar que no esté ya en la lista
                if file_id not in request_ids_incluidos:
                    # Determinar estado: si está procesado, usar "procesado", sino "pendiente"
                    estado_archivo = "procesado" if is_processed else "pendiente"
                    
                    archivo_info = PeriodoArchivoInfo(
                        archivo_id=file_id[:8] if len(file_id) >= 8 else file_id,
                        request_id=file_id,  # Usar file_id como identificador
                        filename=filename,
                        estado=estado_archivo,
                        job_no=None,
                        type=None,
                        source_reference=None,
                        source_ref_id=None,
                        entered_curr=None,
                        entered_amount=None,
                        total_usd=None,
                        fecha_valoracion=None,
                        processed_at=uploaded_file.get("processed_at") if is_processed else None,
                        file_size_bytes=uploaded_file.get("file_size_bytes"),
                        uploaded_at=uploaded_file.get("uploaded_at")
                    )
                    archivos.append(archivo_info)
                    request_ids_incluidos.add(file_id)
        
        # Calcular estado del periodo basado en los 4 estados posibles
        # 1. "procesando" - si hay jobs activos (queued/processing)
        # 2. "procesado" - si todos los archivos están completados
        # 3. "pendiente" - si hay archivos subidos pero no procesados
        # 4. "subiendo" - si hay archivos recién subidos (menos de 5 segundos desde upload)
        
        worker_manager = get_worker_manager()
        jobs_activos = worker_manager.get_jobs_by_periodo_id(periodo_id)
        
        # Contar estados de archivos
        archivos_procesados = sum(1 for a in archivos if a.estado == "procesado")
        archivos_pendientes = sum(1 for a in archivos if a.estado == "pendiente")
        total_archivos = len(archivos)
        
        # Verificar si hay archivos "subiendo" (subidos pero sin job creado aún)
        # Un archivo está "subiendo" si:
        # - Está subido (en uploaded_files)
        # - No tiene job activo asociado
        # - No está procesado
        archivos_subiendo = 0
        file_ids_subidos = {uf.get("file_id") for uf in uploaded_files 
                           if uf.get("metadata", {}).get("periodo_id") == periodo_id}
        file_ids_con_job = {job.file_id for job in jobs_activos}
        
        for file_id in file_ids_subidos:
            # Si el archivo está subido pero no tiene job, está "subiendo"
            if file_id not in file_ids_con_job:
                metadata_file = upload_manager.get_uploaded_metadata(file_id)
                if metadata_file and not metadata_file.get("processed", False):
                    archivos_subiendo += 1
        
        # Verificar si hay jobs activos (queued o processing)
        tiene_jobs_activos = any(
            job.status in ["queued", "processing"] 
            for job in jobs_activos
        )
        
        # Verificar primero si el periodo está "cerrado" en la base de datos
        # Si está cerrado, no calcular dinámicamente, usar "cerrado" directamente
        estado_guardado = periodo_data.get("estado", "")
        
        # CRÍTICO: SIEMPRE respetar el estado "cerrado" - no calcular dinámicamente
        # El estado "cerrado" tiene prioridad absoluta sobre cualquier cálculo dinámico
        if estado_guardado and estado_guardado.lower() == "cerrado":
            # Si el periodo está cerrado, usar "cerrado" directamente sin calcular
            estado_calculado = "cerrado"
            logger.info(f"Periodo {periodo_id} está CERRADO - usando estado 'cerrado' sin calcular")
        else:
            # Calcular estado según prioridad (según requerimientos del usuario):
            # 1. Si no hay archivos → "vacio"
            # 2. Si hay al menos uno pendiente → "pendiente"
            # 3. Si todos están procesados → "procesado"
            estado_calculado = "pendiente"
            if total_archivos == 0:
                # No hay archivos
                estado_calculado = "vacio"
            elif archivos_subiendo > 0:
                # Hay archivos recién subidos (estado "subiendo")
                estado_calculado = "subiendo"
            elif tiene_jobs_activos:
                # Hay jobs en cola o procesando
                estado_calculado = "procesando"
            elif archivos_pendientes > 0:
                # Hay al menos un archivo pendiente
                estado_calculado = "pendiente"
            elif archivos_procesados == total_archivos and archivos_procesados > 0:
                # Todos los archivos están procesados
                estado_calculado = "procesado"
            else:
                # Fallback: si hay archivos pero no se pudo determinar el estado, asumir pendiente
                estado_calculado = "pendiente" if total_archivos > 0 else "vacio"
        
        # Función helper para formatear fechas
        def format_date(date_str: Optional[str]) -> Optional[str]:
            """Formatea fecha ISO a formato DD/MM/YYYY, HH:MM"""
            if not date_str:
                return None
            try:
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return dt.strftime("%d/%m/%Y, %H:%M")
            except Exception:
                return date_str  # Si falla, retornar original
        
        # Calcular registros dinámicamente: total de archivos (procesados + pendientes)
        registros_calculados = total_archivos
        
        periodo_info = PeriodoInfo(
            periodo_id=periodo_data["periodo_id"],
            periodo=periodo_data["periodo"],
            tipo=periodo_data["tipo"],
            estado=estado_calculado,  # Usar estado calculado o "cerrado" si está bloqueado
            registros=registros_calculados,  # Usar registros calculados dinámicamente
            ultimo_procesamiento=format_date(periodo_data.get("ultimo_procesamiento")),
            created_at=format_date(periodo_data.get("created_at"))
        )
        
        return PeriodoDetailResponse(
            success=True,
            periodo=periodo_info,
            archivos=archivos,
            total_archivos=len(archivos)
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error obteniendo detalle de periodo: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener detalle: {str(e)}"
        )


@app.put("/api/v1/periodos/{periodo_id}", response_model=PeriodoResponse, tags=["Periodos"])
@limiter.limit("20/minute")  # Máximo 20 requests por minuto por IP
async def update_periodo(
    request: Request, 
    periodo_id: str, 
    updates: Dict[str, Any],
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Actualiza un periodo.
    
    Args:
        periodo_id: ID del periodo
        updates: Campos a actualizar
        
    Returns:
        Periodo actualizado
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        periodo_manager = get_periodo_manager()
        periodo_data = periodo_manager.update_periodo(periodo_id, updates)
        
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo {periodo_id} no encontrado"
            )
        
        # Función helper para formatear fechas
        def format_date(date_str: Optional[str]) -> Optional[str]:
            """Formatea fecha ISO a formato DD/MM/YYYY, HH:MM"""
            if not date_str:
                return None
            try:
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return dt.strftime("%d/%m/%Y, %H:%M")
            except Exception:
                return date_str  # Si falla, retornar original
        
        periodo_info = PeriodoInfo(
            periodo_id=periodo_data["periodo_id"],
            periodo=periodo_data["periodo"],
            tipo=periodo_data["tipo"],
            estado=periodo_data["estado"],
            registros=periodo_data["registros"],
            ultimo_procesamiento=format_date(periodo_data.get("ultimo_procesamiento")),
            created_at=format_date(periodo_data.get("created_at"))
        )
        
        return PeriodoResponse(success=True, periodo=periodo_info)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error actualizando periodo: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al actualizar periodo: {str(e)}"
        )


@app.delete("/api/v1/periodos/{periodo_id}", tags=["Periodos"])
async def delete_periodo(
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Elimina un periodo y elimina los archivos físicos (PDFs) y metadatas asociados.
    
    Al eliminar un periodo, también se eliminan:
    - Los archivos PDF físicos en uploads/
    - Los archivos de metadata en uploads/metadata/
    - La entrada del periodo en el sistema
    
    Args:
        periodo_id: ID del periodo
        
    Returns:
        Confirmación de eliminación
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        periodo_manager = get_periodo_manager()
        upload_manager = get_upload_manager()
        
        # Eliminar periodo y eliminar archivos físicos asociados
        deleted = periodo_manager.delete_periodo(
            periodo_id, 
            upload_manager=upload_manager, 
            delete_files=True  # Eliminar archivos físicos y metadatas
        )
        
        if not deleted:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo {periodo_id} no encontrado"
            )
        
        return {"success": True, "message": f"Periodo {periodo_id} eliminado"}
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error eliminando periodo: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al eliminar periodo: {str(e)}"
        )


@app.get("/api/v1/periodos/{periodo_id}/resumen-ps", response_model=PeriodoResumenPSResponse, tags=["Periodos"])
async def get_periodo_resumen_ps(
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Obtiene el resumen PS (Off-Shore/On-Shore) de un periodo.
    
    Args:
        periodo_id: ID del periodo
        
    Returns:
        Resumen PS con Department, Discipline, Total US $, Total Horas, Ratios EDP
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        periodo_manager = get_periodo_manager()
        periodo_data = periodo_manager.get_periodo(periodo_id)
        
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo {periodo_id} no encontrado"
            )
        
        # Intentar cargar consolidado desde archivo
        from ..services.resumen_consolidator import ResumenConsolidator
        from ..api.dependencies import get_file_manager
        
        file_manager = get_file_manager()
        consolidator = ResumenConsolidator(output_folder=file_manager.get_output_folder() or Path("./output"))
        consolidado = consolidator.load_consolidado(periodo_id)
        
        if consolidado:
            # Convertir consolidado a formato de respuesta
            periodo_tipo = periodo_data.get("tipo", "offshore").lower()
            items_data = consolidado.get("resumen_ps", {}).get(periodo_tipo, [])
            
            # Convertir items a formato PeriodoResumenPSItem
            items = []
            for item in items_data:
                if periodo_tipo == "onshore":
                    # Para OnShore, incluir todos los campos
                    items.append({
                        "department": item.get("department", "---"),
                        "discipline": item.get("discipline", "---"),
                        "total_us": item.get("total_us", 0.0),
                        "total_horas": item.get("total_hours", 0.0),  # Mapear total_hours a total_horas
                        "ratios_edp": (item.get("total_us", 0.0) / item.get("total_hours", 1.0)) if item.get("total_hours", 0.0) > 0 else 0.0,
                        "job_no": item.get("job_no", "---"),
                        "wages": item.get("wages", 0.0),
                        "expatriate_allowances": item.get("expatriate_allowances", 0.0),
                        "multiplier": item.get("multiplier", 0.0),
                        "odc": item.get("odc", 0.0),
                        "epp": item.get("epp", 0.0),
                        "total_hours": item.get("total_hours", 0.0)
                    })
                else:
                    # Para OffShore
                    items.append({
                        "department": item.get("department", "---"),
                        "discipline": item.get("discipline", "---"),
                        "total_us": item.get("total_us", 0.0),
                        "total_horas": item.get("total_horas", 0.0),
                        "ratios_edp": item.get("ratios_edp", 0.0)
                    })
            
            return PeriodoResumenPSResponse(
                success=True,
                periodo_id=periodo_id,
                tipo=periodo_tipo,
                items=items,
                total=len(items)
            )
        else:
            # Si no hay consolidado, intentar consolidar ahora (puede que los JSONs aún existan)
            request_ids = periodo_manager.get_archivos_from_periodo(periodo_id)
            if request_ids:
                periodo_tipo = periodo_data.get("tipo", "offshore")
                consolidado = consolidator.consolidate_periodo(
                    periodo_id=periodo_id,
                    periodo_tipo=periodo_tipo,
                    request_ids=request_ids
                )
                
                # Convertir y retornar
                items_data = consolidado.get("resumen_ps", {}).get(periodo_tipo.lower(), [])
                items = []
                for item in items_data:
                    if periodo_tipo.lower() == "onshore":
                        items.append({
                            "department": item.get("department", "---"),
                            "discipline": item.get("discipline", "---"),
                            "total_us": item.get("total_us", 0.0),
                            "total_horas": item.get("total_hours", 0.0),
                            "ratios_edp": (item.get("total_us", 0.0) / item.get("total_hours", 1.0)) if item.get("total_hours", 0.0) > 0 else 0.0,
                            "job_no": item.get("job_no", "---"),
                            "wages": item.get("wages", 0.0),
                            "expatriate_allowances": item.get("expatriate_allowances", 0.0),
                            "multiplier": item.get("multiplier", 0.0),
                            "odc": item.get("odc", 0.0),
                            "epp": item.get("epp", 0.0),
                            "total_hours": item.get("total_hours", 0.0)
                        })
                    else:
                        items.append({
                            "department": item.get("department", "---"),
                            "discipline": item.get("discipline", "---"),
                            "total_us": item.get("total_us", 0.0),
                            "total_horas": item.get("total_horas", 0.0),
                            "ratios_edp": item.get("ratios_edp", 0.0)
                        })
                
                return PeriodoResumenPSResponse(
                    success=True,
                    periodo_id=periodo_id,
                    tipo=periodo_tipo.lower(),
                    items=items,
                    total=len(items)
                )
        
        # Si no hay archivos procesados, retornar vacío
        return PeriodoResumenPSResponse(
            success=True,
            periodo_id=periodo_id,
            tipo=periodo_data.get("tipo", "offshore"),
            items=[],
            total=0
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error obteniendo resumen PS: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al obtener resumen PS: {str(e)}"
        )


@app.post("/api/v1/periodos/{periodo_id}/exportar", tags=["Periodos"])
async def exportar_periodo(
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Exporta un periodo (genera ZIP consolidado con todos los ZIPs y Excels de los archivos procesados).
    
    Este endpoint crea un ZIP maestro que contiene:
    - Todos los ZIPs individuales de cada archivo procesado del periodo
    - Todos los Excels individuales de cada archivo procesado del periodo (si existen)
    
    Solo incluye archivos que ya han sido procesados (no incluye pendientes).
    
    Args:
        periodo_id: ID del periodo (ej: "2025-11-onshore")
        
    Returns:
        Redirección a /public/{periodo_id}_export_{timestamp}.zip
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    from fastapi.responses import RedirectResponse
    import zipfile
    
    periodo_manager = get_periodo_manager()
    upload_manager = get_upload_manager()
    processed_tracker = get_processed_tracker()
    archive_manager = get_archive_manager()
    
    # 1. Verificar que el periodo existe
    periodo_data = periodo_manager.get_periodo(periodo_id)
    if not periodo_data:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Periodo '{periodo_id}' no encontrado"
        )
    
    # 2. Obtener todos los request_ids de archivos procesados del periodo
    request_ids = periodo_manager.get_archivos_from_periodo(periodo_id)
    
    if not request_ids:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No hay archivos procesados en el periodo '{periodo_id}' para exportar"
        )
    
    # 3. Buscar zip_filename y excel_filename para cada request_id
    archivos_para_exportar = []
    
    # Buscar en archivos procesados desde upload-pdf
    uploaded_files = upload_manager.list_uploaded_files(processed=True)
    for f in uploaded_files:
        request_id = f.get("request_id")
        if request_id and request_id in request_ids:
            zip_filename = f.get("zip_filename")
            excel_filename = f.get("excel_filename")
            filename = f.get("filename", "unknown")
            
            if zip_filename or excel_filename:
                archivos_para_exportar.append({
                    "request_id": request_id,
                    "filename": filename,
                    "zip_filename": zip_filename,
                    "excel_filename": excel_filename
                })
    
    # Buscar también en archivos procesados directamente (por si acaso)
    direct_files = processed_tracker.get_processed_files()
    for f in direct_files:
        request_id = f.get("request_id")
        if request_id and request_id in request_ids:
            # Verificar que no esté ya en la lista
            if not any(a["request_id"] == request_id for a in archivos_para_exportar):
                zip_filename = f.get("zip_filename")
                excel_filename = f.get("excel_filename")
                filename = f.get("filename", "unknown")
                
                if zip_filename or excel_filename:
                    archivos_para_exportar.append({
                        "request_id": request_id,
                        "filename": filename,
                        "zip_filename": zip_filename,
                        "excel_filename": excel_filename
                    })
    
    if not archivos_para_exportar:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"No se encontraron archivos ZIP/Excel para exportar del periodo '{periodo_id}'"
        )
    
    # 4. Crear ZIP maestro con todos los archivos
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_maestro_filename = f"{periodo_id}_export_{timestamp}.zip"
    zip_maestro_path = archive_manager.public_folder / zip_maestro_filename
    
    archivos_agregados = 0
    archivos_no_encontrados = []
    
    try:
        with zipfile.ZipFile(zip_maestro_path, 'w', zipfile.ZIP_DEFLATED) as zip_maestro:
            for archivo_info in archivos_para_exportar:
                # Agregar ZIP si existe
                if archivo_info.get("zip_filename"):
                    zip_filename = archivo_info["zip_filename"]
                    zip_path = archive_manager.public_folder / zip_filename
                    
                    if zip_path.exists():
                        # Agregar al ZIP maestro (en la raíz, sin subcarpetas)
                        zip_maestro.write(zip_path, zip_filename)
                        archivos_agregados += 1
                        logger.info(f"[Export {periodo_id}] Agregado ZIP: {zip_filename}")
                    else:
                        archivos_no_encontrados.append(f"ZIP: {zip_filename}")
                
                # Agregar Excel si existe
                if archivo_info.get("excel_filename"):
                    excel_filename = archivo_info["excel_filename"]
                    excel_path = archive_manager.public_folder / excel_filename
                    
                    if excel_path.exists():
                        # Agregar al ZIP maestro (en la raíz, sin subcarpetas)
                        zip_maestro.write(excel_path, excel_filename)
                        archivos_agregados += 1
                        logger.info(f"[Export {periodo_id}] Agregado Excel: {excel_filename}")
                    else:
                        archivos_no_encontrados.append(f"Excel: {excel_filename}")
        
        if archivos_agregados == 0:
            # Si no se agregó ningún archivo, eliminar el ZIP vacío
            zip_maestro_path.unlink()
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No se encontraron archivos físicos para exportar del periodo '{periodo_id}'. Archivos no encontrados: {', '.join(archivos_no_encontrados)}"
            )
        
        logger.info(f"[Export {periodo_id}] ZIP maestro creado: {zip_maestro_filename} ({archivos_agregados} archivos)")
        
        if archivos_no_encontrados:
            logger.warning(f"[Export {periodo_id}] Algunos archivos no se encontraron: {', '.join(archivos_no_encontrados)}")
        
        # 5. Redirigir a la descarga del ZIP maestro
        return RedirectResponse(url=f"/public/{zip_maestro_filename}", status_code=302)
    
    except Exception as e:
        logger.exception(f"Error exportando periodo {periodo_id}: {e}")
        # Limpiar ZIP maestro si se creó parcialmente
        if zip_maestro_path.exists():
            try:
                zip_maestro_path.unlink()
            except Exception:
                pass
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error creando exportación del periodo: {str(e)}"
        )


@app.get("/api/v1/periodos/{periodo_id}/resumen-ps/exportar-excel", tags=["Periodos"])
async def exportar_resumen_ps_excel(
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Exporta el resumen PS de un periodo a Excel.
    
    Args:
        periodo_id: ID del periodo
        
    Returns:
        Archivo Excel con el resumen PS
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        from fastapi.responses import FileResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        periodo_manager = get_periodo_manager()
        periodo_data = periodo_manager.get_periodo(periodo_id)
        
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo {periodo_id} no encontrado"
            )
        
        # Cargar consolidado
        from ..services.resumen_consolidator import ResumenConsolidator
        from ..api.dependencies import get_file_manager
        
        file_manager = get_file_manager()
        consolidator = ResumenConsolidator(output_folder=file_manager.get_output_folder() or Path("./output"))
        consolidado = consolidator.load_consolidado(periodo_id)
        
        if not consolidado:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No hay resumen consolidado para el periodo {periodo_id}"
            )
        
        periodo_tipo = periodo_data.get("tipo", "offshore").lower()
        items_data = consolidado.get("resumen_ps", {}).get(periodo_tipo, [])
        
        if not items_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"No hay datos en el resumen PS del periodo {periodo_id}"
            )
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Resumen PS-{periodo_tipo.capitalize()}"
        
        # Estilos
        header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")  # Header oscuro
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        currency_format = '#,##0.00'
        number_format = '#,##0.00'
        
        # Escribir headers según tipo
        row = 1
        if periodo_tipo == "onshore":
            headers = ["Job No", "Department", "Discipline", "Wages", "Expatriate Allowances", 
                      "Multiplier", "ODC", "EPP", "Total US $", "Total Hours"]
            col_widths = [15, 20, 20, 15, 20, 12, 12, 12, 15, 15]
        else:
            headers = ["Department", "Discipline", "Total US $", "Total Horas", "Ratios EDP $/HH"]
            col_widths = [25, 25, 15, 15, 18]
        
        # Escribir headers
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Escribir datos
        for item in items_data:
            row += 1
            if periodo_tipo == "onshore":
                values = [
                    item.get("job_no", "---"),
                    item.get("department", "---"),
                    item.get("discipline", "---"),
                    item.get("wages", 0.0),
                    item.get("expatriate_allowances", 0.0),
                    item.get("multiplier", 0.0),
                    item.get("odc", 0.0),
                    item.get("epp", 0.0),
                    item.get("total_us", 0.0),
                    item.get("total_hours", 0.0)
                ]
            else:
                values = [
                    item.get("department", "---"),
                    item.get("discipline", "---"),
                    item.get("total_us", 0.0),
                    item.get("total_horas", 0.0),
                    item.get("ratios_edp", 0.0)
                ]
            
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = border
                
                # Aplicar formato según tipo de dato
                if isinstance(value, (int, float)):
                    if periodo_tipo == "onshore":
                        if col_idx in [4, 5, 9]:  # Columnas monetarias (Wages, Expatriate Allowances, Total US $)
                            cell.number_format = currency_format
                        elif col_idx == 10:  # Total Hours
                            cell.number_format = '#,##0.00'  # Formato con 2 decimales
                        else:  # Multiplier, ODC, EPP
                            cell.number_format = number_format
                    else:  # OffShore
                        if col_idx == 3:  # Total US $
                            cell.number_format = currency_format
                        elif col_idx == 4:  # Total Horas
                            cell.number_format = '#,##0.00'  # Formato con 2 decimales
                        elif col_idx == 5:  # Ratios EDP $/HH
                            cell.number_format = '#,##0.000'  # Formato con 3 decimales como en la imagen
                        else:
                            cell.number_format = number_format
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
        
        # Guardar Excel
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"resumen_ps_{periodo_id}_{timestamp}.xlsx"
        excel_path = file_manager.get_output_folder() or Path("./output")
        excel_path = excel_path / "public" / excel_filename
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb.save(excel_path)
        
        return FileResponse(
            path=excel_path,
            filename=excel_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error exportando resumen PS a Excel: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error exportando resumen PS a Excel: {str(e)}"
        )


@app.post("/api/v1/periodos/{periodo_id}/bloquear", response_model=PeriodoResponse, tags=["Periodos"])
async def bloquear_periodo(
    periodo_id: str,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Bloquea un periodo (cambia estado a "cerrado").
    Al bloquear, guarda un snapshot de los apartados actuales para mantener histórico.
    
    Args:
        periodo_id: ID del periodo
        
    Returns:
        Periodo actualizado con estado "cerrado"
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        periodo_manager = get_periodo_manager()
        
        # Cargar apartados actuales antes de bloquear para guardar snapshot
        maestros_data = load_maestros_data()
        apartados_actuales = maestros_data.get("apartados", [])
        
        # Actualizar estado a "cerrado"
        periodo_data = periodo_manager.update_periodo(periodo_id, {"estado": "cerrado"})
        
        if not periodo_data:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Periodo {periodo_id} no encontrado"
            )
        
        # Guardar snapshot de apartados para este periodo
        if apartados_actuales:
            periodo_manager.save_apartados_snapshot(periodo_id, apartados_actuales)
            logger.info(f"Snapshot de apartados guardado para periodo cerrado {periodo_id}")
        
        # Asegurar que el estado sea "cerrado" en la respuesta
        periodo_data["estado"] = "cerrado"
        
        # Función helper para formatear fechas
        def format_date(date_str: Optional[str]) -> Optional[str]:
            """Formatea fecha ISO a formato DD/MM/YYYY, HH:MM"""
            if not date_str:
                return None
            try:
                dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                return dt.strftime("%d/%m/%Y, %H:%M")
            except Exception:
                return date_str  # Si falla, retornar original
        
        # Construir respuesta con el periodo actualizado
        periodo_info = PeriodoInfo(
            periodo_id=periodo_data["periodo_id"],
            periodo=periodo_data["periodo"],
            tipo=periodo_data["tipo"],
            estado="cerrado",  # Asegurar que el estado sea "cerrado"
            registros=periodo_data.get("registros", 0),
            ultimo_procesamiento=format_date(periodo_data.get("ultimo_procesamiento")),
            created_at=format_date(periodo_data.get("created_at"))
        )
        
        return PeriodoResponse(success=True, periodo=periodo_info)
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error bloqueando periodo: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Error al bloquear periodo: {str(e)}"
    )


@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """
    Handler global para excepciones no capturadas.
    """
    logger.exception(f"Excepción no capturada: {exc}")
    return JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={
            "success": False,
            "error": "Error interno del servidor",
            "details": str(exc) if os.getenv("DEBUG", "false").lower() == "true" else None
        }
    )


# ===== Maestros Endpoints =====

@app.get("/api/v1/maestros/apartados", response_model=MaestrosResponse, tags=["Maestros"])
async def get_maestros_apartados(
    periodo_id: Optional[str] = Query(None, description="ID del periodo (opcional). Si se proporciona y el periodo está cerrado, retorna el snapshot"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Obtiene la lista de apartados (conceptos) configurados.
    
    Si se proporciona periodo_id y el periodo está cerrado, retorna el snapshot
    de apartados que tenía cuando se cerró, para mantener el histórico.
    
    Args:
        periodo_id: ID del periodo (opcional). Si el periodo está cerrado, retorna su snapshot.
    
    Returns:
        Lista de apartados guardados en maestros_apartados.json o snapshot del periodo si está cerrado
    """
    get_current_user_email(credentials)
    
    try:
        apartados = []
        
        # Si hay periodo_id, verificar si tiene snapshot (periodo cerrado)
        if periodo_id:
            periodo_manager = get_periodo_manager()
            periodo = periodo_manager.get_periodo(periodo_id)
            
            if periodo and periodo.get("estado") == "cerrado":
                # Periodo cerrado: usar snapshot
                snapshot = periodo_manager.get_apartados_snapshot(periodo_id)
                if snapshot:
                    apartados = snapshot
                    logger.info(f"Retornando snapshot de apartados para periodo cerrado {periodo_id}")
                else:
                    # No hay snapshot, usar apartados actuales
                    data = load_maestros_data()
                    apartados = data.get("apartados", [])
            else:
                # Periodo abierto o no existe: usar apartados actuales
                data = load_maestros_data()
                apartados = data.get("apartados", [])
        else:
            # Sin periodo_id: usar apartados actuales
            data = load_maestros_data()
            apartados = data.get("apartados", [])
        
        # Convertir a formato ApartadoInfo
        apartados_info = [
            ApartadoInfo(
                id=a.get("id", ""),
                nombre=a.get("nombre", ""),
                orden=a.get("orden", 0)
            )
            for a in apartados
        ]
        
        return MaestrosResponse(
            success=True,
            apartados=apartados_info
        )
    except Exception as e:
        logger.error(f"Error al obtener apartados: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=ErrorResponse(
                request_id="",
                error="Error al obtener apartados",
                details={"message": str(e)}
            ).model_dump()
        )


@app.post("/api/v1/maestros/apartados", response_model=MaestrosResponse, tags=["Maestros"])
async def save_maestros_apartados(
    request: MaestrosSaveRequest,
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Guarda la lista de apartados (conceptos) configurados.
    
    Args:
        request: Lista de apartados a guardar
        
    Returns:
        Confirmación de guardado exitoso con la lista de apartados guardados
    """
    get_current_user_email(credentials)
    
    try:
        # Convertir ApartadoInfo a dict para guardar en JSON
        apartados_dict = [
            {
                "id": a.id,
                "nombre": a.nombre,
                "orden": a.orden
            }
            for a in request.apartados
        ]
        
        # Guardar en archivo JSON
        data = {"apartados": apartados_dict}
        if save_maestros_data(data):
            return MaestrosResponse(
                success=True,
                apartados=request.apartados,
                message="Apartados guardados exitosamente"
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=ErrorResponse(
                    request_id="",
                    error="Error al guardar apartados",
                    details={"message": "No se pudo escribir en el archivo"}
                ).model_dump()
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error al guardar apartados: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=ErrorResponse(
                request_id="",
                error="Error al guardar apartados",
                details={"message": str(e)}
            ).model_dump()
        )


@app.get("/api/v1/maestros/apartados/exportar-excel", tags=["Maestros"])
async def exportar_apartados_excel(
    periodo_id: Optional[str] = Query(None, description="ID del periodo (opcional)"),
    credentials: Optional[HTTPAuthorizationCredentials] = Depends(security)
):
    """
    Exporta la lista de apartados configurados a Excel.
    
    Args:
        periodo_id: ID del periodo (opcional) para incluir en el nombre del archivo
    
    Returns:
        Archivo Excel con los apartados ordenados según su posición
    """
    # Verificar autenticación y estado del usuario
    get_current_user_email(credentials)
    
    try:
        from fastapi.responses import FileResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from ..api.dependencies import get_file_manager
        
        # Cargar apartados - usar snapshot si el periodo está cerrado, sino usar apartados actuales
        apartados = []
        if periodo_id:
            # Verificar si hay snapshot para este periodo
            periodo_manager = get_periodo_manager()
            snapshot = periodo_manager.get_apartados_snapshot(periodo_id)
            
            if snapshot:
                # Usar snapshot del periodo cerrado
                apartados = snapshot
                logger.info(f"Usando snapshot de apartados para periodo {periodo_id}")
            else:
                # Usar apartados actuales
                data = load_maestros_data()
                apartados = data.get("apartados", [])
        else:
            # Si no hay periodo_id, usar apartados actuales
            data = load_maestros_data()
            apartados = data.get("apartados", [])
        
        if not apartados:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="No hay apartados configurados para exportar"
            )
        
        # Ordenar apartados por orden
        apartados_ordenados = sorted(apartados, key=lambda x: x.get("orden", 0))
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Apartados Configurados"
        
        # Estilos
        header_fill = PatternFill(start_color="1F1F1F", end_color="1F1F1F", fill_type="solid")  # Header oscuro
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Escribir headers
        headers = ["Orden", "Concepto", "Total US $"]
        col_widths = [15, 40, 18]
        
        row = 1
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        
        # Escribir datos
        for apartado in apartados_ordenados:
            row += 1
            values = [
                apartado.get("orden", 0) + 1,  # Mostrar posición desde 1
                apartado.get("nombre", "---"),
                "---"  # Total US $ - placeholder
            ]
            
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col_idx, value=value)
                cell.border = border
                
                # Formato según tipo de columna
                if col_idx == 1:  # Orden - número centrado
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 2:  # Concepto - texto izquierda
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:  # Total US $ - texto centrado
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Guardar Excel
        file_manager = get_file_manager()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if periodo_id:
            excel_filename = f"noreembolsables_{periodo_id}_{timestamp}.xlsx"
        else:
            excel_filename = f"noreembolsables_{timestamp}.xlsx"
        base_output = file_manager.get_output_folder() or "./output"
        excel_path = Path(base_output) / "public" / excel_filename
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        wb.save(excel_path)
        
        return FileResponse(
            path=excel_path,
            filename=excel_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.exception(f"Error exportando apartados a Excel: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=ErrorResponse(
                request_id="",
                error="Error al exportar apartados a Excel",
                details={"message": str(e)}
            ).model_dump()
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)

