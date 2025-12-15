"""
Excel Generator - Generación de archivos Excel consolidados
Responsabilidad: Crear archivos Excel a partir de JSONs estructurados
"""

import json
import logging
from pathlib import Path
from typing import Optional

from ..core.file_manager import truncate_filename_for_path

logger = logging.getLogger(__name__)


def truncate_request_id_for_folder(request_id: str, max_length: int = 30) -> str:
    """
    Trunca un request_id para usarlo como nombre de carpeta.
    
    Args:
        request_id: ID del request (puede ser UUID largo)
        max_length: Longitud máxima permitida (default: 30)
        
    Returns:
        Request ID truncado
    """
    if len(request_id) <= max_length:
        return request_id
    return request_id[:max_length]


def extract_code_from_filename(filename: str) -> str:
    """
    Extrae el código del nombre del archivo (antes del primer '_').
    
    Ejemplos:
    - "000006_EDP Octubre_26443_AYALA SEHLKE, ANA MARIA_063573 IFBQOT.pdf" -> "000006"
    - "00037_EDP Octubre_TRANSLUC SpA_1000026741 757.pdf" -> "00037"
    
    Args:
        filename: Nombre del archivo
        
    Returns:
        Código extraído o cadena vacía si no se encuentra
    """
    if not filename:
        return ""
    
    first_underscore_index = filename.find('_')
    if first_underscore_index == -1:
        # Si no hay '_', devolver el nombre completo sin extensión
        return filename.rsplit('.', 1)[0] if '.' in filename else filename
    
    return filename[:first_underscore_index]


async def generate_excel_for_request(
    request_id: str,
    pdf_name: str,
    timestamp: str,
    archive_manager,
    file_manager
) -> tuple[Optional[str], Optional[str]]:
    """
    Genera un archivo Excel consolidado para un request_id.
    SIEMPRE genera el Excel, incluso si no hay datos (solo con encabezados).
    
    Args:
        request_id: ID del procesamiento
        pdf_name: Nombre del PDF
        timestamp: Timestamp para el nombre del archivo
        archive_manager: Instancia de ArchiveManager
        file_manager: Instancia de FileManager
        
    Returns:
        Tupla (excel_filename, excel_download_url) o (None, None) solo si hay error crítico
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
        from openpyxl.utils import get_column_letter
        
        # Buscar todos los JSONs estructurados en la carpeta específica por request_id
        base_output = file_manager.get_output_folder() or "./output"
        # Carpeta específica por request_id: output/api/{request_id}/structured/ (truncado a 30 chars)
        request_id_folder = truncate_request_id_for_folder(request_id)
        structured_folder = Path(base_output) / "api" / request_id_folder / "structured"
        
        # Crear workbook de Excel (siempre se crea)
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos Consolidados"
        
        # Estilos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Diccionario para almacenar todas las tablas consolidadas
        all_tables = {}
        all_columns = set()
        
        # Lista de campos estándar conocidos (por si no hay JSONs o están vacíos)
        standard_fields = {
            "hoja",  # Siempre presente
            # Columnas extras del periodo (siempre presentes)
            "Periodo", "Año", "Scope", "Codigo", "Pagina", "tDivisaOriginal",
            # Campos comunes de mcomprobante
            "tNumero", "nPrecioTotal", "tFecha", "tRazonSocial",
            # Campos comunes de mcomprobante_detalle
            "tDescripcion", "nCantidad", "nPrecioUnitario", "nImporte",
            # Campos comunes de mresumen
            "tConcepto", "nMonto", "tMoneda",
            # Campos comunes de mproveedor
            "tRazonSocial", "tRUC", "tDireccion",
            # Campos comunes de mjornada
            "tEmpleado", "nHoras", "tFechaTrabajo"
        }
        
        # Buscar JSONs en la carpeta específica por request_id
        json_files = []
        if structured_folder.exists():
            # Buscar todos los JSONs estructurados en esta carpeta específica
            all_json_files = sorted(structured_folder.glob("*_structured.json"))
            json_files = all_json_files  # Todos los JSONs en esta carpeta pertenecen a este request_id
        
        # Procesar cada JSON estructurado (si existen)
        for json_file in json_files:
            try:
                # Extraer número de página
                page_match = json_file.stem.split("_page_")
                if len(page_match) >= 2:
                    page_num_str = page_match[1].replace("_structured", "")
                    try:
                        page_num = int(page_num_str)
                    except ValueError:
                        page_num = 0
                else:
                    page_num = 0
                
                # Leer JSON
                with open(json_file, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
                
                # Obtener metadata para extraer información del periodo
                metadata = json_data.get("metadata", {})
                month = metadata.get("month", "")
                year = metadata.get("year", "")
                onshore_offshore = metadata.get("onshore_offshore", "")
                
                # Capitalizar Scope (OnShore/OffShore)
                scope_value = ""
                if onshore_offshore:
                    scope_value = onshore_offshore.capitalize()
                    if scope_value.lower() == "onshore":
                        scope_value = "OnShore"
                    elif scope_value.lower() == "offshore":
                        scope_value = "OffShore"
                
                # Obtener tablas del nivel raíz (ya no están en additional_data)
                # Lista de tablas principales que queremos procesar
                main_tables = [
                    "mresumen", "mcomprobante", "mcomprobante_detalle",
                    "mjornada", "mjornada_empleado", "mproveedor", "mmaquinaria_equipos"
                ]
                
                # Obtener mcomprobante para propagar tMonedaOriginal y nPrecioTotalOriginal a los detalles
                mcomprobante_list = json_data.get("mcomprobante", [])
                comprobante_currency_map = {}
                if mcomprobante_list and isinstance(mcomprobante_list, list) and len(mcomprobante_list) > 0:
                    comprobante = mcomprobante_list[0]  # Tomar el primer comprobante
                    if isinstance(comprobante, dict):
                        comprobante_currency_map = {
                            "tMonedaOriginal": comprobante.get("tMonedaOriginal", ""),
                            "nPrecioTotalOriginal": comprobante.get("nPrecioTotalOriginal", "")
                        }
                
                # Procesar cada tabla del nivel raíz
                for table_name in main_tables:
                    table_data = json_data.get(table_name, [])
                    if not isinstance(table_data, list):
                        continue
                    
                    if table_name not in all_tables:
                        all_tables[table_name] = []
                    
                    for record in table_data:
                        if isinstance(record, dict):
                            record_with_hoja = {}
                            
                            # Si es mcomprobante_detalle y tiene catalogos, aplanarlos
                            if table_name == "mcomprobante_detalle" and "catalogos" in record:
                                catalogos = record.get("catalogos", {})
                                # Extraer valores de los catálogos y agregarlos como columnas
                                if "mdivisa" in catalogos and catalogos["mdivisa"]:
                                    record_with_hoja["tDivisa"] = catalogos["mdivisa"][0].get("tDivisa", "") if isinstance(catalogos["mdivisa"][0], dict) else ""
                                if "mproveedor" in catalogos and catalogos["mproveedor"]:
                                    record_with_hoja["tProveedor"] = catalogos["mproveedor"][0].get("tRazonSocial", "") if isinstance(catalogos["mproveedor"][0], dict) else ""
                                if "mnaturaleza" in catalogos and catalogos["mnaturaleza"]:
                                    record_with_hoja["tNaturaleza"] = catalogos["mnaturaleza"][0].get("tNaturaleza", "") if isinstance(catalogos["mnaturaleza"][0], dict) else ""
                                if "mdocumento_tipo" in catalogos and catalogos["mdocumento_tipo"]:
                                    doc_tipo = catalogos["mdocumento_tipo"][0]
                                    if isinstance(doc_tipo, dict):
                                        record_with_hoja["iMDocumentoTipo"] = doc_tipo.get("iMDocumentoTipo", "")
                                        record_with_hoja["tTipoDocumento"] = doc_tipo.get("tTipo", "")
                                if "midioma" in catalogos and catalogos["midioma"]:
                                    idioma = catalogos["midioma"][0]
                                    if isinstance(idioma, dict):
                                        record_with_hoja["iMIdioma"] = idioma.get("iMIdioma", "")
                                        record_with_hoja["tIdioma"] = idioma.get("tIdioma", "")
                                # Remover catalogos del record para no duplicarlo
                                record_copy = {k: v for k, v in record.items() if k != "catalogos"}
                            else:
                                record_copy = record
                            
                            for key, value in record_copy.items():
                                if isinstance(value, (list, tuple)):
                                    record_with_hoja[key] = ", ".join(str(v) for v in value) if value else ""
                                elif isinstance(value, dict):
                                    record_with_hoja[key] = json.dumps(value, ensure_ascii=False)
                                else:
                                    record_with_hoja[key] = value
                            
                            # Agregar campo hoja
                            record_with_hoja["hoja"] = f"{pdf_name} - Página {page_num}"
                            
                            # Extraer código del nombre del PDF (antes del primer '_')
                            codigo_value = extract_code_from_filename(pdf_name)
                            
                            # Extraer moneda original (tMonedaOriginal)
                            # Prioridad: del record mismo, luego del mcomprobante padre, luego vacío
                            moneda_original = record.get("tMonedaOriginal") or comprobante_currency_map.get("tMonedaOriginal", "")
                            
                            # Agregar columnas extras del periodo
                            record_with_hoja["Periodo"] = month if month else ""
                            record_with_hoja["Año"] = year if year else ""
                            record_with_hoja["Scope"] = scope_value
                            record_with_hoja["Codigo"] = codigo_value
                            record_with_hoja["Pagina"] = page_num
                            record_with_hoja["tDivisaOriginal"] = moneda_original if moneda_original else ""
                            # NOTA: nPrecioTotalOriginal ya se mapea automáticamente desde el JSON, no necesita columna separada
                            
                            all_tables[table_name].append(record_with_hoja)
                            
                            # Agregar todas las columnas encontradas
                            all_columns.update(record_with_hoja.keys())
            except Exception:
                continue
        
        # Si no hay columnas de datos, usar campos estándar conocidos
        if not all_columns:
            all_columns = standard_fields
        
        # Asegurar que las columnas extras del periodo siempre estén presentes
        # NOTA: nPrecioTotalOriginal ya se mapea automáticamente desde el JSON, no necesita columna separada
        extra_columns = ["Periodo", "Año", "Scope", "Codigo", "Pagina", "tDivisaOriginal"]
        all_columns.update(extra_columns)
        
        # Consolidar todas las tablas en una sola lista
        all_records = []
        for table_name, records in all_tables.items():
            all_records.extend(records)
        
        # Ordenar columnas (hoja y columnas extras del periodo primero)
        # Filtrar las columnas extra que realmente existen en all_columns
        existing_extra_columns = [c for c in extra_columns if c in all_columns]
        columns_to_sort = [c for c in all_columns if c not in ["hoja"] + extra_columns]
        sorted_columns = sorted(columns_to_sort)
        # Orden específico: hoja, periodo info, tDivisaOriginal, luego resto
        period_info_cols = [c for c in ["Periodo", "Año", "Scope", "Codigo", "Pagina"] if c in existing_extra_columns]
        currency_cols = [c for c in ["tDivisaOriginal"] if c in existing_extra_columns]
        column_order = ["hoja"] + period_info_cols + currency_cols + sorted_columns
        
        # Escribir encabezados (siempre se escriben, aunque no haya datos)
        row = 1
        for col_idx, col_name in enumerate(column_order, start=1):
            cell = ws.cell(row=row, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Columnas que deben tener formato numérico
        numeric_columns = ["nPrecioTotal", "nPrecioUnitario"]
        
        # Escribir datos (si hay)
        row = 2
        for record in all_records:
            if isinstance(record, dict):
                for col_idx, col_name in enumerate(column_order, start=1):
                    value = record.get(col_name, "")
                    if value is None:
                        value = ""
                    
                    cell = ws.cell(row=row, column=col_idx, value=value)
                    
                    # Aplicar formato numérico a columnas específicas
                    if col_name in numeric_columns:
                        # Intentar convertir a número si es posible
                        if value != "":
                            try:
                                # Convertir a float si es string numérico
                                if isinstance(value, str):
                                    # Limpiar el string (remover espacios, comas, etc.)
                                    cleaned_value = value.replace(",", "").replace(" ", "").strip()
                                    if cleaned_value:
                                        numeric_value = float(cleaned_value)
                                        cell.value = numeric_value
                                elif isinstance(value, (int, float)):
                                    cell.value = float(value)
                                
                                # Aplicar formato numérico con 2 decimales
                                cell.number_format = '#,##0.00'
                            except (ValueError, TypeError):
                                # Si no se puede convertir, dejar como está
                                pass
                row += 1
        
        # Ajustar ancho de columnas
        for col_idx, col_name in enumerate(column_order, start=1):
            max_length = len(str(col_name))
            if row > 2:  # Solo si hay datos
                for row_idx in range(2, row):
                    cell_value = ws.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max(max_length + 2, 10), 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
        
        # Guardar Excel en carpeta pública (SIEMPRE se guarda, aunque esté vacío)
        # Truncar nombre del Excel a máximo 50 caracteres para evitar rutas largas
        excel_filename_base = f"{pdf_name}_consolidado_{timestamp}_{request_id[:8]}.xlsx"
        excel_filename = truncate_filename_for_path(excel_filename_base, max_length=50)
        excel_path = archive_manager.public_folder / excel_filename
        
        # Asegurar que la carpeta pública existe
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        
        # Guardar el Excel
        wb.save(excel_path)
        
        # Verificar que se guardó correctamente
        if not excel_path.exists():
            logger.error(f"Excel no se guardó correctamente: {excel_path}")
            return None, None
        
        # Generar URL pública
        excel_download_url = archive_manager.get_public_url(excel_path)
        
        logger.info(f"[{request_id}] Excel generado exitosamente: {excel_filename} ({len(all_records)} registros, {len(column_order)} columnas)")
        
        return excel_filename, excel_download_url
    except Exception as e:
        logger.error(f"Error generando Excel para request_id {request_id}: {e}")
        import traceback
        logger.debug(f"Traceback Excel: {traceback.format_exc()}")
        return None, None

